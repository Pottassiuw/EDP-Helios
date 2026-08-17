"""Testes do módulo COFFEE (backend)."""
import os
import tempfile
import time as _time
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

# Blindagem global: impede que a execução de testes afete o banco de dados real
_tmp_test_dir = tempfile.mkdtemp(prefix="edp_coffee_test_")
os.environ.setdefault("COFFEE_DATA_DIR", _tmp_test_dir)
os.environ.setdefault("INPUT_DATA_DIR", _tmp_test_dir)

import pytest

import httpx

from coffee_module import classify, config


def test_classificacao_pendente():
    assert classify.classificar(config.SAP_PENDENTE, None) == "pendente"
    assert classify.classificar(config.SAP_PENDENTE, 17247854) == "pendente"


def test_classificacao_corrigida_na_transicao():
    # anterior era placeholder, atual virou SAP real
    assert classify.classificar(17247854, config.SAP_PENDENTE) == "corrigida"


def test_classificacao_gerada():
    # primeira busca já com SAP real (sem anterior conhecido)
    assert classify.classificar(17247854, None) == "gerada"
    # transição já "consumida": anterior também é real
    assert classify.classificar(17247854, 17247854) == "gerada"


def test_classificacao_duplicada():
    assert classify.classificar(config.SAP_DUPLICATA, None) == "duplicada"
    assert classify.classificar(config.SAP_DUPLICATA, config.SAP_PENDENTE) == "duplicada"
    # sentinel de duplicata nunca deve cair no ramo de "SAP real" (corrigida/gerada)
    assert classify.classificar(config.SAP_DUPLICATA, 17247854) == "duplicada"


@pytest.fixture
def coffee_tmp(monkeypatch, tmp_path):
    """Aponta o módulo para dados temporários, chave fake, e inicializa o banco."""
    monkeypatch.setenv("COFFEE_DATA_DIR", str(tmp_path))
    monkeypatch.setattr(config, "COFFEE_API_KEY", "fake-key")
    monkeypatch.setattr(config, "DELAY_BUSCA", 0)
    monkeypatch.setattr(config, "DELAY_GERACAO", 0)
    from coffee_module import db
    db.inicializar_banco()
    return tmp_path


def test_conexao_configura_timeout_e_foreign_keys_por_conexao(coffee_tmp):
    """Remover a configuração por conexão faria cada conexão voltar aos defaults."""
    from coffee_module import db

    primeira = db.get_db_connection()
    segunda = db.get_db_connection()
    try:
        assert primeira.execute("PRAGMA busy_timeout").fetchone()[0] == 5000
        assert segunda.execute("PRAGMA busy_timeout").fetchone()[0] == 5000
        assert primeira.execute("PRAGMA foreign_keys").fetchone()[0] == 1
        assert segunda.execute("PRAGMA foreign_keys").fetchone()[0] == 1
    finally:
        primeira.close()
        segunda.close()


def test_conexao_concorrente_abre_enquanto_outra_mantem_lock(coffee_tmp):
    """Renegociar WAL na abertura tenta lock exclusivo e quebra jobs concorrentes."""
    from coffee_module import db

    bloqueadora = db.get_db_connection()
    bloqueadora.execute("BEGIN EXCLUSIVE")
    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(db.get_db_connection)
            concorrente = future.result(timeout=1)
        try:
            assert concorrente.execute("PRAGMA journal_mode").fetchone()[0] == "wal"
        finally:
            concorrente.close()
    finally:
        bloqueadora.rollback()
        bloqueadora.close()


def test_upsert_primeira_busca_pendente(coffee_tmp):
    from coffee_module import db
    classe = db.upsert_nota(355617, 10000000, {"id_sap": 10000000})
    assert classe == "pendente"
    notas = db.listar_notas("pendente")
    assert len(notas) == 1
    assert notas[0]["pk"] == 355617
    assert notas[0]["id_sap_anterior"] is None
    assert notas[0]["arquivado"] is False


def test_upsert_transicao_corrigida_depois_gerada(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(355617, 10000000, {"id_sap": 10000000})
    # SAP atribuído: 10000000 -> real
    classe = db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    assert classe == "corrigida"
    nota = db.listar_notas("corrigida")[0]
    assert nota["id_sap_anterior"] == 10000000
    assert nota["arquivado"] is False
    # re-busca: transição consumida -> gerada
    classe = db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    assert classe == "gerada"
    assert db.listar_notas("corrigida") == []
    assert len(db.listar_notas("gerada")) == 1


def test_registrar_erro_e_listar_tudo(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(1, 10000000, {})
    db.registrar_erro(2, "timeout")
    todas = db.listar_notas()
    assert len(todas) == 2
    erro = [n for n in todas if n["pk"] == 2][0]
    assert erro["erro"] == "timeout"


# ---------------------------------------------------------------------------
# Sub-projeto 1 — Sistema de logs (coffee_logs)
# ---------------------------------------------------------------------------


def test_registrar_e_listar_log_roundtrip(coffee_tmp):
    from coffee_module import db
    db.registrar_log("api_call", "buscar_nota", 355617,
                     {"id": 355617, "status_http": 200, "tempo_ms": 12}, True)
    logs = db.listar_logs()
    assert len(logs) == 1
    log = logs[0]
    assert log["tipo"] == "api_call"
    assert log["acao"] == "buscar_nota"
    assert log["nota_pk"] == 355617
    assert log["sucesso"] is True
    assert log["detalhes"]["status_http"] == 200
    assert isinstance(log["id"], int)


def test_listar_logs_filtra_por_nota_e_tipo(coffee_tmp):
    from coffee_module import db
    db.registrar_log("api_call", "buscar_nota", 1, {"id": 1}, True)
    db.registrar_log("api_call", "arquivar", 2, {"id": 2}, True)
    db.registrar_log("transicao", "classificar", 1, {"anterior": "pendente"}, True)
    assert len(db.listar_logs(nota_pk=1)) == 2
    assert len(db.listar_logs(tipo="api_call")) == 2
    assert len(db.listar_logs(nota_pk=1, tipo="transicao")) == 1


def test_listar_logs_ordena_desc_e_respeita_limit(coffee_tmp):
    from coffee_module import db
    for i in range(5):
        db.registrar_log("acao_usuario", "regerar", i, {"i": i}, True)
    logs = db.listar_logs(limit=3)
    assert len(logs) == 3
    # mais recentes primeiro: o último inserido (i=4) deve vir antes
    assert logs[0]["detalhes"]["i"] >= logs[-1]["detalhes"]["i"]


def test_registrar_log_nunca_levanta(coffee_tmp):
    from coffee_module import db
    # detalhes não-serializável não deve quebrar o chamador
    db.registrar_log("api_call", "x", None, {"obj": object()}, False)
    # erro de sucesso=False registrado normalmente continua funcionando
    db.registrar_log("api_call", "y", None, None, False)
    assert any(l["acao"] == "y" for l in db.listar_logs())


# ---------------------------------------------------------------------------
# Task 2 — Transition logs in upsert_nota
# ---------------------------------------------------------------------------


def test_upsert_registra_transicao_de_classificacao(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(355617, 10000000, {"id_sap": 10000000})  # pendente (sem anterior)
    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})   # -> corrigida
    trans = db.listar_logs(tipo="transicao")
    classif = [t for t in trans if t["acao"] == "classificar"]
    assert len(classif) == 1
    assert classif[0]["nota_pk"] == 355617
    assert classif[0]["detalhes"]["anterior"] == "pendente"
    assert classif[0]["detalhes"]["novo"] == "corrigida"



def test_upsert_primeira_busca_nao_gera_transicao(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(355617, 10000000, {"id_sap": 10000000})
    assert db.listar_logs(tipo="transicao") == []


# ---------------------------------------------------------------------------
# Task 3 — client.py (httpx wrapper)
# ---------------------------------------------------------------------------


class _FakeResp:
    def __init__(self, payload=None, status=200):
        self._payload = payload
        self.status_code = status

    def raise_for_status(self):
        if self.status_code != 200:
            raise httpx.HTTPStatusError(
                "erro",
                request=httpx.Request("GET", "test"),
                response=httpx.Response(self.status_code),
            )

    def json(self):
        return self._payload


# json_all retorna uma STRING JSON (duplamente codificada)
_JSON_ALL = (
    '[{"model": "AppDeOlhoNaRede2.informativo", "pk": 355617, '
    '"fields": {"id_sap": 17247854, "arquivado": true, "sintoma": "EEST"}}]'
)


def test_buscar_nota_faz_duplo_parse(coffee_tmp, monkeypatch):
    monkeypatch.setattr(config, "COFFEE_API_KEY", "fake-key")
    capturado = {}

    def fake_get(url, timeout=None, verify=None):
        capturado["url"] = url
        return _FakeResp(payload=_JSON_ALL)

    monkeypatch.setattr(httpx, "get", fake_get)
    from coffee_module import client, db
    nota = client.buscar_nota(355617)
    assert nota["pk"] == 355617
    assert nota["id_sap"] == 17247854
    assert nota["arquivado"] is True
    assert nota["fields"]["sintoma"] == "EEST"
    assert capturado["url"].endswith("/deolhonarede/json_all/355617")
    logs = db.listar_logs(tipo="api_call")
    assert len(logs) == 1 and logs[0]["acao"] == "buscar_nota" and logs[0]["sucesso"] is True
    assert "tempo_ms" in logs[0]["detalhes"]


def test_buscar_nota_propaga_erro_http(coffee_tmp, monkeypatch):
    monkeypatch.setattr(config, "COFFEE_API_KEY", "fake-key")
    monkeypatch.setattr(httpx, "get", lambda url, timeout=None, verify=None: _FakeResp(status=500))
    from coffee_module import client, db
    with pytest.raises(httpx.HTTPStatusError):
        client.buscar_nota(1)
    logs = db.listar_logs(tipo="api_call")
    assert len(logs) == 1 and logs[0]["sucesso"] is False
    assert logs[0]["detalhes"]["status_http"] == 500


def test_buscar_nota_inexistente_erro_claro(coffee_tmp, monkeypatch):
    # json_all devolve 200 com lista vazia quando o id nao existe no COFFEE
    monkeypatch.setattr(config, "COFFEE_API_KEY", "fake-key")
    monkeypatch.setattr(httpx, "get", lambda url, timeout=None, verify=None: _FakeResp(payload="[]"))
    from coffee_module import client, db
    with pytest.raises(client.NotaNaoEncontradaErro):
        client.buscar_nota(999)
    logs = db.listar_logs(tipo="api_call")
    assert len(logs) == 1 and logs[0]["sucesso"] is False
    assert "999" in logs[0]["detalhes"]["erro"]


def test_escritas_montam_url(coffee_tmp, monkeypatch):
    monkeypatch.setattr(config, "COFFEE_API_KEY", "fake-key")
    urls = []

    def fake_get(url, timeout=None, verify=None):
        urls.append(url)
        return _FakeResp(payload="ok")

    monkeypatch.setattr(httpx, "get", fake_get)
    from coffee_module import client, db
    assert client.definir_sap(123321, 10000000) is True
    assert client.desarquivar(123321) is True
    assert client.alterar_local(123321, "701CF12345678") is True
    assert urls[0].endswith("/deolhonarede/sap/123321/10000000")
    assert urls[1].endswith("/deolhonarede/desarquivar/123321")
    assert urls[2].endswith("/deolhonarede/local_instalacao/123321/701CF12345678")
    acoes = {l["acao"] for l in db.listar_logs(tipo="api_call")}
    assert {"definir_sap", "desarquivar", "alterar_local"} <= acoes


def test_ssl_verify_le_env(monkeypatch):
    monkeypatch.delenv("COFFEE_SSL_VERIFY", raising=False)
    assert config.ssl_verify() is False  # padrao: CA corporativo auto-assinado
    monkeypatch.setenv("COFFEE_SSL_VERIFY", "true")
    assert config.ssl_verify() is True
    monkeypatch.setenv("COFFEE_SSL_VERIFY", "/etc/ssl/corp-ca.pem")
    assert config.ssl_verify() == "/etc/ssl/corp-ca.pem"


# ---------------------------------------------------------------------------
# Task 4 — jobs.py (batch fetch runner with progress)
# ---------------------------------------------------------------------------


def _aguardar_job(jobs, job_id, limite_s=3.0):
    """Faz polling do job até concluir (ou estourar o tempo)."""
    fim = _time.time() + limite_s
    while _time.time() < fim:
        j = jobs.obter_job(job_id)
        if j and j["estado"] == "concluido":
            return j
        _time.sleep(0.01)
    raise TimeoutError("job não concluiu a tempo")


def test_job_busca_lote_com_progresso_e_erros(coffee_tmp, monkeypatch):
    from coffee_module import client, db, jobs

    def fake_buscar(id):
        if str(id) == "999":
            raise RuntimeError("timeout")
        return {"pk": int(id), "id_sap": 17247854, "arquivado": False,
                "fields": {"id_sap": 17247854}}

    monkeypatch.setattr(client, "buscar_nota", fake_buscar)
    job_id = jobs.iniciar_busca(["355617", "999", "355618"])
    j = _aguardar_job(jobs, job_id)
    assert j["total"] == 3
    assert j["feitas"] == 3
    assert len(j["erros"]) == 1
    assert j["erros"][0]["pk"] == "999"
    # as duas notas válidas foram persistidas
    assert len(db.listar_notas("gerada")) == 2


def test_obter_job_inexistente(coffee_tmp):
    from coffee_module import jobs
    assert jobs.obter_job("nao-existe") is None


# ---------------------------------------------------------------------------
# Task 5 — routes.py (API router)
# ---------------------------------------------------------------------------

from fastapi.testclient import TestClient


@pytest.fixture
def coffee_cliente(coffee_tmp, monkeypatch):
    from coffee_module import client
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda id: {"pk": int(id), "id_sap": 17247854, "arquivado": False,
                    "local_instalacao": None, "fields": {"id_sap": 17247854}},
    )
    from main import app
    return TestClient(app)


def test_rota_buscar_job_e_notas(coffee_cliente):
    from coffee_module import jobs
    r = coffee_cliente.post("/api/coffee/buscar", json={"ids": ["355617"]})
    assert r.status_code == 200
    job_id = r.json()["job_id"]
    _aguardar_job(jobs, job_id)
    rj = coffee_cliente.get(f"/api/coffee/job/{job_id}")
    assert rj.json()["feitas"] == 1
    notas = coffee_cliente.get("/api/coffee/notas").json()["registros"]
    assert len(notas) == 1 and notas[0]["pk"] == 355617
    assert coffee_cliente.get("/api/coffee/notas?status=gerada").json()["registros"][0]["pk"] == 355617


def test_rota_buscar_lista_vazia_400(coffee_cliente):
    assert coffee_cliente.post("/api/coffee/buscar", json={"ids": []}).status_code == 400


def test_rota_job_inexistente_404(coffee_cliente):
    assert coffee_cliente.get("/api/coffee/job/nao-existe").status_code == 404


def test_rotas_de_escrita(coffee_cliente, monkeypatch):
    from coffee_module import client
    chamadas = []
    local = "701CF12345678"
    monkeypatch.setattr(client, "definir_sap", lambda i, s: chamadas.append(("sap", i, s)) or True)
    monkeypatch.setattr(client, "desarquivar", lambda i: chamadas.append(("des", i)) or True)
    monkeypatch.setattr(client, "alterar_local", lambda i, l: chamadas.append(("loc", i, l)) or True)
    monkeypatch.setattr(
        client,
        "buscar_nota",
        lambda i: {
            "pk": int(i),
            "id_sap": 17247854,
            "arquivado": False,
            "local_instalacao": local,
            "fields": {"id_sap": 17247854},
        },
    )
    assert coffee_cliente.post("/api/coffee/sap", json={"id": 1, "sap": 10000000}).json()["ok"] is True
    assert coffee_cliente.post("/api/coffee/desarquivar", json={"id": 1}).json()["ok"] is True
    resposta_local = coffee_cliente.post(
        "/api/coffee/local-instalacao",
        json={"id": 1, "local": local},
    )
    assert resposta_local.json() == {"ok": True, "local_instalacao": local}
    assert ("sap", 1, 10000000) in chamadas


def test_rota_local_rejeita_formato_invalido_sem_chamar_coffee(
    coffee_cliente,
    monkeypatch,
):
    from coffee_module import client

    chamadas = []
    monkeypatch.setattr(
        client,
        "alterar_local",
        lambda i, local: chamadas.append((i, local)) or True,
    )

    resposta = coffee_cliente.post(
        "/api/coffee/local-instalacao",
        json={"id": 1, "local": "701-CF-123"},
    )

    assert resposta.status_code == 400
    assert "13" in resposta.json()["detail"]
    assert chamadas == []


def test_rota_local_rejeita_sucesso_nao_confirmado(coffee_cliente, monkeypatch):
    from coffee_module import client

    monkeypatch.setattr(client, "alterar_local", lambda i, local: True)
    monkeypatch.setattr(
        client,
        "buscar_nota",
        lambda i: {
            "pk": int(i),
            "id_sap": 17247854,
            "arquivado": False,
            "local_instalacao": "701CF99999999",
            "fields": {"id_sap": 17247854},
        },
    )

    resposta = coffee_cliente.post(
        "/api/coffee/local-instalacao",
        json={"id": 1, "local": "701CF12345678"},
    )

    assert resposta.status_code == 409
    assert "não confirmou" in resposta.json()["detail"]


def test_listar_alimentadores_carrega_do_csv():
    from coffee_module import alimentadores
    registros = alimentadores.listar()
    assert len(registros) > 1000
    assert {"id", "cidade"} <= registros[0].keys()
    assert alimentadores.alimentador_valido(registros[0]["id"])
    assert not alimentadores.alimentador_valido("NAO_EXISTE_999")


def test_listar_municipios_carrega_do_csv():
    from coffee_module import municipios
    registros = municipios.listar()
    assert len(registros) == 28
    assert {"codigo", "nome"} <= registros[0].keys()
    assert len(registros[0]["codigo"]) == 3
    assert municipios.municipio_valido(registros[0]["codigo"])
    assert not municipios.municipio_valido("999")


def test_listar_tipos_equipamento_carrega_do_csv():
    from coffee_module import tipos_equipamento
    registros = tipos_equipamento.listar()
    assert len(registros) == 15
    assert {"id", "descricao"} <= registros[0].keys()
    assert tipos_equipamento.tipo_equipamento_valido(registros[0]["id"])
    assert not tipos_equipamento.tipo_equipamento_valido("NAO_EXISTE")


def test_rota_municipios_expoe_registros(coffee_cliente):
    resposta = coffee_cliente.get("/api/coffee/municipios")
    assert resposta.status_code == 200
    registros = resposta.json()["registros"]
    assert len(registros) == 28


def test_rota_tipos_equipamento_expoe_registros(coffee_cliente):
    resposta = coffee_cliente.get("/api/coffee/tipos-equipamento")
    assert resposta.status_code == 200
    registros = resposta.json()["registros"]
    assert len(registros) == 15


def test_rota_alimentador_confirma_e_atualiza(coffee_cliente, monkeypatch):
    from coffee_module import client
    chamadas = []
    monkeypatch.setattr(client, "alterar_alimentador",
                         lambda i, a: chamadas.append((i, a)) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "local_instalacao": "701CF12345678",
                   "fields": {"id_sap": 17247854, "alimentador": "AFC01"}},
    )

    resposta = coffee_cliente.post("/api/coffee/alimentador", json={"id": 1, "alimentador": "AFC01"})

    assert resposta.json() == {"ok": True, "alimentador": "AFC01"}
    assert (1, "AFC01") in chamadas


def test_rota_alimentador_rejeita_id_desconhecido_sem_chamar_coffee(coffee_cliente, monkeypatch):
    from coffee_module import client
    chamadas = []
    monkeypatch.setattr(client, "alterar_alimentador", lambda i, a: chamadas.append((i, a)) or True)

    resposta = coffee_cliente.post("/api/coffee/alimentador", json={"id": 1, "alimentador": "NAO_EXISTE_999"})

    assert resposta.status_code == 400
    assert chamadas == []


def test_rota_alimentador_rejeita_sucesso_nao_confirmado(coffee_cliente, monkeypatch):
    from coffee_module import client
    monkeypatch.setattr(client, "alterar_alimentador", lambda i, a: True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "local_instalacao": "701CF12345678",
                   "fields": {"id_sap": 17247854, "alimentador": "OUTRO01"}},
    )

    resposta = coffee_cliente.post("/api/coffee/alimentador", json={"id": 1, "alimentador": "AFC01"})

    assert resposta.status_code == 409
    assert "não confirmou" in resposta.json()["detail"]


def test_rota_consultar_expoe_referencia_eletrica_alimentador_e_campos_crus(coffee_cliente, monkeypatch):
    from coffee_module import client
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "local_instalacao": "718ET00026773",
                   "fields": {"id_sap": 17247854, "referencia_fisica": "SER-11",
                              "referencia_eletrica": "ELE-22", "alimentador": "AFC01"}},
    )
    r = coffee_cliente.get("/api/coffee/consultar/355617")
    body = r.json()
    assert body["referencia_fisica"] == "SER-11"
    assert body["referencia_eletrica"] == "ELE-22"
    assert body["alimentador"] == "AFC01"
    assert body["campos"]["referencia_eletrica"] == "ELE-22"


def test_rota_local_nao_encaminha_nota_sozinha(coffee_cliente, monkeypatch):
    from coffee_module import client, db

    local = "701CF12345678"
    monkeypatch.setattr(client, "alterar_local", lambda i, valor: True)
    monkeypatch.setattr(
        client,
        "buscar_nota",
        lambda i: {
            "pk": int(i),
            "id_sap": None,
            "arquivado": False,
            "local_instalacao": local,
            "fields": {
                "id_sap": None,
                "cidade": "701",
                "tipo_local_instalacao": "CF",
                "local_instalacao_numero": 12345678,
            },
        },
    )

    resposta = coffee_cliente.post(
        "/api/coffee/local-instalacao",
        json={"id": 1, "local": local},
        headers={"X-User": "alice"},
    )

    assert resposta.status_code == 200
    assert db.listar_itens_operacao() == []
    espelho = db.obter_nota(1)
    assert espelho["usuario"] == "alice"
    campos = espelho["dados_json"]
    assert campos["cidade"] == "701"
    assert campos["tipo_local_instalacao"] == "CF"
    assert campos["local_instalacao_numero"] == 12345678


def test_rota_consultar_nao_persiste_nota_do_header(coffee_cliente):
    """Consulta usada pela busca de duplicata não deve criar estado local."""
    from coffee_module import db
    r = coffee_cliente.get("/api/coffee/consultar/999", headers={"X-User": "alice"})
    assert r.status_code == 200
    assert db.obter_nota(999) is None
    de_alice = coffee_cliente.get("/api/coffee/notas", headers={"X-User": "alice"}).json()["registros"]
    assert de_alice == []
    de_bob = coffee_cliente.get("/api/coffee/notas", headers={"X-User": "bob"}).json()["registros"]
    assert de_bob == []


def test_rota_exportar_concluidas_gera_xlsx_so_com_notas_concluidas(coffee_cliente):
    """O Excel de concluídas usa somente o espelho local já filtrado por status."""
    from openpyxl import load_workbook
    from coffee_module import db

    db.upsert_nota(101, 17247854, {
        "cidade": "Serra",
        "tipo_local_instalacao": "ET",
        "local_instalacao_numero": "00026773",
        "postes": "P-77",
        "referencia_fisica": "Rua da Consulta",
        "componente": "Chave",
        "sintoma": "Queda",
        "observacoes": "Validada pela engenharia",
    })
    db.upsert_nota(102, 10000000, {"cidade": "Vitória"})

    response = coffee_cliente.post(
        "/api/coffee/notas/concluidas/exportar",
        json={"pks": [101, 102]},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    assert "notas_concluidas" in response.headers["content-disposition"]
    worksheet = load_workbook(BytesIO(response.content)).active
    rows = list(worksheet.values)
    assert rows[0][:4] == ("ID ONR", "ID SAP", "Classificação", "Local de instalação")
    assert rows[1][:8] == (
        101,
        17247854,
        "Gerada",
        "Serra-ET-00026773",
        "P-77",
        "Rua da Consulta",
        "Chave",
        "Queda",
    )
    assert len(rows) == 2
    assert worksheet.freeze_panes == "A2"


def test_rota_exportar_concluidas_nao_expoe_notas_de_outro_usuario(coffee_cliente):
    """O header X-User limita a planilha a notas concluídas do próprio usuário."""
    from openpyxl import load_workbook
    from coffee_module import db

    try:
        db.definir_usuario("alice")
        db.upsert_nota(201, 17247854, {"observacoes": "Nota da Alice"})
        db.definir_usuario("bob")
        db.upsert_nota(202, 17247855, {"observacoes": "Nota do Bob"})

        response = coffee_cliente.post(
            "/api/coffee/notas/concluidas/exportar",
            json={"pks": [201, 202]},
            headers={"X-User": "alice"},
        )

        assert response.status_code == 200
        rows = list(load_workbook(BytesIO(response.content)).active.values)
        assert [row[0] for row in rows[1:]] == [201]
        assert rows[1][8] == "Nota da Alice"
    finally:
        db.definir_usuario(None)


def test_planilha_concluidas_neutraliza_formula_de_texto_externo():
    """Campos vindos do COFFEE não podem executar fórmulas ao abrir no Excel."""
    from openpyxl import load_workbook
    from coffee_module.exportacao import gerar_planilha_concluidas

    arquivo = gerar_planilha_concluidas([{
        "pk": 101,
        "id_sap": 17247854,
        "classificacao": "gerada",
        "dados_json": {"observacoes": "=HYPERLINK(\"https://exemplo.invalid\")"},
    }])

    observacao = load_workbook(BytesIO(arquivo)).active["I2"]
    assert observacao.data_type == "s"
    assert observacao.value == "'=HYPERLINK(\"https://exemplo.invalid\")"


def test_rota_buscar_log_acao_com_usuario_do_header(coffee_cliente):
    from coffee_module import db, jobs
    r = coffee_cliente.post("/api/coffee/buscar", json={"ids": ["777"]},
                            headers={"X-User": "bob"})
    _aguardar_job(jobs, r.json()["job_id"])
    lote = [l for l in db.listar_logs(tipo="acao_usuario") if l["acao"] == "busca_lote"]
    assert lote and lote[0]["usuario"] == "bob"


# ---------------------------------------------------------------------------
# Task 4 — Routes /logs, /regerar + acao_usuario logging
# ---------------------------------------------------------------------------


def test_rota_buscar_registra_acao_usuario(coffee_cliente):
    from coffee_module import jobs, db
    r = coffee_cliente.post("/api/coffee/buscar", json={"ids": ["355617", "355618"]})
    _aguardar_job(jobs, r.json()["job_id"])
    lote = [l for l in db.listar_logs(tipo="acao_usuario") if l["acao"] == "busca_lote"]
    assert len(lote) == 1
    assert lote[0]["detalhes"]["total"] == 2


def test_rota_logs_filtra(coffee_cliente):
    from coffee_module import db
    db.registrar_log("api_call", "buscar_nota", 1, {"id": 1}, True)
    db.registrar_log("transicao", "classificar", 1, {"x": 1}, True)
    todos = coffee_cliente.get("/api/coffee/logs").json()["logs"]
    assert len(todos) >= 2
    so_api = coffee_cliente.get("/api/coffee/logs?tipo=api_call").json()["logs"]
    assert all(l["tipo"] == "api_call" for l in so_api)
    so_nota = coffee_cliente.get("/api/coffee/logs?nota_pk=1").json()["logs"]
    assert all(l["nota_pk"] == 1 for l in so_nota)


def test_rota_regerar(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    chamadas = []
    monkeypatch.setattr(client, "desarquivar", lambda i: chamadas.append(("des", i)) or True)
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: chamadas.append(("sap", i, sap)) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 10000000, "arquivado": False,
                   "fields": {"id_sap": 10000000}},
    )
    r = coffee_cliente.post("/api/coffee/regerar",
                            json={"id": 355617, "justificativa": "reprocessar"})
    assert r.status_code == 200
    body = r.json()
    assert body["ok"] is True
    assert body["nota"]["pk"] == 355617
    # define SAP=10000000 E desarquiva: o COFFEE so gera notas desarquivadas
    assert ("sap", 355617, 10000000) in chamadas
    assert ("des", 355617) in chamadas
    # nota re-buscada com 10000000 fica pendente
    assert db.listar_notas("pendente")[0]["pk"] == 355617
    log = [l for l in db.listar_logs(tipo="acao_usuario") if l["acao"] == "regerar"]
    assert log and log[0]["detalhes"]["justificativa"] == "reprocessar"


# ---------------------------------------------------------------------------
# Sub-projeto 3 — usuario nos logs
# ---------------------------------------------------------------------------


def test_log_grava_usuario(coffee_tmp, monkeypatch):
    from coffee_module import db
    monkeypatch.setattr(db.getpass, "getuser", lambda: "operador.teste")
    db.registrar_log("acao_usuario", "x", None, None, True)
    logs = db.listar_logs()
    assert logs[0]["usuario"] == "operador.teste"


def test_usuario_atual_fallback_nunca_levanta(coffee_tmp, monkeypatch):
    from coffee_module import db

    def boom():
        raise OSError("sem tty")

    monkeypatch.setattr(db.getpass, "getuser", boom)
    monkeypatch.setenv("USERNAME", "via.env")
    assert db._usuario_atual() == "via.env"


# ---------------------------------------------------------------------------
# Sub-projeto 3 — flag a_gerar
# ---------------------------------------------------------------------------


def test_marcar_gerar_e_listar(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    assert db.listar_notas("a_gerar") == []
    db.marcar_gerar(355617, True)
    aged = db.listar_notas("a_gerar")
    assert len(aged) == 1 and aged[0]["pk"] == 355617
    assert aged[0]["a_gerar"] is True


def test_marcar_gerar_falso_remove_da_lista(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(1, 17247854, {})
    db.marcar_gerar(1, True)
    db.marcar_gerar(1, False)
    assert db.listar_notas("a_gerar") == []


def test_a_gerar_preservado_em_refetch(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(1, 10000000, {"id_sap": 10000000})
    db.marcar_gerar(1, True)
    db.upsert_nota(1, 17247854, {"id_sap": 17247854})  # re-busca
    assert db.listar_notas("a_gerar")[0]["pk"] == 1


def test_nota_existe(coffee_tmp):
    from coffee_module import db
    assert db.nota_existe(99) is False
    db.upsert_nota(99, 10000000, {})
    assert db.nota_existe(99) is True


def test_obter_nota(monkeypatch, tmp_path):
    monkeypatch.setenv("COFFEE_DATA_DIR", str(tmp_path))
    from coffee_module import db
    db.inicializar_banco()
    db.upsert_nota(4242, 12345678, {"prioridade": 3, "observacoes": "Trocar poste"})
    nota = db.obter_nota(4242)
    assert nota is not None
    assert nota["id_sap"] == 12345678
    assert nota["dados_json"]["observacoes"] == "Trocar poste"
    assert nota["a_gerar"] is False
    assert db.obter_nota(999999) is None


def test_obter_nota_ignora_arquivada(monkeypatch, tmp_path):
    """obter_nota respeita o mesmo filtro de arquivamento local de listar_notas."""
    monkeypatch.setenv("COFFEE_DATA_DIR", str(tmp_path))
    from coffee_module import db
    db.inicializar_banco()
    db.upsert_nota(4242, 12345678, {"prioridade": 3, "observacoes": "Trocar poste"})
    assert db.obter_nota(4242) is not None
    db.arquivar_nota(4242)
    assert db.obter_nota(4242) is None


def test_desarquivar_nota_reverte_arquivar_nota(monkeypatch, tmp_path):
    """desarquivar_nota devolve a nota arquivada localmente pra visibilidade normal."""
    monkeypatch.setenv("COFFEE_DATA_DIR", str(tmp_path))
    from coffee_module import db
    db.inicializar_banco()
    db.upsert_nota(4242, 12345678, {"prioridade": 3, "observacoes": "Trocar poste"})
    db.arquivar_nota(4242)
    assert db.obter_nota(4242) is None
    db.desarquivar_nota(4242)
    assert db.obter_nota(4242) is not None


# ---------------------------------------------------------------------------
# Task 3 — Routes /marcar-gerar + limpeza de a_gerar no /regerar
# ---------------------------------------------------------------------------


def test_rota_marcar_gerar_sap_real_existente_sai_da_fila(coffee_cliente):
    from coffee_module import db
    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    r = coffee_cliente.post("/api/coffee/marcar-gerar", json={"id": 355617, "a_gerar": True})
    assert r.status_code == 200 and r.json()["ok"] is True
    assert db.listar_notas("a_gerar") == []
    assert db.obter_nota(355617)["a_gerar"] is False
    assert any(l["acao"] == "marcar_gerar" for l in db.listar_logs(tipo="acao_usuario"))


def test_rota_marcar_gerar_sap_real_busca_e_desmarca(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "fields": {"id_sap": 17247854}},
    )
    r = coffee_cliente.post("/api/coffee/marcar-gerar", json={"id": 355617, "a_gerar": True})
    assert r.status_code == 200
    assert db.nota_existe(355617) is True
    assert db.listar_notas("a_gerar") == []
    assert db.obter_nota(355617)["a_gerar"] is False


def test_rota_marcar_gerar_falha_busca_502(coffee_cliente, monkeypatch):
    from coffee_module import client, db

    def boom(i):
        raise RuntimeError("falha API")

    monkeypatch.setattr(client, "buscar_nota", boom)
    r = coffee_cliente.post("/api/coffee/marcar-gerar", json={"id": 999, "a_gerar": True})
    assert r.status_code == 502
    assert any(l["acao"] == "marcar_gerar" and l["sucesso"] is False
               for l in db.listar_logs(tipo="acao_usuario"))


def test_rota_marcar_gerar_nota_inexistente_404(coffee_cliente, monkeypatch):
    from coffee_module import client

    def nao_encontrada(i):
        raise client.NotaNaoEncontradaErro(i)

    monkeypatch.setattr(client, "buscar_nota", nao_encontrada)
    r = coffee_cliente.post("/api/coffee/marcar-gerar", json={"id": 999, "a_gerar": True})
    assert r.status_code == 404
    assert "999" in r.json()["detail"]


def test_rota_marcar_gerar_false_com_justificativa_tira_da_fila(coffee_cliente):
    from coffee_module import db
    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    db.marcar_gerar(355617, True)
    r = coffee_cliente.post("/api/coffee/marcar-gerar",
                            json={"id": 355617, "a_gerar": False,
                                  "justificativa": "Nota reaberta na Verificar"})
    assert r.status_code == 200
    assert db.listar_notas("a_gerar") == []


def test_rota_regerar_limpa_a_gerar(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    db.upsert_nota(355617, 10000000, {"id_sap": 10000000})
    db.marcar_gerar(355617, True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 10000000, "arquivado": False,
                   "fields": {"id_sap": 10000000}},
    )
    r = coffee_cliente.post("/api/coffee/regerar", json={"id": 355617})
    assert r.status_code == 200
    assert db.listar_notas("a_gerar") == []


# ---------------------------------------------------------------------------
# 2026-06-27 — consulta sincrona para o modal
# ---------------------------------------------------------------------------


def test_rota_consultar_retorna_campos(coffee_cliente, monkeypatch):
    from coffee_module import client
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "local_instalacao": "718ET00026773",
                   "fields": {"id_sap": 17247854}},
    )
    r = coffee_cliente.get("/api/coffee/consultar/355617")
    assert r.status_code == 200
    body = r.json()
    assert body["pk"] == 355617
    assert body["id_sap"] == 17247854
    assert body["local_instalacao"] == "718ET00026773"
    assert body["classificacao"] == "gerada"
    assert body["arquivado"] is False


def test_rota_consultar_retorna_poste_e_referencia(coffee_cliente, monkeypatch):
    from coffee_module import client
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "local_instalacao": "718ET00026773",
                   "fields": {"id_sap": 17247854, "postes": "TR-088",
                              "referencia_fisica": "SER-11"}},
    )
    r = coffee_cliente.get("/api/coffee/consultar/355617")
    assert r.status_code == 200
    body = r.json()
    assert body["poste"] == "TR-088"
    assert body["referencia"] == "SER-11"


def test_rota_consultar_poste_referencia_ausentes_vira_none(coffee_cliente, monkeypatch):
    from coffee_module import client
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "local_instalacao": None},
    )
    r = coffee_cliente.get("/api/coffee/consultar/355617")
    body = r.json()
    assert body["poste"] is None
    assert body["referencia"] is None
    assert body["problema"] is None
    assert body["observacao"] is None


def test_rota_consultar_projeta_problema_e_observacao_sem_persistir(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "local_instalacao": "718ET00026773",
                   "fields": {"componente_novo": "Luminaria", "sintoma": "Apagada",
                              "causa": "Lampada queimada", "observacoes": "Trocar lampada"}},
    )

    body = coffee_cliente.get("/api/coffee/consultar/355617").json()

    assert body["problema"] == "Luminaria · Apagada · Lampada queimada"
    assert body["observacao"] == "Trocar lampada"
    assert db.obter_nota(355617) is None


def test_compor_local_instalacao():
    from coffee_module import client
    # cidade(3) + tipo(2) + numero(8, zero a esquerda) — formato real da API COFFEE
    assert client.compor_local_instalacao(
        {"cidade": "718", "tipo_local_instalacao": "ET", "local_instalacao_numero": 26773}
    ) == "718ET00026773"
    # cidade com menos de 3 digitos recebe zero a esquerda
    assert client.compor_local_instalacao(
        {"cidade": "45", "tipo_local_instalacao": "CF", "local_instalacao_numero": 25416}
    ) == "045CF00025416"
    # falta componente -> None
    assert client.compor_local_instalacao({"cidade": "718", "tipo_local_instalacao": "ET"}) is None
    assert client.compor_local_instalacao({}) is None


def test_rota_consultar_nota_inexistente_404(coffee_cliente, monkeypatch):
    from coffee_module import client

    def nao_encontrada(i):
        raise client.NotaNaoEncontradaErro(i)

    monkeypatch.setattr(client, "buscar_nota", nao_encontrada)
    r = coffee_cliente.get("/api/coffee/consultar/999")
    assert r.status_code == 404
    assert "999" in r.json()["detail"]


def test_rota_consultar_falha_502(coffee_cliente, monkeypatch):
    from coffee_module import client

    def boom(i):
        raise RuntimeError("falha API")

    monkeypatch.setattr(client, "buscar_nota", boom)
    assert coffee_cliente.get("/api/coffee/consultar/999").status_code == 502


# ---------------------------------------------------------------------------
# Sub-projeto 4 — status nao_gerada
# ---------------------------------------------------------------------------


def test_classificacao_nao_gerada():
    from coffee_module import classify
    assert classify.classificar(None, None) == "nao_gerada"
    assert classify.classificar(0, None) == "nao_gerada"
    assert classify.classificar("", None) == "nao_gerada"
    # sem SAP atual = nao_gerada mesmo com anterior conhecido
    assert classify.classificar(None, config.SAP_PENDENTE) == "nao_gerada"


def test_upsert_nota_sem_sap_classifica_nao_gerada(coffee_tmp):
    from coffee_module import db
    classe = db.upsert_nota(355617, None, {"id_sap": None})
    assert classe == "nao_gerada"
    assert db.listar_notas("nao_gerada")[0]["pk"] == 355617


# ---------------------------------------------------------------------------
# Geração em lote (jobs.iniciar_geracao + /gerar-lote)
# ---------------------------------------------------------------------------


def test_job_geracao_define_sap_e_isola_erro(coffee_tmp, monkeypatch):
    from coffee_module import client, db, jobs
    saps = []
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, sap: saps.append((int(i), sap)) or True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)

    def fake_buscar(id):
        if str(id) == "999":
            raise RuntimeError("timeout")
        return {"pk": int(id), "id_sap": 10000000, "arquivado": False,
                "fields": {"id_sap": 10000000}}

    monkeypatch.setattr(client, "buscar_nota", fake_buscar)
    job_id = jobs.iniciar_geracao([355617, 999, 355618], justificativa="lote x")
    j = _aguardar_job(jobs, job_id)
    assert j["total"] == 3
    assert j["feitas"] == 3
    assert len(j["erros"]) == 1
    assert j["erros"][0]["pk"] == 999
    # SAP=10000000 definido para as duas válidas
    assert (355617, 10000000) in saps and (355618, 10000000) in saps
    # válidas persistidas como pendentes e fora da fila
    assert len(db.listar_notas("pendente")) == 2
    assert db.listar_notas("a_gerar") == []


def test_rota_gerar_lote(coffee_cliente, monkeypatch):
    from coffee_module import client, db, jobs
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 10000000, "arquivado": False,
                   "fields": {"id_sap": 10000000}},
    )
    r = coffee_cliente.post("/api/coffee/gerar-lote",
                            json={"ids": [355617, 355618], "justificativa": "j"})
    assert r.status_code == 200
    _aguardar_job(jobs, r.json()["job_id"])
    lote = [l for l in db.listar_logs(tipo="acao_usuario") if l["acao"] == "geracao_lote"]
    assert lote and lote[0]["detalhes"]["total"] == 2
    assert lote[0]["detalhes"]["justificativa"] == "j"


def test_rota_gerar_lote_vazio_400(coffee_cliente):
    assert coffee_cliente.post("/api/coffee/gerar-lote", json={"ids": []}).status_code == 400


# ---------------------------------------------------------------------------
# Task 4 — /marcar-gerar com justificativa obrigatoria ao remover da fila
# ---------------------------------------------------------------------------


def test_marcar_gerar_remover_exige_justificativa(coffee_cliente):
    from coffee_module import db
    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    db.marcar_gerar(355617, True)
    # sem justificativa ao remover → 400
    r = coffee_cliente.post("/api/coffee/marcar-gerar",
                            json={"id": 355617, "a_gerar": False})
    assert r.status_code == 400
    assert db.listar_notas("a_gerar")[0]["pk"] == 355617  # continua na fila


def test_marcar_gerar_remover_com_justificativa(coffee_cliente):
    from coffee_module import db
    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    db.marcar_gerar(355617, True)
    r = coffee_cliente.post("/api/coffee/marcar-gerar",
                            json={"id": 355617, "a_gerar": False,
                                  "justificativa": "posta por engano"})
    assert r.status_code == 200
    assert db.listar_notas("a_gerar") == []
    log = [l for l in db.listar_logs(tipo="acao_usuario") if l["acao"] == "marcar_gerar"]
    assert log and log[0]["detalhes"]["justificativa"] == "posta por engano"


# ---------------------------------------------------------------------------
# Verify batch — Task 5: geração checa arquivado antes de gerar
# ---------------------------------------------------------------------------


def test_geracao_nota_arquivada_nao_define_sap(coffee_tmp, monkeypatch):
    from coffee_module import client, db, jobs
    saps = []
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, sap: saps.append((int(i), sap)) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": True,
                   "local_instalacao": "701CF999",
                   "fields": {"id_sap": 17247854}},
    )
    job_id = jobs.iniciar_geracao([355617])
    j = _aguardar_job(jobs, job_id)
    assert saps == []  # arquivada: nunca define SAP
    assert j["arquivadas"] == [
        {"pk": 355617, "id_sap": 17247854, "local_instalacao": "701CF999"}
    ]
    assert db.listar_notas("pendente") == []


def test_geracao_nota_arquivada_remove_da_fila(coffee_tmp, monkeypatch):
    """Nota arquivada deve ter a_gerar desmarcado apos geracao, saindo da fila."""
    from coffee_module import client, db, jobs

    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    db.marcar_gerar(355617, True)
    assert db.listar_notas("a_gerar")[0]["pk"] == 355617  # esta na fila

    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": True,
                   "local_instalacao": "701CF999",
                   "fields": {"id_sap": 17247854}},
    )
    job_id = jobs.iniciar_geracao([355617])
    _aguardar_job(jobs, job_id)
    assert db.listar_notas("a_gerar") == []  # deve ter saido da fila


def test_geracao_busca_antes_de_definir_sap(coffee_tmp, monkeypatch):
    from coffee_module import client, db, jobs
    ordem = []
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, sap: ordem.append("sap") or True)
    monkeypatch.setattr(client, "desarquivar", lambda i: ordem.append("des") or True)

    def fake_buscar(i):
        ordem.append("buscar")
        return {"pk": int(i), "id_sap": 10000000, "arquivado": False,
                "fields": {"id_sap": 10000000}}

    monkeypatch.setattr(client, "buscar_nota", fake_buscar)
    job_id = jobs.iniciar_geracao([355617])
    _aguardar_job(jobs, job_id)
    assert ordem[0] == "buscar"  # GET antes de definir_sap
    assert "sap" in ordem
    assert db.listar_notas("pendente")[0]["pk"] == 355617


# ---------------------------------------------------------------------------
# Verify batch — Task 3a: diagnóstico de transição (caracterização)
# ---------------------------------------------------------------------------


def test_diagnosticar_nota_retorna_estado_e_logs(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(356322, 10000000, {"id_sap": 10000000})  # pendente
    diag = db.diagnosticar_nota(356322)
    assert diag["pk"] == 356322
    assert diag["id_sap"] == 10000000
    assert diag["classificacao"] == "pendente"
    assert isinstance(diag["logs"], list)
    assert db.diagnosticar_nota(999999) is None


def test_caracteriza_avulsa_atualmente_vira_corrigida(coffee_tmp):
    """Task 3b: nota avulsa (pendente -> SAP real) com origem='avulsa' é
    rotulada 'gerada', não 'corrigida'."""
    from coffee_module import db
    db.upsert_nota(355617, 10000000, {"id_sap": 10000000})
    db.definir_origem(355617, "avulsa")
    classe = db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    assert classe == "gerada"  # após Task 3b: avulsa corretamente vira gerada


# ---------------------------------------------------------------------------
# Verify batch — Task 3b: origem distingue avulsa (gerada) de corrigida
# ---------------------------------------------------------------------------


def test_classificacao_avulsa_vira_gerada():
    from coffee_module import classify, config
    assert classify.classificar(17247854, config.SAP_PENDENTE, "avulsa") == "gerada"


def test_classificacao_sem_origem_mantem_corrigida():
    from coffee_module import classify, config
    # backwards-compat: origem desconhecida continua corrigida
    assert classify.classificar(17247854, config.SAP_PENDENTE) == "corrigida"
    assert classify.classificar(17247854, config.SAP_PENDENTE, None) == "corrigida"


def test_upsert_avulsa_vira_gerada_apos_pendente(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(1, 10000000, {"id_sap": 10000000})  # pendente
    db.definir_origem(1, "avulsa")
    classe = db.upsert_nota(1, 17247854, {"id_sap": 17247854})
    assert classe == "gerada"
    assert db.listar_notas("corrigida") == []


def test_geracao_marca_origem_avulsa(coffee_tmp, monkeypatch):
    from coffee_module import client, db, jobs
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 10000000, "arquivado": False,
                   "fields": {"id_sap": 10000000}},
    )
    job_id = jobs.iniciar_geracao([355617])
    _aguardar_job(jobs, job_id)
    diag = db.diagnosticar_nota(355617)
    assert diag is not None
    # origem persistida; re-busca com SAP real classifica como gerada
    classe = db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    assert classe == "gerada"


# ---------------------------------------------------------------------------
# 2026-06-27 — bug 1: marcar-gerar liga a_gerar no pk resolvido (não no id)
# ---------------------------------------------------------------------------


def test_marcar_gerar_desmarca_pk_resolvido_nao_o_id(coffee_cliente, monkeypatch):
    """id de entrada (999) != pk real (355617): desmarca o pk resolvido."""
    from coffee_module import client, db
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": 355617, "id_sap": 17247854, "arquivado": False,
                   "fields": {"id_sap": 17247854}},
    )
    r = coffee_cliente.post("/api/coffee/marcar-gerar", json={"id": 999, "a_gerar": True})
    assert r.status_code == 200
    assert db.listar_notas("a_gerar") == []
    nota = db.obter_nota(355617)
    assert nota["a_gerar"] is False
    assert nota["verificar_id"] == 999
    assert nota["verificar_ativa"] is True


def test_marcar_gerar_grava_origem_verificar(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "fields": {"id_sap": 17247854}},
    )
    coffee_cliente.post("/api/coffee/marcar-gerar", json={"id": 355617, "a_gerar": True})
    assert db.origem_atual(355617) == "verificar"


# ---------------------------------------------------------------------------
# 2026-06-27 — bug 2: gerar nao toca SAP real; bug 3: origem preservada
# ---------------------------------------------------------------------------


def _SAP_REAL():
    return 17247854


def test_geracao_pula_nota_com_sap_real(coffee_tmp, monkeypatch):
    """Nota nao-arquivada com SAP real nao recebe placeholder; sai da fila."""
    from coffee_module import client, db, jobs
    saps = []
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, sap: saps.append((int(i), sap)) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": _SAP_REAL(), "arquivado": False,
                   "fields": {"id_sap": _SAP_REAL()}},
    )
    db.upsert_nota(355617, _SAP_REAL(), {"id_sap": _SAP_REAL()})
    db.marcar_gerar(355617, True)
    job_id = jobs.iniciar_geracao([355617])
    _aguardar_job(jobs, job_id)
    assert saps == []                              # nao definiu SAP
    assert db.listar_notas("a_gerar") == []        # saiu da fila
    ignorada = [l for l in db.listar_logs(tipo="acao_usuario")
                if l["acao"] == "geracao_ignorada_sap_real"]
    assert ignorada and ignorada[0]["nota_pk"] == 355617


def test_geracao_nao_sobrescreve_origem_verificar(coffee_tmp, monkeypatch):
    """Nota da Verificar gerada via lote mantem origem='verificar' (-> corrigida)."""
    from coffee_module import client, db, jobs
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": config.SAP_PENDENTE, "arquivado": False,
                   "fields": {"id_sap": config.SAP_PENDENTE}},
    )
    db.upsert_nota(355617, config.SAP_PENDENTE, {"id_sap": config.SAP_PENDENTE})
    db.definir_origem(355617, "verificar")
    job_id = jobs.iniciar_geracao([355617])
    _aguardar_job(jobs, job_id)
    assert db.origem_atual(355617) == "verificar"
    # re-busca com SAP real -> corrigida (origem != avulsa)
    classe = db.upsert_nota(355617, _SAP_REAL(), {"id_sap": _SAP_REAL()})
    assert classe == "corrigida"


def test_rota_regerar_pula_sap_real(coffee_cliente, monkeypatch):
    """regerar nao re-define SAP quando a nota ja tem SAP real (nao arquivada)."""
    from coffee_module import client, db
    saps = []
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: saps.append((int(i), sap)) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": False,
                   "fields": {"id_sap": 17247854}},
    )
    db.upsert_nota(355617, 17247854, {"id_sap": 17247854})
    db.marcar_gerar(355617, True)
    r = coffee_cliente.post("/api/coffee/regerar", json={"id": 355617})
    assert r.status_code == 200 and r.json()["ok"] is True
    assert saps == []  # nao definiu SAP
    assert db.listar_notas("a_gerar") == []  # saiu da fila
    ignorada = [l for l in db.listar_logs(tipo="acao_usuario")
                if l["acao"] == "geracao_ignorada_sap_real" and l["nota_pk"] == 355617]
    assert ignorada


# ---------------------------------------------------------------------------
# Logs git-graph — trace_id
# ---------------------------------------------------------------------------


def test_registrar_log_carimba_trace(coffee_tmp):
    from coffee_module import db
    db.definir_trace("abc123")
    db.registrar_log("acao_usuario", "x", None, {"k": 1}, True)
    db.definir_trace(None)
    db.registrar_log("acao_usuario", "y", None, None, True)
    logs = db.listar_logs()
    x = next(l for l in logs if l["acao"] == "x")
    y = next(l for l in logs if l["acao"] == "y")
    assert x["trace_id"] == "abc123"
    assert y["trace_id"] is None


def test_middleware_carimba_trace_na_requisicao(coffee_cliente):
    from coffee_module import db, jobs

    resposta = coffee_cliente.post("/api/coffee/buscar", json={"ids": ["1"]})
    _aguardar_job(jobs, resposta.json()["job_id"])
    lote = [l for l in db.listar_logs(tipo="acao_usuario") if l["acao"] == "busca_lote"]
    assert lote and lote[0]["trace_id"] is not None


def test_job_geracao_propaga_trace_aos_filhos(coffee_cliente, monkeypatch):
    from coffee_module import client, db, jobs
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda id: (
            db.registrar_log("api_call", "buscar_nota", int(id), {"id": id}, True),
            {"pk": int(id), "id_sap": 10000000, "arquivado": False,
             "local_instalacao": None, "fields": {"id_sap": 10000000}},
        )[-1],
    )
    monkeypatch.setattr(client, "definir_sap", lambda i, s: True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)
    r = coffee_cliente.post("/api/coffee/gerar-lote", json={"ids": [355617]})
    _aguardar_job(jobs, r.json()["job_id"])
    logs = db.listar_logs()
    lote = next(l for l in logs if l["acao"] == "geracao_lote")
    filhos = [l for l in logs if l["acao"] == "buscar_nota"]
    assert lote["trace_id"] is not None
    assert filhos and all(f["trace_id"] == lote["trace_id"] for f in filhos)


# ---------------------------------------------------------------------------
# Task 3 — consultar/alterar_local acao_usuario + filtro nota_pk inclui lote
# ---------------------------------------------------------------------------


def test_consultar_sucesso_loga_acao_usuario(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda id: {"pk": int(id), "id_sap": 17247854, "arquivado": False,
                    "local_instalacao": "718ET00026773", "fields": {"id_sap": 17247854}},
    )
    coffee_cliente.get("/api/coffee/consultar/44421")
    consultas = [l for l in db.listar_logs(tipo="acao_usuario")
                 if l["acao"] == "consultar" and l["sucesso"]]
    assert consultas and consultas[0]["nota_pk"] == 44421


def test_alterar_local_loga_acao_usuario(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    monkeypatch.setattr(client, "alterar_local", lambda i, l: True)
    coffee_cliente.post("/api/coffee/local-instalacao",
                        json={"id": 44421, "local": "718ET00026773"})
    locs = [l for l in db.listar_logs(tipo="acao_usuario") if l["acao"] == "alterar_local"]
    assert locs and locs[0]["nota_pk"] == 44421


def test_listar_logs_nota_inclui_cabecalho_de_lote(coffee_tmp):
    from coffee_module import db
    db.definir_trace("t1")
    db.registrar_log("acao_usuario", "geracao_lote", None, {"total": 2}, True)
    db.registrar_log("api_call", "buscar_nota", 44421, {"id": 44421}, True)
    db.definir_trace(None)
    db.registrar_log("acao_usuario", "outra", None, {}, True)
    acoes = {l["acao"] for l in db.listar_logs(nota_pk=44421)}
    assert "buscar_nota" in acoes       # filho da nota
    assert "geracao_lote" in acoes      # cabecalho do trace (nota_pk NULL)
    assert "outra" not in acoes         # sem trace, nao relacionado


# ---------------------------------------------------------------------------
# Task 1 — classificacao_em (idade da pendência)
# ---------------------------------------------------------------------------


def test_classificacao_em_gravada_e_preservada(coffee_tmp):
    from coffee_module import db
    db.upsert_nota(1, 10000000, {})
    t1 = db.listar_notas("pendente")[0]["classificacao_em"]
    assert t1  # gravada no nascimento da linha

    _time.sleep(0.01)
    db.upsert_nota(1, 10000000, {})  # re-busca, mesma classe
    assert db.listar_notas("pendente")[0]["classificacao_em"] == t1  # idade preservada

    _time.sleep(0.01)
    db.upsert_nota(1, 17247854, {})  # pendente -> corrigida
    t2 = db.listar_notas("corrigida")[0]["classificacao_em"]
    assert t2 > t1  # reclassificação atualiza


# ---------------------------------------------------------------------------
# Task 2 — param since em /coffee/logs
# ---------------------------------------------------------------------------


def test_listar_logs_since(coffee_tmp):
    from coffee_module import db
    db.registrar_log("acao_usuario", "primeira", None, None, True)
    _time.sleep(0.01)
    db.registrar_log("acao_usuario", "segunda", None, None, True)
    todos = db.listar_logs()
    corte = todos[0]["timestamp"]  # ordem DESC: [0] é "segunda"
    filtrados = db.listar_logs(since=corte)
    assert [l["acao"] for l in filtrados] == ["segunda"]

# ---------------------------------------------------------------------------
# 2026-07-06 — forcar geracao deixa a nota DESARQUIVADA no COFFEE
# (o COFFEE so gera notas desarquivadas: da o SAP real e arquiva sozinho)
# ---------------------------------------------------------------------------


def test_geracao_desarquiva_pendente_arquivada(coffee_tmp, monkeypatch):
    """SAP=10000000 + arquivada: define placeholder E desarquiva."""
    from coffee_module import client, jobs
    chamadas = []
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, sap: chamadas.append(("sap", int(i), sap)) or True)
    monkeypatch.setattr(client, "desarquivar",
                        lambda i: chamadas.append(("des", int(i))) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": config.SAP_PENDENTE, "arquivado": True,
                   "local_instalacao": None,
                   "fields": {"id_sap": config.SAP_PENDENTE}},
    )
    job_id = jobs.iniciar_geracao([355617])
    j = _aguardar_job(jobs, job_id)
    assert ("sap", 355617, config.SAP_PENDENTE) in chamadas
    assert ("des", 355617) in chamadas
    assert "arquivadas" not in j  # nao foi pulada


def test_geracao_desarquiva_nota_sem_sap_arquivada(coffee_tmp, monkeypatch):
    """Arquivada SEM SAP nao e 'fora da fila': forca placeholder + desarquiva."""
    from coffee_module import client, jobs
    chamadas = []
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, sap: chamadas.append(("sap", int(i), sap)) or True)
    monkeypatch.setattr(client, "desarquivar",
                        lambda i: chamadas.append(("des", int(i))) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": None, "arquivado": True,
                   "local_instalacao": None, "fields": {"id_sap": None}},
    )
    job_id = jobs.iniciar_geracao([355617])
    j = _aguardar_job(jobs, job_id)
    assert ("sap", 355617, config.SAP_PENDENTE) in chamadas
    assert ("des", 355617) in chamadas
    assert "arquivadas" not in j


def test_rota_regerar_arquivada_com_sap_real_desarquiva(coffee_cliente, monkeypatch):
    """Regerar explicito de nota ja gerada (SAP real + arquivada):
    volta ao placeholder e desarquiva para o COFFEE re-gerar."""
    from coffee_module import client
    chamadas = []
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, sap: chamadas.append(("sap", int(i), sap)) or True)
    monkeypatch.setattr(client, "desarquivar",
                        lambda i: chamadas.append(("des", int(i))) or True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": 17247854, "arquivado": True,
                   "fields": {"id_sap": 17247854}},
    )
    r = coffee_cliente.post("/api/coffee/regerar", json={"id": 355617})
    assert r.status_code == 200
    assert ("sap", 355617, config.SAP_PENDENTE) in chamadas
    assert ("des", 355617) in chamadas


def test_rota_regerar_grava_origem_avulsa(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": config.SAP_PENDENTE, "arquivado": False,
                   "fields": {"id_sap": config.SAP_PENDENTE}},
    )
    coffee_cliente.post("/api/coffee/regerar", json={"id": 355617})
    assert db.origem_atual(355617) == "avulsa"
    # re-busca com SAP real -> classificada como gerada (nao corrigida)
    assert db.upsert_nota(355617, 17247854, {"id_sap": 17247854}) == "gerada"


def test_rota_regerar_preserva_origem_existente(coffee_cliente, monkeypatch):
    from coffee_module import client, db
    monkeypatch.setattr(client, "definir_sap", lambda i, sap: True)
    monkeypatch.setattr(client, "desarquivar", lambda i: True)
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda i: {"pk": int(i), "id_sap": config.SAP_PENDENTE, "arquivado": False,
                   "fields": {"id_sap": config.SAP_PENDENTE}},
    )
    db.upsert_nota(355617, config.SAP_PENDENTE, {"id_sap": config.SAP_PENDENTE})
    db.definir_origem(355617, "verificar")
    coffee_cliente.post("/api/coffee/regerar", json={"id": 355617})
    assert db.origem_atual(355617) == "verificar"


# ---------------------------------------------------------------------------
# Malha fina — job de correção de local com 9 extra
# ---------------------------------------------------------------------------

def _nota_fake(pk, local, id_sap=None, arquivado=False):
    return {"pk": int(pk), "id_sap": id_sap, "arquivado": arquivado,
            "local_instalacao": local, "fields": {"id_sap": id_sap}}


def test_job_correcao_local_corrige_e_pula(coffee_tmp, monkeypatch):
    """Corrige quem tem 9 extra; pula ja_corrigidas e divergentes."""
    from coffee_module import client, jobs

    locais = {1: "718ET000267739",   # errado -> corrige
              2: "718ET00026773",    # já é o proposto -> ja_corrigidas
              3: "718XX99999999"}    # nem errado nem proposto -> divergente
    alterados = []
    monkeypatch.setattr(client, "buscar_nota", lambda i: _nota_fake(i, locais[int(i)]))
    monkeypatch.setattr(client, "alterar_local",
                        lambda i, l: alterados.append((int(i), l)) or True)

    itens = [{"id": 1, "local": "718ET00026773"},
             {"id": 2, "local": "718ET00026773"},
             {"id": 3, "local": "718ET00026773"}]
    j = _aguardar_job(jobs, jobs.iniciar_correcao_local(itens))

    assert j["total"] == 3 and j["feitas"] == 3
    assert alterados == [(1, "718ET00026773")]
    assert j["corrigidas"] == [1]
    assert j["ja_corrigidas"] == [2]
    assert j["divergentes"] == [{"id": 3, "local_atual": "718XX99999999"}]
    assert j["geradas"] == [] and j["erros"] == []


def test_job_correcao_local_erro_isolado_nao_derruba_lote(coffee_tmp, monkeypatch):
    from coffee_module import client, jobs

    def fake_buscar(i):
        if int(i) == 99:
            raise RuntimeError("timeout")
        return _nota_fake(i, "718ET000267739")

    alterados = []
    monkeypatch.setattr(client, "buscar_nota", fake_buscar)
    monkeypatch.setattr(client, "alterar_local",
                        lambda i, l: alterados.append(int(i)) or True)

    itens = [{"id": 1, "local": "718ET00026773"},
             {"id": 99, "local": "718ET00026773"},
             {"id": 2, "local": "718ET00026773"}]
    j = _aguardar_job(jobs, jobs.iniciar_correcao_local(itens))

    assert j["feitas"] == 3
    assert j["corrigidas"] == [1, 2]
    assert len(j["erros"]) == 1 and j["erros"][0]["pk"] == 99


def test_job_correcao_local_gerar_apos_encadeia_so_corrigidas(coffee_tmp, monkeypatch):
    """gerar_apos: SAP placeholder + desarquivar só para quem foi corrigido."""
    from coffee_module import client, config, jobs

    locais = {1: "718ET000267739", 2: "718ET00026773"}
    chamadas = []

    def fake_buscar(i):
        # Após alterar_local, a re-busca devolve o local corrigido.
        corrigido = ("alterar", int(i)) in [c[:2] for c in chamadas]
        local = "718ET00026773" if corrigido else locais[int(i)]
        return _nota_fake(i, local)

    monkeypatch.setattr(client, "buscar_nota", fake_buscar)
    monkeypatch.setattr(client, "alterar_local",
                        lambda i, l: chamadas.append(("alterar", int(i), l)) or True)
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, s: chamadas.append(("sap", int(i), s)) or True)
    monkeypatch.setattr(client, "desarquivar",
                        lambda i: chamadas.append(("desarq", int(i))) or True)

    itens = [{"id": 1, "local": "718ET00026773"},
             {"id": 2, "local": "718ET00026773"}]
    j = _aguardar_job(jobs, jobs.iniciar_correcao_local(itens, gerar_apos=True))

    assert j["corrigidas"] == [1] and j["ja_corrigidas"] == [2]
    assert j["geradas"] == [1]
    assert ("sap", 1, config.SAP_PENDENTE) in chamadas
    assert ("desarq", 1) in chamadas
    # nota 2 não entrou na geração
    assert ("sap", 2, config.SAP_PENDENTE) not in chamadas


def test_job_correcao_local_gerar_apos_ignora_sap_real(coffee_tmp, monkeypatch):
    """Corrigida mas com SAP real: não re-gera (loga geracao_ignorada_sap_real)."""
    from coffee_module import client, config, jobs

    chamadas = []
    monkeypatch.setattr(client, "buscar_nota",
                        lambda i: _nota_fake(i, "718ET000267739", id_sap=17247854))
    monkeypatch.setattr(client, "alterar_local", lambda i, l: True)
    monkeypatch.setattr(client, "definir_sap",
                        lambda i, s: chamadas.append(("sap", int(i))) or True)
    monkeypatch.setattr(client, "desarquivar",
                        lambda i: chamadas.append(("desarq", int(i))) or True)

    itens = [{"id": 1, "local": "718ET00026773"}]
    j = _aguardar_job(jobs, jobs.iniciar_correcao_local(itens, gerar_apos=True))

    assert j["corrigidas"] == [1]
    assert j["geradas"] == []
    assert chamadas == []


def test_job_correcao_local_gerar_apos_falha_registra_corrigida_e_erro(coffee_tmp, monkeypatch):
    """Se _gerar_apos_correcao falha, a nota fica em corrigidas E erros, com contexto."""
    from coffee_module import client, jobs

    monkeypatch.setattr(client, "buscar_nota",
                        lambda i: _nota_fake(i, "718ET000267739"))
    monkeypatch.setattr(client, "alterar_local", lambda i, l: True)

    def fake_definir_sap(i, s):
        raise RuntimeError("falha ao definir sap")

    monkeypatch.setattr(client, "definir_sap", fake_definir_sap)

    itens = [{"id": 1, "local": "718ET00026773"}]
    j = _aguardar_job(jobs, jobs.iniciar_correcao_local(itens, gerar_apos=True))

    assert j["corrigidas"] == [1]
    assert j["geradas"] == []
    assert len(j["erros"]) == 1
    assert j["erros"][0]["pk"] == 1
    assert "gera" in j["erros"][0]["msg"].lower()


# ---------------------------------------------------------------------------
# Task 2 — Rota /corrigir-local-lote
# ---------------------------------------------------------------------------


def test_rota_corrigir_local_lote_validacoes(coffee_cliente):
    r = coffee_cliente.post("/api/coffee/corrigir-local-lote",
                            json={"itens": []})
    assert r.status_code == 400

    r = coffee_cliente.post("/api/coffee/corrigir-local-lote",
                            json={"itens": [{"id": 1, "local": "curto"}]})
    assert r.status_code == 400
    assert "13" in r.json()["detail"]


# ---------------------------------------------------------------------------
# Tarefa B — notas COFFEE por usuario (visibilidade estrita do dono)
# ---------------------------------------------------------------------------


def test_upsert_grava_dono_no_primeiro_upsert_e_preserva(coffee_tmp):
    from coffee_module import db
    db.definir_usuario("alice")
    db.upsert_nota(1, 10000000, {"id_sap": 10000000})
    assert db.listar_notas()[0]["usuario"] == "alice"
    # re-busca por outro usuario nao rouba a posse
    db.definir_usuario("bob")
    db.upsert_nota(1, 17247854, {"id_sap": 17247854})
    assert db.listar_notas()[0]["usuario"] == "alice"
    db.definir_usuario(None)


def test_listar_notas_filtra_estrito_por_usuario_e_inclui_null(coffee_tmp):
    from coffee_module import db
    db.definir_usuario("alice")
    db.upsert_nota(1, 10000000, {"id_sap": 10000000})
    db.definir_usuario("bob")
    db.upsert_nota(2, 10000000, {"id_sap": 10000000})
    db.definir_usuario(None)
    db.upsert_nota(3, 10000000, {"id_sap": 10000000})
    # simula nota legada (pre-migracao), sem dono gravado
    conn = db.get_db_connection()
    conn.execute("UPDATE notas_coffee SET usuario = NULL WHERE pk = 3")
    conn.commit()
    conn.close()

    vistas_por_alice = {n["pk"] for n in db.listar_notas(usuario="alice")}
    assert vistas_por_alice == {1, 3}

    todas = {n["pk"] for n in db.listar_notas()}
    assert todas == {1, 2, 3}


def test_rota_notas_filtra_por_x_user(coffee_cliente, monkeypatch):
    from coffee_module import client, db, jobs
    monkeypatch.setattr(
        client, "buscar_nota",
        lambda id: {"pk": int(id), "id_sap": 17247854, "arquivado": False,
                    "local_instalacao": None, "fields": {"id_sap": 17247854}},
    )
    r = coffee_cliente.post("/api/coffee/buscar", json={"ids": ["1"]},
                            headers={"X-User": "alice"})
    _aguardar_job(jobs, r.json()["job_id"])
    assert db.listar_notas()[0]["usuario"] == "alice"

    r2 = coffee_cliente.post("/api/coffee/buscar", json={"ids": ["2"]},
                             headers={"X-User": "bob"})
    _aguardar_job(jobs, r2.json()["job_id"])

    vistas_por_alice = coffee_cliente.get(
        "/api/coffee/notas", headers={"X-User": "alice"}).json()["registros"]
    assert {n["pk"] for n in vistas_por_alice} == {1}


def test_rota_corrigir_local_lote_dispara_job(coffee_cliente, monkeypatch):
    from coffee_module import client, db, jobs

    monkeypatch.setattr(client, "buscar_nota",
                        lambda i: _nota_fake(i, "718ET000267739"))
    monkeypatch.setattr(client, "alterar_local", lambda i, l: True)

    r = coffee_cliente.post("/api/coffee/corrigir-local-lote",
                            json={"itens": [{"id": 1, "local": "718ET00026773"}],
                                  "gerar_apos": False})
    assert r.status_code == 200
    job_id = r.json()["job_id"]
    j = _aguardar_job(jobs, job_id)
    assert j["corrigidas"] == [1]

    logs = db.listar_logs(tipo="acao_usuario")
    assert any(l["acao"] == "correcao_local_lote" for l in logs)


def test_listar_notas_separa_geradas_corrigidas_e_concluidas(coffee_tmp):
    from coffee_module import db

    db.upsert_nota(101, 17100101, {"id_sap": 17100101})
    db.upsert_nota(202, 10000000, {"id_sap": 10000000})
    db.definir_origem(202, "verificar")
    db.upsert_nota(202, 17100202, {"id_sap": 17100202})

    assert [n["pk"] for n in db.listar_notas("gerada")] == [101]
    assert [n["pk"] for n in db.listar_notas("corrigida")] == [202]
    assert {n["pk"] for n in db.listar_notas("concluida")} == {101, 202}
