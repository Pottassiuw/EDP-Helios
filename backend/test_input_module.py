"""Testes do módulo Input (backend)."""
import datetime
import io
import os
import tempfile
import threading
import time
from pathlib import Path

# Blindagem global: impede que a execução de testes afete o banco de dados real
_tmp_test_dir = tempfile.mkdtemp(prefix="edp_input_test_")
os.environ.setdefault("INPUT_DATA_DIR", _tmp_test_dir)

from input_module import config


def test_config_dicionarios_completos():
    assert config.STATUS_MAP[99] == "99 Encerrado"
    assert config.STATUS_MAP[0] == "00 Pendente"
    assert config.INV_STATUS_MAP["99 Encerrado"] == 99
    assert config.DE_PARA_REGIONAL["045"] == "Guarulhos"
    assert config.DE_PARA_CIDADES["130"] == "Mogi das Cruzes - SP"
    assert config.DE_PARA_CJ_ANEEL["POA"] == "POA"
    assert config.MAP_FILTROS["Status"] == "Status_Nota"
    assert config.MAP_ORDEM_EXECUTADA["JAND INVE"] == "SIM"
    assert config.MAP_REGIONAL_CSD["POA"] == "Poa/Suzano"
    assert len(config.BASES_REDE) == 7
    assert len(config.BASES_APOIO) == 4
    assert "Emergente" in config.PRIORIDADES
    assert config.NOMES_AMIGAVEIS["Numero_Nota"] == "Nº Nota (ID)"
    assert "Numero_Nota" in config.COLUNAS_PAINEL


def test_extracoes_sap_compartilham_raiz_arquivos_sap():
    arquivos = {
        config.CAMINHO_BASE_IW28: "Gerada_base_IW28.XLSX",
        config.CAMINHO_CUSTO_ORD_IW38: "Gerada_custo_ord_IW38.XLSX",
        config.CAMINHO_BASE_IW66: "Gerada_medidas_IW66.XLSX",
    }

    for caminho, arquivo in arquivos.items():
        assert caminho == config.REDE_ARQUIVOS_SAP + f"\\{arquivo}"


def test_data_dir_respeita_env(monkeypatch, tmp_path):
    monkeypatch.setenv("INPUT_DATA_DIR", str(tmp_path))
    assert config.data_dir() == tmp_path


import sqlite3

import pytest


@pytest.fixture
def banco_temporario(monkeypatch, tmp_path):
    """Aponta o módulo para um diretório de dados temporário e inicializa o banco."""
    monkeypatch.setenv("INPUT_DATA_DIR", str(tmp_path))
    from input_module import db
    db.inicializar_banco()
    return tmp_path


def test_inicializar_banco_cria_tabelas(banco_temporario):
    from input_module import db
    conn = db.get_db_connection()
    tabelas = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    conn.close()
    assert {"notas", "log_alteracoes", "log_arquivos", "bloqueios"} <= tabelas


def test_inicializar_banco_cria_indices(banco_temporario):
    from input_module import db
    conn = db.get_db_connection()
    indices = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='index'").fetchall()}
    conn.close()
    assert {"idx_log_alteracoes_nota", "idx_log_alteracoes_data",
            "idx_log_arquivos_data"} <= indices


def test_migracao_copia_banco_da_rede(monkeypatch, tmp_path):
    from input_module import config, db
    # Simula o banco "da rede" como um sqlite real noutro tmp
    origem = tmp_path / "rede.db"
    conn = sqlite3.connect(origem)
    conn.execute("CREATE TABLE notas (Numero_Nota INTEGER PRIMARY KEY)")
    conn.execute("INSERT INTO notas VALUES (123)")
    conn.commit(); conn.close()
    monkeypatch.setenv("INPUT_DATA_DIR", str(tmp_path / "dados"))
    monkeypatch.setattr(config, "REDE_DB_ORIGEM", str(origem))
    resultado = db.migrar_da_rede_se_preciso()
    assert resultado == "migrado"
    conn = db.get_db_connection()
    assert conn.execute("SELECT COUNT(*) FROM notas").fetchone()[0] == 1
    conn.close()
    # Segunda chamada: banco já existe, não migra de novo
    assert db.migrar_da_rede_se_preciso() == "ja-existe"


def test_migracao_sem_rede_retorna_indisponivel(monkeypatch, tmp_path):
    from input_module import config, db
    monkeypatch.setenv("INPUT_DATA_DIR", str(tmp_path))
    monkeypatch.setattr(config, "REDE_DB_ORIGEM", str(tmp_path / "nao_existe.db"))
    assert db.migrar_da_rede_se_preciso() == "rede-indisponivel"


# ── Tarefa 3: CRUD, logs, undo, backups, responsáveis e edição com diff ──
import pandas as pd


def _nota(numero=1000, **extras):
    base = {
        "ID_Cronologia": 1, "Numero_Nota": numero, "Status_Obra": "-",
        "Conjunto": "POA", "Circuito": "POA 123", "Local_Instalacao": "045 RL TESTE",
        "Regional": "Guarulhos", "Planejado_DDPM": 2.0,
        "Mes_Execucao_Planejado": "jun-2026", "Data_Envio_Projeto": "01/06/2026",
        "Status_Nota": "10 Em planejamento", "Prioridade_Nota": "Programável",
        "Observacao": "", "Check": "-", "Status_Anterior": "-",
        "Centro_Responsavel": "-",
    }
    base.update(extras)
    return base


def test_status10_resumo_serializa_ausentes_com_json_estrito(monkeypatch):
    import json
    from input_module import status10_service

    monkeypatch.setattr(
        status10_service.db,
        "carregar_dados",
        lambda: pd.DataFrame([
            _nota(1000, Planejado_DDPM=float("nan"), Modular=float("nan"),
                  Regional=None, Conjunto=None, Observacao=None,
                  Data_Envio_Projeto=float("nan")),
        ]),
    )

    resumo = status10_service.obter_resumo_status10()
    payload = json.dumps(resumo, allow_nan=False)

    assert json.loads(payload)["registros"][0]["Observacao"] is None


def test_status10_rotulo_planejado_usa_unidade_de_quantidade():
    from input_module.status10_service import rotulos_resumo_status10

    assert rotulos_resumo_status10()["Total_Planejado"] == "Total Planejado (un)"


def test_upsert_e_carregar(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(1000), _nota(1001, Conjunto="SUZANO")]))
    df = db.carregar_dados()
    assert len(df) == 2
    linha = df[df["Numero_Nota"] == 1000].iloc[0]
    assert linha["Status_Nota"] == "10 Em planejamento"
    assert linha["Cidade"] == "Guarulhos"
    db.salvar_em_massa(pd.DataFrame([_nota(1000, Observacao="editada")]))
    df = db.carregar_dados()
    assert len(df) == 2
    assert df[df["Numero_Nota"] == 1000].iloc[0]["Observacao"] == "editada"


def test_aplicar_edicoes_gera_diff_log_e_status_anterior(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    resultado = db.aplicar_edicoes(
        [{"Numero_Nota": 2000, "Status_Nota": "99 Encerrado", "Observacao": "feita"}],
        usuario="tester")
    assert resultado["alteradas"] == 1
    assert resultado["campos"] == 2
    df = db.carregar_dados()
    linha = df[df["Numero_Nota"] == 2000].iloc[0]
    assert linha["Status_Nota"] == "99 Encerrado"
    assert str(linha["Status_Anterior"]).startswith("10")  # status antigo preservado (numérico)
    logs = db.carregar_logs()
    assert set(logs["Campo_Alterado"]) == {"Status_Nota", "Observacao"}
    assert logs.iloc[0]["Usuario"] == "tester"
    resultado = db.aplicar_edicoes([{"Numero_Nota": 2000, "Observacao": "feita"}], usuario="tester")
    assert resultado["alteradas"] == 0


def test_aplicar_edicoes_aceita_schema_legado_sem_status_anterior(
    banco_temporario
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(2001)]))
    conn = db.get_db_connection()
    try:
        conn.execute("ALTER TABLE notas DROP COLUMN Status_Anterior")
        conn.commit()
    finally:
        conn.close()

    resultado = db.aplicar_edicoes(
        [{"Numero_Nota": 2001, "Status_Nota": "99 Encerrado"}],
        usuario="tester",
    )

    assert resultado == {"alteradas": 1, "campos": 1, "bloqueadas": []}
    nota = db.carregar_dados().set_index("Numero_Nota").loc[2001]
    assert nota["Status_Nota"] == "99 Encerrado"
    assert list(db.carregar_logs()["Campo_Alterado"]) == ["Status_Nota"]


def test_aplicar_edicoes_consolida_linhas_duplicadas_da_mesma_nota(
    banco_temporario
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(2002)]))

    resultado = db.aplicar_edicoes(
        [
            {"Numero_Nota": 2002, "Observacao": "primeira"},
            {"Numero_Nota": 2002, "Observacao": "final", "Check": "feito"},
        ],
        usuario="tester",
    )

    assert resultado == {"alteradas": 1, "campos": 2, "bloqueadas": []}
    nota = db.carregar_dados().set_index("Numero_Nota").loc[2002]
    assert nota["Observacao"] == "final"
    assert nota["Check"] == "feito"
    logs = db.carregar_logs()
    assert set(logs["Campo_Alterado"]) == {"Observacao", "Check"}


def test_aplicar_edicoes_nota_inexistente_da_erro(banco_temporario):
    from input_module import db
    with pytest.raises(ValueError):
        db.aplicar_edicoes([{"Numero_Nota": 999999, "Observacao": "x"}], usuario="t")


def test_reverter_ultima_alteracao(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(3000)]))
    db.aplicar_edicoes([{"Numero_Nota": 3000, "Status_Nota": "99 Encerrado"}], usuario="t")
    ok, _msg = db.reverter_ultima_alteracao("t")
    assert ok
    df = db.carregar_dados()
    assert df[df["Numero_Nota"] == 3000].iloc[0]["Status_Nota"] == "10 Em planejamento"
    ok, _msg = db.reverter_ultima_alteracao("t")
    assert not ok


def test_reverter_nao_desfaz_alteracao_de_outro_usuario(banco_temporario):
    """O undo é por usuário: com o banco compartilhado, desfazer o próprio
    trabalho não pode reverter o da colega que salvou depois."""
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(3100)]))
    db.aplicar_edicoes([{"Numero_Nota": 3100, "Observacao": "minha"}], usuario="eu")
    db.aplicar_edicoes([{"Numero_Nota": 3100, "Observacao": "dela"}], usuario="outra")

    ok, _msg = db.reverter_ultima_alteracao("eu")
    assert ok
    df = db.carregar_dados()
    # A edição da outra pessoa (mais recente) permanece intocada.
    assert df[df["Numero_Nota"] == 3100].iloc[0]["Observacao"] == "dela"


def test_reverter_protege_comparacao_e_update_na_mesma_transacao(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(3101)]))
    db.aplicar_edicoes(
        [{"Numero_Nota": 3101, "Observacao": "minha"}],
        usuario="eu",
    )
    comparar_original = db._mesmo_valor
    tentativa_concorrente = {}

    def comparar_com_escrita_concorrente(gravado, esperado):
        conexao = sqlite3.connect(db.obter_caminho_banco(), timeout=0)
        try:
            conexao.execute(
                "UPDATE notas SET Observacao = ? WHERE Numero_Nota = ?",
                ("concorrente", 3101),
            )
            conexao.commit()
            tentativa_concorrente["gravou"] = True
        except sqlite3.OperationalError as erro:
            assert "locked" in str(erro).lower()
            tentativa_concorrente["gravou"] = False
        finally:
            conexao.close()
        return comparar_original(gravado, esperado)

    monkeypatch.setattr(db, "_mesmo_valor", comparar_com_escrita_concorrente)

    ok, _mensagem = db.reverter_ultima_alteracao("eu")

    assert ok
    assert tentativa_concorrente == {"gravou": False}
    nota = db.carregar_dados().set_index("Numero_Nota").loc[3101]
    assert nota["Observacao"] == ""


def test_deletar_notas(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(4000), _nota(4001)]))
    assert db.deletar_notas([4000]) == 1
    assert list(db.carregar_dados()["Numero_Nota"]) == [4001]


def test_carregar_dados_qualidade(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([
        _nota(4300, Prioridade_Nota="Programavel", Mes_Execucao_Planejado="jun-2026"),
        _nota(4301, Prioridade_Nota="Prioritario", Mes_Execucao_Planejado="2026-12-01 00:00:00"),
    ]))
    df = db.carregar_dados()
    assert len(df) == 2
    pri = dict(zip(df["Numero_Nota"], df["Prioridade_Nota"]))
    assert pri[4300] == "Programável"
    assert pri[4301] == "Prioritário"


def test_carregar_logs_fallback_em_erro(banco_temporario, monkeypatch):
    from input_module import db

    def boom(*args, **kwargs):
        raise RuntimeError("falha simulada de leitura")

    monkeypatch.setattr(db.pd, "read_sql", boom)
    logs = db.carregar_logs()
    assert logs.empty
    assert "Campo_Alterado" in logs.columns
    arquivos = db.carregar_log_arquivos()
    assert arquivos.empty
    assert "Nome_Arquivo" in arquivos.columns


def test_deletar_notas_gera_log(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(4100)]))
    assert db.deletar_notas([4100], usuario="tester") == 1
    logs = db.carregar_logs()
    linha = logs[logs["Numero_Nota"] == 4100].iloc[0]
    assert linha["Campo_Alterado"] == "EXCLUSÃO DE NOTA"
    assert linha["Usuario"] == "tester"


# ── Fase 2: bloqueios (edição concorrente no banco compartilhado) ────────
def test_travar_nota_bloqueia_outro_usuario(banco_temporario):
    from input_module import db
    assert db.travar_nota(4200, "ana") == {"ok": True}
    resultado = db.travar_nota(4200, "bob")
    assert resultado["ok"] is False
    assert resultado["usuario"] == "ana"
    assert "desde" in resultado


def test_travar_nota_concorrente_tem_apenas_um_vencedor(
    banco_temporario, monkeypatch
):
    from input_module import db

    get_connection_original = db.get_db_connection
    conexoes_prontas = threading.Barrier(2)
    resultados = []
    erros = []

    def get_connection_sincronizada():
        conexao = get_connection_original()
        conexoes_prontas.wait(timeout=5)
        return conexao

    def travar(usuario):
        try:
            resultados.append((usuario, db.travar_nota(4210, usuario)))
        except Exception as erro:
            erros.append(erro)

    monkeypatch.setattr(db, "get_db_connection", get_connection_sincronizada)
    threads = [
        threading.Thread(target=travar, args=("ana",)),
        threading.Thread(target=travar, args=("bob",)),
    ]

    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)

    monkeypatch.setattr(db, "get_db_connection", get_connection_original)
    assert all(not thread.is_alive() for thread in threads)
    assert erros == []
    assert sum(resultado["ok"] for _, resultado in resultados) == 1
    dono = db.obter_bloqueios([4210])[4210]["usuario"]
    assert next(usuario for usuario, resultado in resultados if resultado["ok"]) == dono


def test_travar_nota_mesmo_usuario_renova(banco_temporario):
    from input_module import db
    assert db.travar_nota(4201, "ana")["ok"] is True
    # Segunda chamada da MESMA pessoa não é bloqueio — é renovação do TTL.
    assert db.travar_nota(4201, "ana")["ok"] is True


def test_destravar_libera_para_outro_usuario(banco_temporario):
    from input_module import db
    db.travar_nota(4202, "ana")
    assert db.destravar_notas([4202], "ana") == 1
    assert db.travar_nota(4202, "bob")["ok"] is True


def test_destravar_nao_derruba_lock_de_outro(banco_temporario):
    """Um release tardio de quem perdeu a corrida não pode apagar o lock de
    quem já assumiu a nota no meio tempo."""
    from input_module import db
    db.travar_nota(4203, "ana")
    assert db.destravar_notas([4203], "bob") == 0  # bob nunca foi o dono
    assert db.obter_bloqueios([4203])[4203]["usuario"] == "ana"


def test_bloqueio_expira_por_ttl(banco_temporario, monkeypatch):
    from input_module import db
    import datetime
    db.travar_nota(4204, "ana")
    # Simula um lock antigo sem esperar o TTL de verdade.
    expirado = datetime.datetime.now() - datetime.timedelta(minutes=db.BLOQUEIO_TTL_MINUTOS + 1)
    conn = db.get_db_connection()
    conn.execute("UPDATE bloqueios SET Data_Hora = ? WHERE Numero_Nota = ?", (expirado, 4204))
    conn.commit()
    conn.close()
    assert db.obter_bloqueios([4204]) == {}
    assert db.travar_nota(4204, "bob")["ok"] is True


def test_aplicar_edicoes_pula_nota_travada_por_outro(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(4205)]))
    db.travar_nota(4205, "outra")
    resultado = db.aplicar_edicoes(
        [{"Numero_Nota": 4205, "Observacao": "tentativa"}], usuario="eu")
    assert resultado["alteradas"] == 0
    assert resultado["bloqueadas"] == [4205]
    df = db.carregar_dados()
    assert df[df["Numero_Nota"] == 4205].iloc[0]["Observacao"] == ""


def test_aplicar_edicoes_permite_dono_do_bloqueio(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(4206)]))
    db.travar_nota(4206, "eu")
    resultado = db.aplicar_edicoes(
        [{"Numero_Nota": 4206, "Observacao": "minha edicao"}], usuario="eu")
    assert resultado["alteradas"] == 1
    assert resultado["bloqueadas"] == []


def test_aplicar_edicoes_concorrentes_preserva_campos_diferentes(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(4209)]))
    carregar_original = db.carregar_dados
    get_connection_original = db.get_db_connection
    conexoes_prontas = threading.Barrier(2)
    erros = []

    def get_connection_sincronizada():
        conexao = get_connection_original()
        conexoes_prontas.wait(timeout=5)
        return conexao

    def editar(linha):
        try:
            db.aplicar_edicoes([linha], usuario="mesmo-usuario")
        except Exception as erro:
            erros.append(erro)

    monkeypatch.setattr(db, "get_db_connection", get_connection_sincronizada)
    threads = [
        threading.Thread(
            target=editar,
            args=({"Numero_Nota": 4209, "Observacao": "observacao concorrente"},),
        ),
        threading.Thread(
            target=editar,
            args=({"Numero_Nota": 4209, "Check": "check concorrente"},),
        ),
    ]

    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)

    monkeypatch.setattr(db, "get_db_connection", get_connection_original)
    assert all(not thread.is_alive() for thread in threads)
    assert erros == []
    nota = carregar_original().set_index("Numero_Nota").loc[4209]
    assert nota["Observacao"] == "observacao concorrente"
    assert nota["Check"] == "check concorrente"


def test_aplicar_edicoes_reverte_log_quando_update_falha(banco_temporario):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(4211)]))
    conn = db.get_db_connection()
    try:
        conn.execute(
            """CREATE TRIGGER falhar_update_observacao
               BEFORE UPDATE OF Observacao ON notas
               WHEN NEW.Observacao = 'falha-forcada'
               BEGIN
                 SELECT RAISE(ABORT, 'falha simulada no update');
               END"""
        )
        conn.commit()
    finally:
        conn.close()

    with pytest.raises(sqlite3.IntegrityError, match="falha simulada"):
        db.aplicar_edicoes(
            [{"Numero_Nota": 4211, "Observacao": "falha-forcada"}],
            usuario="teste",
        )

    nota = db.carregar_dados().set_index("Numero_Nota").loc[4211]
    logs = db.carregar_logs()
    assert nota["Observacao"] == ""
    assert logs[logs["Numero_Nota"] == 4211].empty


def test_deletar_notas_pula_travada_por_outro(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(4207), _nota(4208)]))
    db.travar_nota(4207, "outra")
    assert db.deletar_notas([4207, 4208], usuario="eu") == 1
    numeros = set(db.carregar_dados()["Numero_Nota"])
    assert 4207 in numeros   # travada: sobreviveu
    assert 4208 not in numeros  # livre: excluída


def test_deletar_notas_revalida_lock_adquirido_antes_do_delete(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(4212)]))
    obter_original = db.obter_bloqueios
    consulta_concluida = threading.Event()
    continuar_delete = threading.Event()
    resultado_delete = {}
    erros = []

    def obter_pausado(numeros):
        bloqueios = obter_original(numeros)
        consulta_concluida.set()
        assert continuar_delete.wait(timeout=5)
        return bloqueios

    def deletar():
        try:
            resultado_delete["quantidade"] = db.deletar_notas(
                [4212], usuario="ana"
            )
        except Exception as erro:
            erros.append(erro)

    monkeypatch.setattr(db, "obter_bloqueios", obter_pausado)
    thread = threading.Thread(target=deletar)
    thread.start()
    assert consulta_concluida.wait(timeout=5)
    resultado_lock = db.travar_nota(4212, "bob")
    continuar_delete.set()
    thread.join(timeout=10)
    monkeypatch.setattr(db, "obter_bloqueios", obter_original)

    assert not thread.is_alive()
    assert erros == []
    assert resultado_lock == {"ok": True}
    assert resultado_delete["quantidade"] == 0
    assert 4212 in set(db.carregar_dados()["Numero_Nota"])
    assert db.obter_bloqueios([4212])[4212]["usuario"] == "bob"


def test_deletar_nota_concorrente_gera_um_unico_log(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(4213)]))
    obter_original = db.obter_bloqueios
    consultas_concluidas = threading.Barrier(2)
    resultados = []
    erros = []

    def obter_sincronizado(numeros):
        bloqueios = obter_original(numeros)
        consultas_concluidas.wait(timeout=5)
        return bloqueios

    def deletar(usuario):
        try:
            resultados.append(db.deletar_notas([4213], usuario=usuario))
        except Exception as erro:
            erros.append(erro)

    monkeypatch.setattr(db, "obter_bloqueios", obter_sincronizado)
    threads = [
        threading.Thread(target=deletar, args=("ana",)),
        threading.Thread(target=deletar, args=("bob",)),
    ]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)
    monkeypatch.setattr(db, "obter_bloqueios", obter_original)

    assert all(not thread.is_alive() for thread in threads)
    assert erros == []
    assert sum(resultados) == 1
    logs = db.carregar_logs()
    logs_exclusao = logs[
        (logs["Numero_Nota"] == 4213)
        & (logs["Campo_Alterado"] == "EXCLUSÃO DE NOTA")
    ]
    assert len(logs_exclusao) == 1


def test_deletar_notas_ignora_numeros_duplicados(banco_temporario):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(4214)]))

    assert db.deletar_notas([4214, 4214], usuario="ana") == 1
    logs = db.carregar_logs()
    logs_exclusao = logs[
        (logs["Numero_Nota"] == 4214)
        & (logs["Campo_Alterado"] == "EXCLUSÃO DE NOTA")
    ]
    assert len(logs_exclusao) == 1


def test_backup_rotativo(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(5000)]))
    pasta = config_backups_dir()
    for f in pasta.glob("notas_departamento_*.db"):
        try:
            os.remove(f)
        except Exception:
            pass
    db.realizar_backup(limite=20, intervalo_horas=0)
    arquivos = list(pasta.glob("notas_departamento_*.db"))
    assert len(arquivos) == 1
    db.realizar_backup(limite=20, intervalo_horas=2)
    assert len(list(pasta.glob("notas_departamento_*.db"))) == 1


def test_backup_rotativo_inclui_commit_no_wal_de_conexao_concorrente(banco_temporario):
    """Protege contra cópia do .db sem o conteúdo confirmado no arquivo -wal."""
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(5001)]))
    pasta = config_backups_dir()
    for arquivo in pasta.glob("notas_departamento_*.db"):
        arquivo.unlink()
    caminho_origem = db.obter_caminho_banco()
    conexao_gravadora = sqlite3.connect(caminho_origem)
    try:
        journal_mode = conexao_gravadora.execute(
            "PRAGMA journal_mode = WAL"
        ).fetchone()[0].lower()
        assert journal_mode == "wal"
        conexao_gravadora.execute(
            "UPDATE notas SET Observacao = ? WHERE Numero_Nota = ?",
            ("gravado por conexão concorrente", 5001),
        )
        conexao_gravadora.commit()

        db.realizar_backup(limite=20, intervalo_horas=0)

        caminho_backup = next(pasta.glob("notas_departamento_*.db"))
        conexao_backup = sqlite3.connect(caminho_backup)
        try:
            observacao = conexao_backup.execute(
                "SELECT Observacao FROM notas WHERE Numero_Nota = ?", (5001,)
            ).fetchone()[0]
        finally:
            conexao_backup.close()
    finally:
        conexao_gravadora.close()

    assert observacao == "gravado por conexão concorrente"


def test_backup_rotativo_remove_arquivo_parcial_quando_snapshot_falha(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(5002)]))
    pasta = config_backups_dir()
    for arquivo in pasta.glob("notas_departamento_*.db"):
        arquivo.unlink()

    class ConexaoOrigemComFalha:
        def backup(self, conexao_destino):
            raise sqlite3.OperationalError("falha simulada no snapshot")

        def close(self):
            pass

    monkeypatch.setattr(
        db, "_conectar_origem_backup", lambda caminho: ConexaoOrigemComFalha()
    )

    db.realizar_backup(limite=20, intervalo_horas=0)

    assert list(pasta.glob("notas_departamento_*.db")) == []


def test_backups_no_mesmo_segundo_recebem_nomes_unicos(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(5003)]))
    pasta = config_backups_dir()
    for arquivo in pasta.glob("notas_departamento_*.db"):
        arquivo.unlink()

    class DataHoraFixa(datetime.datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 8, 9, 18, 15, 30, tzinfo=tz)

    monkeypatch.setattr(db.datetime, "datetime", DataHoraFixa)

    db.realizar_backup(limite=20, intervalo_horas=0)
    db.realizar_backup(limite=20, intervalo_horas=0)

    assert len(list(pasta.glob("notas_departamento_*.db"))) == 2


def test_backup_falho_nao_remove_snapshot_valido_do_mesmo_instante(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(5004)]))
    pasta = config_backups_dir()
    for arquivo in pasta.glob("notas_departamento_*.db"):
        arquivo.unlink()

    class DataHoraFixa(datetime.datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 8, 9, 18, 15, 30, tzinfo=tz)

    monkeypatch.setattr(db.datetime, "datetime", DataHoraFixa)
    db.realizar_backup(limite=20, intervalo_horas=0)
    snapshot_valido = next(pasta.glob("notas_departamento_*.db"))

    class ConexaoOrigemComFalha:
        def backup(self, conexao_destino):
            raise sqlite3.OperationalError("falha concorrente simulada")

        def close(self):
            pass

    monkeypatch.setattr(
        db, "_conectar_origem_backup", lambda caminho: ConexaoOrigemComFalha()
    )

    db.realizar_backup(limite=20, intervalo_horas=0)

    assert snapshot_valido.exists()
    assert list(pasta.glob("notas_departamento_*.db")) == [snapshot_valido]


def test_backup_nao_recria_origem_removida_antes_da_abertura(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(5005)]))
    pasta = config_backups_dir()
    for arquivo in pasta.glob("notas_departamento_*.db"):
        arquivo.unlink()
    caminho_origem = Path(db.obter_caminho_banco())
    caminho_origem.unlink()
    existe_original = db.os.path.exists

    def existe(caminho):
        if str(caminho) == str(caminho_origem):
            return True
        return existe_original(caminho)

    monkeypatch.setattr(db.os.path, "exists", existe)

    db.realizar_backup(limite=20, intervalo_horas=0)

    assert not existe_original(str(caminho_origem))
    assert list(pasta.glob("notas_departamento_*.db")) == []


def test_conexao_backup_converte_unc_para_uri_sem_authority(monkeypatch):
    from input_module import db

    caminho_unc = "\\\\servidor\\share\\notas departamento.db"
    conectar_original = db.sqlite3.connect
    chamada = {}

    def capturar(database, **kwargs):
        chamada["uri"] = database
        chamada["kwargs"] = kwargs
        return object()

    monkeypatch.setattr(db.sqlite3, "connect", capturar)

    db._conectar_origem_backup(caminho_unc)

    assert chamada == {
        "uri": "file:////servidor/share/notas%20departamento.db?mode=ro",
        "kwargs": {"uri": True, "timeout": 30},
    }
    uri_memoria = chamada["uri"].replace("mode=ro", "mode=memory")
    conexao = conectar_original(uri_memoria, uri=True)
    conexao.close()


def test_backups_concorrentes_respeitam_limite_e_nao_deixam_parcial(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(5006)]))
    pasta = config_backups_dir()
    for arquivo in pasta.glob("notas_departamento_*"):
        arquivo.unlink()
    lock_medicao = threading.Lock()
    ciclos_ativos = 0
    maximo_ciclos_ativos = 0
    executar_original = db._realizar_backup_serializado

    def executar(limite, intervalo_horas):
        nonlocal ciclos_ativos, maximo_ciclos_ativos
        with lock_medicao:
            ciclos_ativos += 1
            maximo_ciclos_ativos = max(maximo_ciclos_ativos, ciclos_ativos)
        time.sleep(0.05)
        try:
            executar_original(limite, intervalo_horas)
        finally:
            with lock_medicao:
                ciclos_ativos -= 1

    monkeypatch.setattr(db, "_realizar_backup_serializado", executar)
    threads = [
        threading.Thread(
            target=db.realizar_backup,
            kwargs={"limite": 1, "intervalo_horas": 2},
        )
        for _ in range(2)
    ]

    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)

    assert all(not thread.is_alive() for thread in threads)
    assert maximo_ciclos_ativos == 1
    assert len(list(pasta.glob("notas_departamento_*.db"))) == 1
    assert list(pasta.glob("*.partial")) == []


def test_backup_so_publica_db_depois_do_snapshot_concluido(
    banco_temporario, monkeypatch
):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(5007)]))
    pasta = config_backups_dir()
    for arquivo in pasta.glob("notas_departamento_*"):
        arquivo.unlink()
    iniciou_snapshot = threading.Event()
    liberar_snapshot = threading.Event()
    conectar_original = db._conectar_origem_backup

    class ConexaoControlada:
        def __init__(self, conexao):
            self._conexao = conexao

        def backup(self, conexao_destino):
            iniciou_snapshot.set()
            assert liberar_snapshot.wait(timeout=5)
            self._conexao.backup(conexao_destino)

        def close(self):
            self._conexao.close()

    def conectar(caminho):
        return ConexaoControlada(conectar_original(caminho))

    monkeypatch.setattr(db, "_conectar_origem_backup", conectar)
    thread = threading.Thread(
        target=db.realizar_backup,
        kwargs={"limite": 20, "intervalo_horas": 0},
    )
    thread.start()
    try:
        assert iniciou_snapshot.wait(timeout=5)
        assert list(pasta.glob("notas_departamento_*.db")) == []
        assert len(list(pasta.glob("*.partial"))) == 1
    finally:
        liberar_snapshot.set()
        thread.join(timeout=10)

    assert not thread.is_alive()
    assert len(list(pasta.glob("notas_departamento_*.db"))) == 1
    assert list(pasta.glob("*.partial")) == []


def config_backups_dir():
    from input_module import config
    return config.data_dir() / "backups"


def test_responsaveis_roundtrip(banco_temporario):
    from input_module import db
    padrao = db.carregar_responsaveis()
    assert padrao["Mogi das Cruzes"] == "Fabricio"
    db.salvar_responsaveis({"Mogi das Cruzes": "Maria"})
    assert db.carregar_responsaveis() == {"Mogi das Cruzes": "Maria"}


# ── Task 2 (integração Coffee↔Input): service.py — caminho canônico ──────
def test_service_criar_notas(banco_temporario):
    from input_module import db, service
    nota = service.NovaNota(
        Numero_Nota=555001, Status_Nota="00 Pendente",
        Prioridade_Nota="Programável", Local_Instalacao="045RL00000001",
    )
    assert service.criar_notas([nota], usuario="teste") == 1
    df = db.carregar_dados()
    linha = df[df["Numero_Nota"] == 555001].iloc[0]
    assert linha["Regional"] == "Guarulhos"          # derivada de Local_Instalacao[:3]
    assert linha["ID_Cronologia"] == 1

    logs = db.carregar_logs()
    assert (logs["Numero_Nota"] == 555001).any()
    log_criacao = logs[logs["Numero_Nota"] == 555001].iloc[0]
    assert log_criacao["Usuario"] == "teste"
    assert log_criacao["Campo_Alterado"] == "CRIAÇÃO DE NOTA"

    with pytest.raises(service.NotasDuplicadasErro):
        service.criar_notas([nota], usuario="teste")


def test_service_criar_notas_duplicata_no_lote(banco_temporario):
    from input_module import service
    n = service.NovaNota(Numero_Nota=7, Status_Nota="00 Pendente", Prioridade_Nota="Programável")
    with pytest.raises(service.NotasDuplicadasErro):
        service.criar_notas([n, n], usuario="teste")


def test_service_criar_nota_filha_diretamente(banco_temporario):
    from input_module import db, service
    mae = service.NovaNota(
        Numero_Nota=555100, Status_Nota="00 Pendente",
        Prioridade_Nota="Programável", Planejado_DDPM=5.0,
    )
    filha = service.NovaNota(
        Numero_Nota=555101, Status_Nota="00 Pendente",
        Prioridade_Nota="Programável", Planejado_DDPM=1.0,
        Nota_Mae="555100",
    )
    assert service.criar_notas([mae, filha], usuario="teste") == 2
    df = db.carregar_dados()
    linha_filha = df[df["Numero_Nota"] == 555101].iloc[0]
    assert str(linha_filha["Nota_Mae"]) == "555100"
    logs = db.carregar_logs()
    log_filha = logs[logs["Numero_Nota"] == 555101].iloc[0]
    assert "Mãe: 555100" in log_filha["Valor_Novo"]


# ── Task 13: versão do dataset (cache/ETag de GET /notas) ────────────────
def test_versao_dataset_muda_com_escritas(banco_temporario):
    from input_module import db, service
    import datetime
    v0 = db.obter_versao_dataset()
    nota = service.NovaNota(Numero_Nota=777001, Status_Nota="00 Pendente", Prioridade_Nota="Programável")
    service.criar_notas([nota], usuario="teste")           # criação não loga: pega pelo COUNT(notas)
    v1 = db.obter_versao_dataset()
    assert v1 != v0
    db.aplicar_edicoes([{"Numero_Nota": 777001, "Observacao": "editada"}], usuario="teste")
    v2 = db.obter_versao_dataset()
    assert v2 != v1
    db.salvar_log_arquivo("Gerada_base_IW28.XLSX", "robo-sap", datetime.datetime.now(), "Sync SAP")
    assert db.obter_versao_dataset() != v2


def test_salvar_log_arquivo_retorna_se_gravou(banco_temporario, monkeypatch):
    """O log de arquivos é best-effort, mas quem chama precisa saber se gravou.

    Sem esse retorno, uma auditoria perdida é indistinguível de uma gravada — e
    a rota de upload responde "ok" para um upload que ninguém consegue rastrear.
    Falha ao abrir o banco conta como não gravou, não como estouro."""
    from input_module import db
    import datetime
    agora = datetime.datetime.now()
    assert db.salvar_log_arquivo("Base.xlsx", "ana", agora, "Substituição") is True

    def banco_fora_do_ar():
        raise sqlite3.OperationalError("unable to open database file")

    monkeypatch.setattr(db, "get_db_connection", banco_fora_do_ar)
    assert db.salvar_log_arquivo("Base.xlsx", "ana", agora, "Substituição") is False


def test_salvar_base_dataframe_separa_conexao_de_gravacao(banco_temporario, monkeypatch):
    """Banco que nem abriu é uma falha diferente de gravação que quebrou.

    `to_sql(if_exists="replace")` dropa a tabela antiga, então uma falha DURANTE
    a gravação pode ter mexido no banco; uma conexão que não abriu não tocou em
    nada. Sem essa distinção quem trata o erro não sabe se tem o que desfazer."""
    from input_module import db
    df = pd.DataFrame({"CONJUNTO_DESC": ["POA"]})

    def banco_fora_do_ar():
        raise sqlite3.OperationalError("unable to open database file")

    monkeypatch.setattr(db, "get_db_connection", banco_fora_do_ar)
    with pytest.raises(db.GravacaoNaoIniciadaErro):
        db.salvar_base_dataframe("base_clientes", df)

    def conexao_somente_leitura():
        return sqlite3.connect(f"file:{db.obter_caminho_banco()}?mode=ro", uri=True)

    # Conexão que abre e só falha no DROP/CREATE do `to_sql`: a gravação começou,
    # então o erro NÃO é GravacaoNaoIniciadaErro (que não é OperationalError).
    monkeypatch.setattr(db, "get_db_connection", conexao_somente_leitura)
    with pytest.raises(sqlite3.OperationalError, match="readonly"):
        db.salvar_base_dataframe("base_clientes", df)


# ── Task 14: cache do engine revalidado por versão do dataset ───────────
def test_get_dataset_revalida_por_versao(banco_temporario, monkeypatch):
    from input_module import db, engine, service
    engine.invalidar_cache()
    df1 = engine.get_dataset()
    chamadas = {"n": 0}
    original = engine.enriquecer_dados
    def contando():
        chamadas["n"] += 1
        return original()
    monkeypatch.setattr(engine, "enriquecer_dados", contando)
    engine.get_dataset()                      # versão igual: serve do cache
    assert chamadas["n"] == 0
    nota = service.NovaNota(Numero_Nota=888001, Status_Nota="00 Pendente", Prioridade_Nota="Programável")
    service.criar_notas([nota], usuario="teste")   # muda a versão (sem invalidar_cache manual)
    df2 = engine.get_dataset()
    assert chamadas["n"] == 1
    assert 888001 in df2["Numero_Nota"].values


# ── Tarefa 4: motor de enriquecimento, auditoria, cache e cópia Excel ────
def _excel_iw28(caminho):
    pd.DataFrame({
        "Nota": [2000], "Status usuário": ["LIBE"],
        "CenTrabalho princ.": ["CT-01"], "Ordem": [777],
        "Encerram.por data": [pd.Timestamp("2026-05-10")],
    }).to_excel(caminho, index=False)


def _excel_iw38(caminho):
    pd.DataFrame({
        "Ordem": [777], "Status usuário": ["JAND INVE"],
        "Status do sistema": ["ENTE"], "Total planejado": [1000.0],
        "Total real": [800.0],
    }).to_excel(caminho, index=False)


# engine.py lê IW28/IW38/IW66 do SQLite nativo (base_iw28/base_iw38/base_iw66),
# não mais do Excel de rede — os helpers acima continuam servindo test_status_bases
# (que checa existência do arquivo, não o carregamento dos dados).
def _sqlite_iw28():
    from input_module import db
    db.salvar_base_dataframe("base_iw28", pd.DataFrame({
        "Nota": [2000], "Status usuário": ["LIBE"],
        "CenTrabalho princ.": ["CT-01"], "Ordem": [777],
        "Encerram.por data": [pd.Timestamp("2026-05-10")],
    }))


def _sqlite_iw38():
    from input_module import db
    db.salvar_base_dataframe("base_iw38", pd.DataFrame({
        "Ordem": [777], "Status usuário": ["JAND INVE"],
        "Status do sistema": ["ENTE"], "Total planejado": [1000.0],
        "Total real": [800.0],
    }))


@pytest.fixture
def engine_isolado(banco_temporario, monkeypatch, tmp_path):
    """Banco temporário + caminhos de rede apontando para tmp (inexistentes por padrão)."""
    from input_module import config, engine
    for attr in ["CAMINHO_INDICADOR_CONTINUIDADE", "CAMINHO_BASE_IW28",
                 "CAMINHO_CUSTO_ORD_IW38", "CAMINHO_CLIENTES_CONJUNTO",
                 "CAMINHO_CUSTO_MODULAR", "CAMINHO_GANHOS",
                 "CAMINHO_PROJETO_CONSTRUCAO", "CAMINHO_BASE_IW66"]:
        monkeypatch.setattr(config, attr, str(tmp_path / f"{attr}.xlsx"))
    monkeypatch.setattr(config, "BASES_REDE", {
        "IW28": config.CAMINHO_BASE_IW28, "IW38": config.CAMINHO_CUSTO_ORD_IW38})
    engine.invalidar_cache()
    return tmp_path


def test_engine_fallbacks_sem_rede(engine_isolado):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    df = engine.enriquecer_dados()
    linha = df[df["Numero_Nota"] == 2000].iloc[0]
    assert linha["Export_status"] == "Pendente Extração SAP"
    assert linha["Conj.critico"] == "-"
    assert linha["Cidade"] == "Guarulhos"
    assert "Auditoria_Cronograma" in df.columns


def test_engine_cruza_iw28_iw38(engine_isolado):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(2000, Status_Nota="99 Encerrado")]))
    _sqlite_iw28()
    _sqlite_iw38()
    df = engine.enriquecer_dados()
    linha = df[df["Numero_Nota"] == 2000].iloc[0]
    assert linha["Export_status"] == "LIBE"
    assert linha["Ordem"] == "777"
    assert linha["Ordem_Executada"] == "SIM"
    assert float(linha["Total_real_ordem"]) == 800.0
    assert float(linha["Exec_percentagem_ordem"]) == pytest.approx(80.0)


def test_auditoria_cronograma(engine_isolado):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(2000, Status_Nota="99 Encerrado")]))
    _sqlite_iw28()
    _sqlite_iw38()
    df = engine.enriquecer_dados()
    assert df.iloc[0]["Auditoria_Cronograma"] == "🔵 Adiantado"


def test_engine_totais_numericos_e_modular(engine_isolado):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(2000, Status_Nota="99 Encerrado")]))
    _sqlite_iw28()
    _sqlite_iw38()
    df = engine.enriquecer_dados()
    linha = df[df["Numero_Nota"] == 2000].iloc[0]
    assert isinstance(linha["Total_planejado_ordem"], (int, float))
    assert isinstance(linha["Total_real_ordem"], (int, float))
    assert float(linha["Total_planejado_ordem"]) == 1000.0
    assert float(linha["Total_real_ordem"]) == 800.0
    assert "Total_planejado_modular" in df.columns
    assert float(linha["Total_planejado_modular"]) == 0.0


def test_status_map_grupo_c():
    assert config.STATUS_MAP[53].startswith("53 "), "53 deve ter prefixo numérico"
    assert 997 in config.STATUS_MAP, "997 (SUPR CANC) ausente do STATUS_MAP"
    assert config.STATUS_MAP[997] == "SUPR CANC"
    assert config.INV_STATUS_MAP[config.STATUS_MAP[53]] == 53
    assert config.INV_STATUS_MAP["SUPR CANC"] == 997


def test_status_para_int_grupo_c(banco_temporario):
    from input_module.db import status_para_int
    assert status_para_int("53 Programado Execução") == 53
    assert status_para_int("SUPR CANC") == 997
    assert status_para_int("SUPR") == 998
    assert status_para_int("ENCE EXEC") == 999


def test_converter_para_iso_data(banco_temporario):
    from input_module.db import converter_para_iso_data
    assert converter_para_iso_data("-") == "-"
    assert converter_para_iso_data("") == "-"
    assert converter_para_iso_data("jun-2026") == "2026-06-01"
    assert converter_para_iso_data("junho-2026") == "2026-06-01"
    assert converter_para_iso_data("2026-06-01") == "2026-06-01"
    assert converter_para_iso_data("2026-06-01 00:00:00") == "2026-06-01"
    assert converter_para_iso_data("dez-2025") == "2025-12-01"


def test_salvar_em_massa_preserva_mes_iso(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(9900, Mes_Execucao_Planejado="jun-2026")]))
    conn = db.get_db_connection()
    row = conn.execute("SELECT Mes_Execucao_Planejado FROM notas WHERE Numero_Nota=9900").fetchone()
    conn.close()
    assert row[0] == "2026-06-01", f"DB deve guardar ISO, encontrado: {row[0]}"


def test_config_iw66():
    assert hasattr(config, "CAMINHO_BASE_IW66")
    assert "IW66" in config.CAMINHO_BASE_IW66.upper() or "medidas" in config.CAMINHO_BASE_IW66.lower()
    assert len(config.BASES_REDE) == 7
    assert "Medida_SAP" in config.COLUNAS_PAINEL
    assert "Medida_vs_Planejado" in config.COLUNAS_PAINEL
    assert "Medida SAP" in config.MAP_FILTROS


def _excel_iw66(caminho):
    pd.DataFrame({
        "Nota": [2000, 2000, 2000],
        "Denominação do conjunto": ["REDE", "POSTE", "REDE"],
        "Texto medida": ["CABO", "POSTE", "CONDUTOR"],
        "Descrição": ["", "", ""],
        "Nº de ordenação": [500.0, 2.0, 300.0],
    }).to_excel(caminho, index=False)


def _sqlite_iw66():
    from input_module import db
    db.salvar_base_dataframe("base_iw66", pd.DataFrame({
        "Nota": [2000, 2000, 2000],
        "Denominação do conjunto": ["REDE", "POSTE", "REDE"],
        "Texto medida": ["CABO", "POSTE", "CONDUTOR"],
        "Descrição": ["", "", ""],
        "Nº de ordenação": [500.0, 2.0, 300.0],
    }))


def test_engine_medidas_iw66_sem_arquivo(engine_isolado):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    df = engine.enriquecer_dados()
    assert "Medida_SAP" in df.columns
    assert "Medida_vs_Planejado" in df.columns
    assert df.iloc[0]["Medida_SAP"] == "-"
    assert df.iloc[0]["Medida_vs_Planejado"] == "-"


def test_engine_medidas_iw66_com_dados(engine_isolado):
    from input_module import db, engine
    _sqlite_iw66()
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    engine.invalidar_cache()
    df = engine.enriquecer_dados()
    linha = df[df["Numero_Nota"] == 2000].iloc[0]
    assert "km" in str(linha["Medida_SAP"])
    assert "un" in str(linha["Medida_SAP"])
    assert linha["Medida_vs_Planejado"] in ("Sim", "Não")


def test_comparar_medida_planejado():
    from input_module.engine import _comparar_medida_planejado
    assert _comparar_medida_planejado("-", 2.0) == "-"
    assert _comparar_medida_planejado("0.8 km", float("nan")) == "-"
    assert _comparar_medida_planejado("0.8 km", 800.0) == "Sim"
    assert _comparar_medida_planejado("0.8 km", 900.0) == "Não"
    assert _comparar_medida_planejado("2 un", 2.0) == "Sim"
    assert _comparar_medida_planejado("0.8 km / 2 un", 800.0) == "Sim"


def test_cache_e_invalidacao(engine_isolado):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    df1 = engine.get_dataset()
    assert len(engine.get_dataset()) == len(df1)  # mesma versão: cache segura
    db.salvar_em_massa(pd.DataFrame([_nota(2001)]))
    assert len(engine.get_dataset()) == len(df1) + 1  # nova versão: revalida sozinho
    engine.invalidar_cache()
    assert len(engine.get_dataset()) == len(df1) + 1  # invalidação manual continua funcionando


def test_status_bases(engine_isolado):
    from input_module import config, engine
    engine.invalidar_status_bases()
    _excel_iw28(config.CAMINHO_BASE_IW28)
    bases = engine.status_bases()
    por_nome = {b["nome"]: b for b in bases}
    assert por_nome["IW28"]["encontrada"] is True
    assert por_nome["IW38"]["encontrada"] is False


def test_invalidar_status_bases_forca_releitura(engine_isolado):
    """Sem invalidar_status_bases(), um arquivo criado após o primeiro status_bases()
    fica invisível até o TTL de 60s expirar — a invalidação explícita evita isso."""
    from input_module import config, engine
    engine.invalidar_status_bases()
    bases = engine.status_bases()
    por_nome = {b["nome"]: b for b in bases}
    assert por_nome["IW28"]["encontrada"] is False
    _excel_iw28(config.CAMINHO_BASE_IW28)
    engine.invalidar_status_bases()
    bases = engine.status_bases()
    por_nome = {b["nome"]: b for b in bases}
    assert por_nome["IW28"]["encontrada"] is True


# ── Tarefa 5: endpoints de leitura /api/input/* ──────────────────────────
from fastapi.testclient import TestClient


@pytest.fixture
def cliente(engine_isolado):
    from main import app
    from input_module import service
    service.resetar_migracao()
    return TestClient(app)


def test_get_notas_traz_registros_e_meta(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    from input_module import engine
    engine.invalidar_cache()
    r = cliente.get("/api/input/notas")
    assert r.status_code == 200
    corpo = r.json()
    assert len(corpo["registros"]) == 1
    assert corpo["registros"][0]["Numero_Nota"] == 2000
    meta = corpo["meta"]
    assert "99 Encerrado" in meta["status_opcoes"]
    assert "Emergente" in meta["prioridade_opcoes"]
    assert isinstance(meta["bases"], list)
    assert "ultima_alteracao" in meta
    assert meta["migracao"] in ("ja-existe", "migrado", "rede-indisponivel")


def test_get_sync(cliente):
    r = cliente.get("/api/input/sync")
    assert r.status_code == 200
    assert "ultima_alteracao" in r.json()


def test_notas_etag_304(banco_temporario):
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
    from input_module.routes import router
    app = FastAPI(); app.include_router(router)
    client = TestClient(app)
    r1 = client.get("/api/input/notas")
    assert r1.status_code == 200
    etag = r1.headers["etag"]
    assert r1.json()["meta"]["versao"]
    r2 = client.get("/api/input/notas", headers={"If-None-Match": etag})
    assert r2.status_code == 304
    r3 = client.get("/api/input/sync")
    assert "versao" in r3.json()


def test_get_logs_e_timeline(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    db.aplicar_edicoes([{"Numero_Nota": 2000, "Observacao": "oi"}], usuario="ana")
    assert len(cliente.get("/api/input/logs").json()["registros"]) == 1
    assert len(cliente.get("/api/input/logs/nota/2000").json()["registros"]) == 1
    assert cliente.get("/api/input/logs/nota/999").json()["registros"] == []
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []


# ── Tarefa 6: endpoints de escrita /api/input/* ──────────────────────────
CABECALHO_USER = {"X-User": "ana"}


def test_escrita_exige_x_user(cliente):
    r = cliente.patch("/api/input/notas", json={"linhas": []})
    assert r.status_code == 400
    assert "X-User" in r.json()["detail"]


def test_patch_edita_e_loga(cliente):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(2000)]))
    engine.invalidar_cache()
    r = cliente.patch("/api/input/notas", headers=CABECALHO_USER,
                      json={"linhas": [{"Numero_Nota": 2000, "Observacao": "via api"}]})
    assert r.status_code == 200
    assert r.json()["alteradas"] == 1
    registros = cliente.get("/api/input/notas").json()["registros"]
    assert registros[0]["Observacao"] == "via api"


def test_patch_nota_inexistente_404(cliente):
    r = cliente.patch("/api/input/notas", headers=CABECALHO_USER,
                      json={"linhas": [{"Numero_Nota": 31337, "Observacao": "x"}]})
    assert r.status_code == 404


def test_post_cria_e_rejeita_duplicata(cliente):
    nova = {"Numero_Nota": 6000, "Status_Nota": "00 Pendente",
            "Prioridade_Nota": "Programável", "Local_Instalacao": "045 RL X"}
    r = cliente.post("/api/input/notas", headers=CABECALHO_USER, json=nova)
    assert r.status_code == 200
    from input_module import db
    df = db.carregar_dados()
    linha = df[df["Numero_Nota"] == 6000].iloc[0]
    assert linha["Regional"] == "Guarulhos"
    r = cliente.post("/api/input/notas", headers=CABECALHO_USER, json=nova)
    assert r.status_code == 409


def test_bulk_valida_duplicatas(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(7000)]))
    lote = {"notas": [
        {"Numero_Nota": 7000, "Status_Nota": "00 Pendente", "Prioridade_Nota": "Programável"},
        {"Numero_Nota": 7001, "Status_Nota": "00 Pendente", "Prioridade_Nota": "Programável"},
    ]}
    r = cliente.post("/api/input/notas/bulk", headers=CABECALHO_USER, json=lote)
    assert r.status_code == 409
    assert "7000" in r.json()["detail"]
    lote = {"notas": [
        {"Numero_Nota": 7002, "Status_Nota": "00 Pendente", "Prioridade_Nota": "Programável"},
        {"Numero_Nota": 7002, "Status_Nota": "00 Pendente", "Prioridade_Nota": "Programável"},
    ]}
    assert cliente.post("/api/input/notas/bulk", headers=CABECALHO_USER, json=lote).status_code == 409
    lote = {"notas": [
        {"Numero_Nota": 7003, "Status_Nota": "00 Pendente", "Prioridade_Nota": "Programável"},
        {"Numero_Nota": 7004, "Status_Nota": "00 Pendente", "Prioridade_Nota": "Programável"},
    ]}
    r = cliente.post("/api/input/notas/bulk", headers=CABECALHO_USER, json=lote)
    assert r.status_code == 200
    assert r.json()["inseridas"] == 2


def test_delete_e_desfazer(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(8000)]))
    cliente.patch("/api/input/notas", headers=CABECALHO_USER,
                  json={"linhas": [{"Numero_Nota": 8000, "Observacao": "antes do undo"}]})
    r = cliente.post("/api/input/desfazer", headers=CABECALHO_USER, json={})
    assert r.status_code == 200 and r.json()["ok"] is True
    r = cliente.request("DELETE", "/api/input/notas", headers=CABECALHO_USER,
                        json={"numeros": [8000]})
    assert r.status_code == 200 and r.json()["excluidas"] == 1


def test_travar_e_listar_bloqueios_api(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(8100)]))
    r = cliente.post("/api/input/notas/8100/travar", headers=CABECALHO_USER, json={})
    assert r.status_code == 200 and r.json()["ok"] is True

    ativos = cliente.get("/api/input/bloqueios").json()["bloqueios"]
    assert any(b["Numero_Nota"] == 8100 and b["Usuario"] == "ana" for b in ativos)

    r = cliente.post("/api/input/notas/8100/travar", headers={"X-User": "bob"}, json={})
    assert r.status_code == 200  # não é erro HTTP — o conflito vem no corpo, como /desfazer
    assert r.json()["ok"] is False
    assert r.json()["usuario"] == "ana"


def test_destravar_api(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(8101)]))
    cliente.post("/api/input/notas/8101/travar", headers=CABECALHO_USER, json={})
    r = cliente.post("/api/input/notas/destravar", headers=CABECALHO_USER,
                     json={"numeros": [8101]})
    assert r.status_code == 200 and r.json()["liberadas"] == 1
    assert cliente.get("/api/input/bloqueios").json()["bloqueios"] == []


def test_patch_retorna_notas_bloqueadas(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(8102)]))
    cliente.post("/api/input/notas/8102/travar", headers={"X-User": "outra"}, json={})
    r = cliente.patch("/api/input/notas", headers=CABECALHO_USER,
                      json={"linhas": [{"Numero_Nota": 8102, "Observacao": "via api"}]})
    assert r.status_code == 200
    assert r.json()["alteradas"] == 0
    assert r.json()["bloqueadas"] == [8102]


def test_export_gera_xlsx(cliente):
    from input_module import db, engine
    db.salvar_em_massa(pd.DataFrame([_nota(9000)]))
    engine.invalidar_cache()
    r = cliente.post("/api/input/export",
                     json={"numeros": [9000], "colunas": ["Numero_Nota", "Status_Nota"]})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    df = pd.read_excel(io.BytesIO(r.content))
    assert list(df.columns) == ["Nº Nota (ID)", "Status Nota"]


# ── Tarefa 7: endpoints de configuração (responsáveis, bases, backups, migração) ──
def test_responsaveis_api(cliente):
    r = cliente.get("/api/input/responsaveis")
    assert r.json()["Mogi das Cruzes"] == "Fabricio"
    r = cliente.put("/api/input/responsaveis", headers=CABECALHO_USER,
                    json={"Mogi das Cruzes": "Maria"})
    assert r.status_code == 200
    assert cliente.get("/api/input/responsaveis").json() == {"Mogi das Cruzes": "Maria"}


def test_bases_lista_download_upload(cliente, monkeypatch, tmp_path):
    from input_module import config
    caminho = tmp_path / "Clientes_Conjunto.xlsx"
    pd.DataFrame({"CONJUNTO_DESC": ["POA"], "QTDE_CONJUNTO": [10]}).to_excel(caminho, index=False)
    monkeypatch.setattr(config, "BASES_APOIO", {"Clientes por Conjunto": str(caminho)})
    r = cliente.get("/api/input/bases")
    assert r.json()["bases"][0]["encontrada"] is True
    r = cliente.get("/api/input/bases/Clientes_Conjunto.xlsx/download")
    assert r.status_code == 200
    # Upload substitui o arquivo e registra no log
    conteudo = caminho.read_bytes()
    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", conteudo)})
    assert r.status_code == 200
    logs = cliente.get("/api/input/logs/arquivos").json()["registros"]
    assert logs[0]["Nome_Arquivo"] == "Clientes_Conjunto.xlsx"
    assert logs[0]["Acao"] == "Substituição"
    # Base desconhecida -> 404
    assert cliente.get("/api/input/bases/nao_existe.xlsx/download").status_code == 404


def _bytes_clientes(conjunto: str) -> bytes:
    """Bytes de um Clientes_Conjunto.xlsx válido com um único conjunto."""
    buffer = io.BytesIO()
    pd.DataFrame({"CONJUNTO_DESC": [conjunto], "QTDE_CONJUNTO": [10]}).to_excel(buffer, index=False)
    return buffer.getvalue()


def _base_apoio_temporaria(monkeypatch, tmp_path):
    """Aponta BASES_APOIO para um Clientes_Conjunto.xlsx válido em tmp_path."""
    from input_module import config
    caminho = tmp_path / "Clientes_Conjunto.xlsx"
    caminho.write_bytes(_bytes_clientes("POA"))
    monkeypatch.setattr(config, "BASES_APOIO", {"Clientes por Conjunto": str(caminho)})
    return caminho


def _base_apoio_ja_importada(monkeypatch, tmp_path):
    """Base de apoio em tmp_path com o MESMO conteúdo já refletido no SQLite."""
    from input_module import db
    caminho = _base_apoio_temporaria(monkeypatch, tmp_path)
    db.salvar_base_dataframe("base_clientes", pd.read_excel(caminho))
    return caminho


def _conjuntos_no_sqlite() -> list:
    from input_module import db
    return db.carregar_base_dataframe("base_clientes")["CONJUNTO_DESC"].tolist()


def test_upload_base_invalida_preserva_arquivo_anterior(cliente, monkeypatch, tmp_path):
    """Arquivo que não é um Excel válido é rejeitado ANTES de tocar o alvo."""
    caminho = _base_apoio_temporaria(monkeypatch, tmp_path)
    original = caminho.read_bytes()
    antes = set(tmp_path.iterdir())

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", b"isto nao e um xlsx")})

    assert r.status_code == 422
    assert "Clientes_Conjunto.xlsx" in r.json()["detail"]
    assert caminho.read_bytes() == original
    assert set(tmp_path.iterdir()) == antes
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []


def test_upload_base_ilegivel_nao_toca_o_sqlite(cliente, monkeypatch, tmp_path):
    """Parse que falha ANTES de qualquer gravação deixa o SQLite saudável quieto.

    Sem nada gravado não há o que realinhar, e reimportar o alvo só para "voltar
    ao normal" dropa e recria (`to_sql` com `if_exists="replace"`) uma tabela que
    estava certa — uma falha nessa releitura destruiria dados que o upload
    recusado nem chegou a tocar."""
    from input_module import db
    _base_apoio_ja_importada(monkeypatch, tmp_path)
    salvar_real = db.salvar_base_dataframe
    gravacoes = []

    def espiar_gravacao(nome_tabela, df):
        gravacoes.append(nome_tabela)
        salvar_real(nome_tabela, df)

    monkeypatch.setattr(db, "salvar_base_dataframe", espiar_gravacao)

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", b"isto nao e um xlsx")})

    assert r.status_code == 422
    assert gravacoes == [], "o SQLite foi reescrito sem que o upload tivesse gravado nada"
    assert _conjuntos_no_sqlite() == ["POA"]


def test_upload_base_com_banco_inalcancavel_nao_realinha_o_sqlite(cliente, monkeypatch, tmp_path):
    """Banco que não abre não conta como gravação parcial.

    A tabela continua com o conteúdo do alvo — não há o que desfazer. Realinhar
    aqui dropava e recriava (`to_sql(if_exists="replace")`) uma tabela sã, e uma
    falha nessa releitura destruiria dados que o upload nem chegou a tocar."""
    from input_module import db
    _base_apoio_ja_importada(monkeypatch, tmp_path)
    tentativas = []

    def banco_fora_do_ar(nome_tabela, _df):
        tentativas.append(nome_tabela)
        raise db.GravacaoNaoIniciadaErro("Banco indisponível ao salvar tabela base_clientes")

    monkeypatch.setattr(db, "salvar_base_dataframe", banco_fora_do_ar)

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_clientes("SUZ"))})

    assert r.status_code == 422
    assert "consistência" not in r.json()["detail"]
    assert tentativas == ["base_clientes"], "o realinhamento reimportou uma tabela intocada"
    assert _conjuntos_no_sqlite() == ["POA"]


def test_upload_base_falha_no_meio_preserva_arquivo_anterior(cliente, monkeypatch, tmp_path):
    """Excel válido cuja gravação no SQLite falha no meio não substitui o alvo.

    Aqui o SQLite está fora do ar para gravação, então nem o realinhamento passa:
    a gravação que levantou pode ter dropado a tabela, e ninguém consegue provar
    que o banco voltou ao conteúdo do alvo — a resposta tem que dizer isso."""
    from input_module import db
    caminho = _base_apoio_temporaria(monkeypatch, tmp_path)
    original = caminho.read_bytes()

    novo = io.BytesIO()
    pd.DataFrame({"CONJUNTO_DESC": ["SUZ"], "QTDE_CONJUNTO": [99]}).to_excel(novo, index=False)

    def falhar(*_args, **_kwargs):
        raise sqlite3.OperationalError("database is locked")

    monkeypatch.setattr(db, "salvar_base_dataframe", falhar)
    antes = set(tmp_path.iterdir())

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", novo.getvalue())})

    assert r.status_code == 500
    assert "database is locked" in r.json()["detail"]
    assert "consistência" in r.json()["detail"]
    assert caminho.read_bytes() == original
    assert set(tmp_path.iterdir()) == antes
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []


def test_upload_base_bem_sucedido_substitui_o_alvo(cliente, monkeypatch, tmp_path):
    """Só depois do import bem-sucedido o conteúdo novo aparece no alvo."""
    caminho = _base_apoio_temporaria(monkeypatch, tmp_path)

    novo = io.BytesIO()
    pd.DataFrame({"CONJUNTO_DESC": ["SUZ"], "QTDE_CONJUNTO": [99]}).to_excel(novo, index=False)
    antes = set(tmp_path.iterdir())

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", novo.getvalue())})

    assert r.status_code == 200
    assert pd.read_excel(caminho)["CONJUNTO_DESC"].tolist() == ["SUZ"]
    assert set(tmp_path.iterdir()) == antes


def test_upload_base_ainda_ausente_na_rede_cria_o_arquivo(cliente, monkeypatch, tmp_path):
    """Base gerenciada que ainda não existe na rede: o upload a cria.

    Sem alvo para guardar de lado não há nada a preservar — o upload é a
    primeira versão da base, não um erro de rede."""
    from input_module import config
    caminho = tmp_path / "Clientes_Conjunto.xlsx"
    monkeypatch.setattr(config, "BASES_APOIO", {"Clientes por Conjunto": str(caminho)})
    antes = set(tmp_path.iterdir())

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_clientes("SUZ"))})

    assert r.status_code == 200
    assert pd.read_excel(caminho)["CONJUNTO_DESC"].tolist() == ["SUZ"]
    assert _conjuntos_no_sqlite() == ["SUZ"]
    assert set(tmp_path.iterdir()) == antes | {caminho}
    assert cliente.get("/api/input/logs/arquivos").json()["registros"][0]["Acao"] == "Substituição"


def test_upload_base_falha_no_replace_preserva_excel_e_sqlite(cliente, monkeypatch, tmp_path):
    """Falha ao trocar o arquivo não pode deixar o SQLite com a base nova."""
    caminho = _base_apoio_ja_importada(monkeypatch, tmp_path)
    original = caminho.read_bytes()
    antes = set(tmp_path.iterdir())

    def replace_falha(*_args, **_kwargs):
        raise OSError("a rede caiu no meio da troca")

    monkeypatch.setattr(os, "replace", replace_falha)

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_clientes("SUZ"))})

    assert r.status_code == 502
    assert caminho.read_bytes() == original
    assert _conjuntos_no_sqlite() == ["POA"]
    assert set(tmp_path.iterdir()) == antes
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []


def test_upload_base_nunca_deixa_o_alvo_ausente(cliente, monkeypatch, tmp_path):
    """O alvo só muda no `os.replace` final — nunca sai do lugar antes disso.

    Enquanto o import para o SQLite roda, o Excel da rede tem que continuar
    legível com o conteúdo antigo: uma queda do processo nessa janela deixa a
    base anterior no lugar, e não um alvo ausente que ninguém recupera."""
    from input_module import db
    caminho = _base_apoio_ja_importada(monkeypatch, tmp_path)
    original = caminho.read_bytes()
    antes = set(tmp_path.iterdir())
    salvar_real = db.salvar_base_dataframe
    replace_real = os.replace
    alvo_durante_import = []
    origens_dos_replaces = []

    def espiar_replace(origem, destino, *args, **kwargs):
        origens_dos_replaces.append(str(origem))
        return replace_real(origem, destino, *args, **kwargs)

    def espiar_gravacao(nome_tabela, df):
        alvo_durante_import.append(caminho.read_bytes() if caminho.exists() else None)
        salvar_real(nome_tabela, df)

    monkeypatch.setattr(os, "replace", espiar_replace)
    monkeypatch.setattr(db, "salvar_base_dataframe", espiar_gravacao)

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_clientes("SUZ"))})

    assert r.status_code == 200
    assert alvo_durante_import == [original], "o alvo mudou (ou sumiu) antes do import terminar"
    assert str(caminho) not in origens_dos_replaces, "o alvo foi movido do lugar"
    assert pd.read_excel(caminho)["CONJUNTO_DESC"].tolist() == ["SUZ"]
    assert set(tmp_path.iterdir()) == antes


def _bytes_custo_modular(item: str) -> bytes:
    """Bytes de um Custo_Modular.xlsx válido (aba 'Modulares')."""
    buffer = io.BytesIO()
    df = pd.DataFrame({"ITEM": [item, item, item], "VALOR": [1, 2, 3]})
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Modulares")
    return buffer.getvalue()


def test_upload_base_multi_aba_com_import_parcial_realinha_o_sqlite(cliente, monkeypatch, tmp_path):
    """Custo_Modular grava DUAS tabelas a partir do mesmo Excel.

    Se a segunda gravação falhar, a primeira não pode ficar com dados de um
    arquivo que nunca entrou na rede: o SQLite volta a refletir o alvo."""
    from input_module import config, db, routes
    caminho = tmp_path / "Custo_Modular.xlsx"
    caminho.write_bytes(_bytes_custo_modular("ANTIGO"))
    monkeypatch.setattr(config, "BASES_APOIO", {"Custo Modular": str(caminho)})
    routes._importar_base_para_sqlite("Custo_Modular.xlsx", str(caminho))
    original = caminho.read_bytes()
    antes = set(tmp_path.iterdir())

    salvar_real = db.salvar_base_dataframe
    ja_falhou = []

    def falhar_uma_vez_na_sazonal(nome_tabela, df):
        if nome_tabela == "base_sazonal" and not ja_falhou:
            ja_falhou.append(nome_tabela)
            raise sqlite3.OperationalError("database is locked")
        salvar_real(nome_tabela, df)

    monkeypatch.setattr(db, "salvar_base_dataframe", falhar_uma_vez_na_sazonal)

    r = cliente.post("/api/input/bases/Custo_Modular.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_custo_modular("NOVO"))})

    assert r.status_code == 422
    assert caminho.read_bytes() == original
    assert db.carregar_base_dataframe("base_custo_modular")["ITEM"].tolist() == ["ANTIGO"] * 3
    assert set(tmp_path.iterdir()) == antes
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []


def test_upload_base_multi_aba_le_o_workbook_inteiro_antes_de_gravar(monkeypatch, tmp_path):
    """Custo_Modular lê duas abas do MESMO Excel.

    Uma falha de leitura na segunda aba tem que acontecer antes de o SQLite ser
    tocado: um arquivo que nem foi lido por inteiro não pode ter gravado tabela
    nenhuma, nem deixar o realinhamento pós-falha com o que desfazer."""
    from input_module import db, routes
    caminho = tmp_path / "Custo_Modular.xlsx"
    leituras = []
    gravacoes = []

    def ler_e_falhar_na_segunda_aba(*args, **kwargs):
        leituras.append(kwargs.get("skiprows"))
        if len(leituras) > 1:
            raise OSError("aba ilegível")
        return pd.DataFrame({"ITEM": ["NOVO"], "VALOR": [1]})

    monkeypatch.setattr(routes.pd, "read_excel", ler_e_falhar_na_segunda_aba)
    monkeypatch.setattr(db, "salvar_base_dataframe",
                        lambda nome_tabela, df: gravacoes.append(nome_tabela))

    tabelas_tocadas: list = []
    with pytest.raises(OSError):
        routes._importar_base_para_sqlite(
            "Custo_Modular.xlsx", str(caminho), tabelas_tocadas)

    assert leituras == [None, 1], "as duas abas têm que ser lidas antes da 1ª gravação"
    assert gravacoes == [], "gravou no SQLite com o workbook ainda por ler"
    assert tabelas_tocadas == []


def _espiar_invalidacoes(monkeypatch) -> list:
    """Registra as invalidações de cache que a rota dispara."""
    from input_module import engine
    invalidacoes = []
    monkeypatch.setattr(engine, "invalidar_cache", lambda: invalidacoes.append("cache"))
    monkeypatch.setattr(engine, "invalidar_status_bases",
                        lambda: invalidacoes.append("status_bases"))
    return invalidacoes


def test_upload_base_com_realinhamento_falho_avisa_consistencia_incerta(
        cliente, monkeypatch, tmp_path):
    """Realinhamento que também falha não pode ser engolido.

    Sobra um SQLite com metade dos dados de um arquivo que nunca entrou na rede.
    A resposta tem que dizer que a consistência não pôde ser confirmada, e os
    caches — que ainda servem o estado anterior — têm que ser invalidados."""
    from input_module import config, db, routes
    caminho = tmp_path / "Custo_Modular.xlsx"
    caminho.write_bytes(_bytes_custo_modular("ANTIGO"))
    monkeypatch.setattr(config, "BASES_APOIO", {"Custo Modular": str(caminho)})
    routes._importar_base_para_sqlite("Custo_Modular.xlsx", str(caminho))
    original = caminho.read_bytes()
    antes = set(tmp_path.iterdir())

    salvar_real = db.salvar_base_dataframe

    def gravacao_instavel(nome_tabela, df):
        """A sazonal nunca grava; no realinhamento (conteúdo ANTIGO) a modular
        também falha — o banco fica com a modular do arquivo recusado."""
        if nome_tabela == "base_sazonal" or df["ITEM"].iloc[0] == "ANTIGO":
            raise sqlite3.OperationalError("database is locked")
        salvar_real(nome_tabela, df)

    monkeypatch.setattr(db, "salvar_base_dataframe", gravacao_instavel)
    invalidacoes = _espiar_invalidacoes(monkeypatch)

    r = cliente.post("/api/input/bases/Custo_Modular.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_custo_modular("NOVO"))})

    assert r.status_code == 500
    assert "consistência" in r.json()["detail"]
    assert "database is locked" in r.json()["detail"]
    assert invalidacoes == ["cache", "status_bases"]
    # O banco realmente ficou com dados que não estão no alvo — é isso que a
    # resposta 500 está avisando.
    assert db.carregar_base_dataframe("base_custo_modular")["ITEM"].tolist() == ["NOVO"] * 3
    assert caminho.read_bytes() == original
    assert set(tmp_path.iterdir()) == antes
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []


def test_upload_base_ausente_com_import_parcial_avisa_consistencia_incerta(
        cliente, monkeypatch, tmp_path):
    """Primeiro upload que quebra no meio não tem alvo de onde realinhar.

    Sem arquivo na rede, as tabelas já gravadas ficam órfãs — a rota não pode
    responder como se só a planilha tivesse sido recusada."""
    from input_module import config, db
    caminho = tmp_path / "Custo_Modular.xlsx"
    monkeypatch.setattr(config, "BASES_APOIO", {"Custo Modular": str(caminho)})
    antes = set(tmp_path.iterdir())

    salvar_real = db.salvar_base_dataframe

    def falhar_sempre_na_sazonal(nome_tabela, df):
        if nome_tabela == "base_sazonal":
            raise sqlite3.OperationalError("database is locked")
        salvar_real(nome_tabela, df)

    monkeypatch.setattr(db, "salvar_base_dataframe", falhar_sempre_na_sazonal)
    invalidacoes = _espiar_invalidacoes(monkeypatch)

    r = cliente.post("/api/input/bases/Custo_Modular.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_custo_modular("NOVO"))})

    assert r.status_code == 500
    assert "consistência" in r.json()["detail"]
    assert invalidacoes == ["cache", "status_bases"]
    assert not caminho.exists()
    assert set(tmp_path.iterdir()) == antes
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []


def _conjunto_do_xlsx(caminho):
    """Conjunto dentro de um Clientes_Conjunto.xlsx, ou None se não for um."""
    try:
        return pd.read_excel(caminho)["CONJUNTO_DESC"].iloc[0]
    except Exception:
        return None


def test_uploads_concorrentes_da_mesma_base_nao_cruzam(cliente, monkeypatch, tmp_path):
    """Dois uploads simultâneos da mesma base não podem cruzar Excel e SQLite.

    O primeiro upload é congelado exatamente na janela crítica — SQLite já
    gravado, `os.replace` ainda não feito — e o segundo é disparado ali dentro.
    A trava por alvo tem que barrar o segundo nessa janela; sem ela, ele importa
    AAA→BBB no SQLite e troca o Excel antes de o primeiro trocar o dele, e o
    par final fica de uploads diferentes."""
    caminho = _base_apoio_ja_importada(monkeypatch, tmp_path)
    replace_real = os.replace
    primeiro_na_janela = threading.Event()
    liberar_primeiro = threading.Event()
    esperas_do_primeiro = []

    def congelar_primeiro_no_replace(origem, destino, *args, **kwargs):
        if _conjunto_do_xlsx(origem) == "AAA":
            primeiro_na_janela.set()
            esperas_do_primeiro.append(liberar_primeiro.wait(timeout=30))
        return replace_real(origem, destino, *args, **kwargs)

    monkeypatch.setattr(os, "replace", congelar_primeiro_no_replace)

    respostas = {}

    def enviar(conjunto: str) -> None:
        respostas[conjunto] = cliente.post(
            "/api/input/bases/Clientes_Conjunto.xlsx", headers=CABECALHO_USER,
            files={"arquivo": ("novo.xlsx", _bytes_clientes(conjunto))})

    primeiro = threading.Thread(target=enviar, args=("AAA",))
    primeiro.start()
    assert primeiro_na_janela.wait(timeout=30), "primeiro upload não chegou na janela crítica"

    segundo = threading.Thread(target=enviar, args=("BBB",))
    segundo.start()
    # Observação, não sincronização: o segundo já entrou e tem que estar preso na
    # trava enquanto o primeiro segura a janela. Sem trava ele termina aqui.
    segundo.join(timeout=1.0)
    assert segundo.is_alive(), "o segundo upload passou por cima da janela do primeiro"
    assert _conjuntos_no_sqlite() == ["AAA"], "o segundo upload gravou no SQLite dentro da janela"

    liberar_primeiro.set()
    primeiro.join(timeout=30)
    segundo.join(timeout=30)
    assert not primeiro.is_alive() and not segundo.is_alive()

    assert esperas_do_primeiro == [True]
    assert respostas["AAA"].status_code == 200
    assert respostas["BBB"].status_code == 200
    # Serializados, o último upload inteiro vence nos dois lados.
    assert pd.read_excel(caminho)["CONJUNTO_DESC"].tolist() == ["BBB"]
    assert _conjuntos_no_sqlite() == ["BBB"]


def test_upload_base_com_auditoria_perdida_ainda_muda_o_etag_das_notas(
        cliente, monkeypatch, tmp_path):
    """A revalidação de GET /notas não pode depender do log de auditoria.

    O log de arquivos é best-effort: se ele não gravar, a versão do dataset
    (derivada dele) fica idêntica à de antes do upload e o ETag também. O
    navegador então continua recebendo 304 e servindo a base ANTIGA, sem nada
    na tela indicando que os dados mudaram."""
    from input_module import db
    _base_apoio_ja_importada(monkeypatch, tmp_path)
    quente = cliente.get("/api/input/notas")
    assert quente.status_code == 200
    etag_antigo = quente.headers["etag"]

    # Best-effort que não gravou: é exatamente o que a auditoria faz hoje quando
    # o INSERT falha — não levanta, só não deixa rastro.
    monkeypatch.setattr(db, "salvar_log_arquivo", lambda *_a, **_k: None)

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_clientes("SUZ"))})

    assert r.status_code == 200
    assert _conjuntos_no_sqlite() == ["SUZ"]
    assert cliente.get("/api/input/logs/arquivos").json()["registros"] == []

    depois = cliente.get("/api/input/notas", headers={"If-None-Match": etag_antigo})
    assert depois.status_code != 304, "GET /notas revalidou como inalterado depois de um upload"
    assert depois.headers["etag"] != etag_antigo


def test_upload_base_com_auditoria_que_levanta_nao_desfaz_a_publicacao(
        cliente, monkeypatch, tmp_path):
    """Auditoria que estoura DEPOIS da publicação não vira um 500 falso.

    Nesse ponto o Excel e o SQLite já estão trocados e consistentes: responder
    500 faria o usuário reenviar uma base que já entrou. A resposta é 200 com
    aviso explícito de que só a auditoria falhou, e os caches são invalidados
    do mesmo jeito — senão a tela segue no conteúdo antigo."""
    from input_module import db
    caminho = _base_apoio_ja_importada(monkeypatch, tmp_path)
    antes = set(tmp_path.iterdir())

    def auditoria_fora_do_ar(*_a, **_k):
        raise sqlite3.OperationalError("unable to open database file")

    monkeypatch.setattr(db, "salvar_log_arquivo", auditoria_fora_do_ar)
    invalidacoes = _espiar_invalidacoes(monkeypatch)

    r = cliente.post("/api/input/bases/Clientes_Conjunto.xlsx",
                     headers=CABECALHO_USER,
                     files={"arquivo": ("novo.xlsx", _bytes_clientes("SUZ"))})

    assert r.status_code == 200
    corpo = r.json()
    assert corpo["ok"] is True
    assert "auditoria" in corpo["aviso"].lower()
    assert invalidacoes == ["cache", "status_bases"]
    assert pd.read_excel(caminho)["CONJUNTO_DESC"].tolist() == ["SUZ"]
    assert _conjuntos_no_sqlite() == ["SUZ"]
    assert set(tmp_path.iterdir()) == antes


def test_backups_lista_e_download(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(9500)]))
    db.realizar_backup(limite=20, intervalo_horas=0)
    r = cliente.get("/api/input/backups")
    backups = r.json()["backups"]
    assert len(backups) >= 1
    nome = backups[0]["arquivo"]
    assert cliente.get(f"/api/input/backups/{nome}/download").status_code == 200
    legado = config_backups_dir() / "notas_departamento_20260809_181530.db"
    legado.write_bytes(b"backup legado")
    assert cliente.get(f"/api/input/backups/{legado.name}/download").status_code == 200
    assert cliente.get("/api/input/backups/..%2Fhack.db/download").status_code in (400, 404)


def test_migrar_endpoint(cliente):
    r = cliente.post("/api/input/migrar", headers=CABECALHO_USER)
    assert r.status_code == 200
    assert r.json()["resultado"] in ("ja-existe", "migrado", "rede-indisponivel")


def _iw66_temporaria(caminho: Path, valor: int) -> None:
    pd.DataFrame({"Nota": [123], "Nº de ordenação": [valor]}).to_excel(
        caminho, index=False
    )


def test_publicacao_iw66_falha_preserva_alvo_e_propaga_erro(
        banco_temporario, monkeypatch, tmp_path):
    from input_module import config, db, service

    caminho = tmp_path / "Gerada_medidas_IW66.XLSX"
    _iw66_temporaria(caminho, 10)
    original = caminho.read_bytes()
    arquivos_antes = set(tmp_path.iterdir())
    monkeypatch.setattr(config, "CAMINHO_BASE_IW66", str(caminho))
    monkeypatch.setattr(
        service.os, "replace",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("volume indisponível")),
    )
    monkeypatch.setattr(
        db, "salvar_base_dataframe",
        lambda *_args, **_kwargs: pytest.fail("SQLite não deve mudar sem publicação"),
    )

    with pytest.raises(OSError, match="volume indisponível"):
        service.atualizar_medidas_excel_local(
            [{"nota": 123, "quantidade": 2, "unidade": "m"}],
            [{"Nota": 123, "Status": "OK"}],
        )

    assert caminho.read_bytes() == original
    assert set(tmp_path.iterdir()) == arquivos_antes


def test_publicacao_iw66_valida_temporario_antes_de_substituir(
        banco_temporario, monkeypatch, tmp_path):
    from input_module import config, db, service

    caminho = tmp_path / "Gerada_medidas_IW66.XLSX"
    _iw66_temporaria(caminho, 10)
    original = caminho.read_bytes()
    arquivos_antes = set(tmp_path.iterdir())
    monkeypatch.setattr(config, "CAMINHO_BASE_IW66", str(caminho))
    read_excel_real = pd.read_excel
    leituras = []

    def validar_temporario(caminho_lido, *args, **kwargs):
        leituras.append(Path(caminho_lido))
        if Path(caminho_lido) != caminho:
            raise ValueError("temporário inválido")
        return read_excel_real(caminho_lido, *args, **kwargs)

    monkeypatch.setattr(service.pd, "read_excel", validar_temporario)
    monkeypatch.setattr(
        db, "salvar_base_dataframe",
        lambda *_args, **_kwargs: pytest.fail("SQLite não deve mudar sem publicação"),
    )

    with pytest.raises(ValueError, match="temporário inválido"):
        service.atualizar_medidas_excel_local(
            [{"nota": 123, "quantidade": 2, "unidade": "m"}],
            [{"Nota": 123, "Status": "OK"}],
        )

    assert len(leituras) == 2
    assert leituras[1].parent == caminho.parent
    assert caminho.read_bytes() == original
    assert set(tmp_path.iterdir()) == arquivos_antes


# ── Fase 4 (Grupo D): Ramal + Nota_Mae + Hierarquia ─────────────────────────
def test_inicializar_banco_cria_notas_ramal(banco_temporario):
    from input_module import db
    conn = db.get_db_connection()
    tabelas = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    colunas_notas = [r[1] for r in conn.execute("PRAGMA table_info(notas)").fetchall()]
    conn.close()
    assert "notas_ramal" in tabelas
    assert "Nota_Mae" in colunas_notas


def _nota_ramal(numero=5000, **extras):
    base = {
        "ID_Cronologia": 1, "Numero_Nota": numero,
        "Status_Obra": "-", "Conjunto": "POA", "Circuito": "POA 123",
        "Local_Instalacao": "045 RL TESTE", "Planejado_DDPM": 1.0,
        "Mes_Execucao_Planejado": "jun-2026", "CenTrab_Respon": "-",
        "Prioridade_Nota": "Programável", "Observacao": "",
        "Extracao_Antiga": "-", "Status_Nota": "00 Pendente",
        "Status_Anterior": "-", "Check_Btzero": "-", "Plano": "-",
    }
    base.update(extras)
    return base


def test_carregar_dados_ramal_vazio(banco_temporario):
    from input_module import db
    df = db.carregar_dados_ramal()
    assert df.empty
    assert "Numero_Nota" in df.columns


def test_salvar_e_carregar_ramal(banco_temporario):
    from input_module import db
    db.salvar_ramal_em_massa(pd.DataFrame([_nota_ramal(5001), _nota_ramal(5002)]))
    df = db.carregar_dados_ramal()
    assert len(df) == 2
    assert set(df["Numero_Nota"].tolist()) == {5001, 5002}
    db.salvar_ramal_em_massa(pd.DataFrame([_nota_ramal(5001, Observacao="atualizada")]))
    assert len(db.carregar_dados_ramal()) == 2  # upsert, sem duplicata


def test_deletar_notas_ramal(banco_temporario):
    from input_module import db
    db.salvar_ramal_em_massa(pd.DataFrame([_nota_ramal(5010), _nota_ramal(5011)]))
    assert db.deletar_notas_ramal([5010], usuario="tester") == 1
    assert len(db.carregar_dados_ramal()) == 1


def test_vincular_nota_mae(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(6001), _nota(6002)]))
    n = db.vincular_nota_mae_lote({"6001": [6002]}, usuario="tester")
    assert n >= 1
    df = db.carregar_dados()
    assert df[df["Numero_Nota"] == 6002].iloc[0]["Nota_Mae"] == "6001"


def test_vincular_nota_mae_respeita_lock_de_outro_usuario(banco_temporario):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(6020), _nota(6021)]))
    assert db.travar_nota(6021, "bob") == {"ok": True}

    atualizadas = db.vincular_nota_mae_lote({"6020": [6021]}, usuario="alice")

    assert atualizadas == 0
    filha = db.carregar_dados().set_index("Numero_Nota").loc[6021]
    assert filha["Nota_Mae"] == "-"
    assert db.obter_bloqueios([6021])[6021]["usuario"] == "bob"


def test_vincular_nota_mae_usa_begin_immediate(
        banco_temporario, monkeypatch):
    from input_module import db

    db.salvar_em_massa(pd.DataFrame([_nota(6030), _nota(6031)]))
    conn = db.get_db_connection()
    comandos = []
    conn.set_trace_callback(comandos.append)
    monkeypatch.setattr(db, "realizar_backup", lambda: None)
    monkeypatch.setattr(db, "get_db_connection", lambda: conn)

    assert db.vincular_nota_mae_lote({"6030": [6031]}, usuario="alice") == 1
    assert any(comando.upper() == "BEGIN IMMEDIATE" for comando in comandos)


def test_nota_mae_nao_sobrescrita_por_salvar(banco_temporario):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([_nota(6010), _nota(6011)]))
    db.vincular_nota_mae_lote({"6010": [6011]}, usuario="tester")
    db.salvar_em_massa(pd.DataFrame([_nota(6011, Observacao="editada")]))
    df = db.carregar_dados()
    assert df[df["Numero_Nota"] == 6011].iloc[0]["Nota_Mae"] == "6010"


def test_api_ramal_crud(cliente):
    from input_module import db, engine
    ramal_payload = {"notas": [{"Numero_Nota": 5100, "Conjunto": "POA"}]}
    r = cliente.post("/api/input/ramal/bulk", headers=CABECALHO_USER, json=ramal_payload)
    assert r.status_code == 200
    assert r.json()["inseridas"] == 1
    r = cliente.get("/api/input/ramal")
    assert r.status_code == 200
    assert len(r.json()["registros"]) == 1
    r = cliente.request("DELETE", "/api/input/ramal", headers=CABECALHO_USER,
                        json={"numeros": [5100]})
    assert r.status_code == 200
    assert r.json()["excluidas"] == 1
    assert cliente.get("/api/input/ramal").json()["registros"] == []


def test_api_hierarquia(cliente, monkeypatch):
    from input_module import db, engine, routes
    db.salvar_em_massa(pd.DataFrame([_nota(7010), _nota(7011)]))
    engine.invalidar_cache()
    pos_escritas = []
    monkeypatch.setattr(routes, "pos_escrita", lambda tasks: pos_escritas.append(tasks))
    r = cliente.post("/api/input/hierarquia", headers=CABECALHO_USER,
                     json={"dados": {"7010": [7011]}})
    assert r.status_code == 200
    assert r.json()["atualizadas"] >= 1
    assert len(pos_escritas) == 1
    r = cliente.get("/api/input/hierarquia/7011")
    assert r.status_code == 200
    assert r.json()["nota_mae"] == "7010"


# ── Tarefa 3: contrato de leitura IW28 + consulta de nota do plano ──
def test_iw28_obter_por_nota(banco_temporario):
    from input_module import db, iw28
    assert iw28.obter_por_nota(12345678) is None  # tabela ainda não existe
    db.salvar_base_dataframe("base_iw28", pd.DataFrame([{
        "Nota": 12345678.0, "Status usuário": "PLAN",
        "CenTrabalho princ.": "POA", "Ordem": 900001, "Encerram.por data": None,
    }]))
    registro = iw28.obter_por_nota(12345678)
    assert registro is not None
    assert registro["Status usuário"] == "PLAN"
    assert registro["Encerram.por data"] is None      # NaN vira None (JSON-safe)
    assert iw28.obter_por_nota(99999999) is None


def test_obter_nota_plano(banco_temporario):
    from input_module import db
    assert db.obter_nota_plano(1000) is None
    db.salvar_em_massa(pd.DataFrame([_nota(1000)]))
    registro = db.obter_nota_plano(1000)
    assert registro is not None
    assert registro["Status_Nota"] == "10 Em planejamento"   # formatado, não int


# ── Relatórios Home: schema e helpers de metas do Plano de Recomposição ──
def test_metas_schema_e_helpers(banco_temporario):
    from input_module import db
    conn = db.get_db_connection()
    tabelas = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    conn.close()
    assert {"metas_plano", "planos_depara", "metas_sync_estado"} <= tabelas

    metas = pd.DataFrame([
        {"Ano": 2026, "Mes": 1, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Meta": 17.0},
        {"Ano": 2026, "Mes": 2, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Meta": 19.0},
    ])
    depara = pd.DataFrame([
        {"Plano": "POSTES - CAPEX", "Nome_Curto": "POSTE", "Unidade": "Und.",
         "Area": "Construção", "Modular_RS": 6921.0, "Ordem_Exibicao": 1},
    ])
    db.substituir_metas(metas, depara)
    assert len(db.carregar_metas(2026)) == 2
    assert db.carregar_metas(2025).empty
    dp = db.carregar_planos_depara()
    assert dp.iloc[0]["Nome_Curto"] == "POSTE"

    # replace: segunda chamada substitui, não acumula
    db.substituir_metas(metas.head(1), depara)
    assert len(db.carregar_metas(2026)) == 1

    # estado de sync sobrevive e guarda erro
    assert db.obter_estado_metas() is None
    db.gravar_estado_metas(arquivo_mtime=1234.5, erro=None)
    estado = db.obter_estado_metas()
    assert estado["arquivo_mtime"] == 1234.5 and estado["erro"] is None
    db.gravar_estado_metas(arquivo_mtime=1234.5, erro="lock")
    assert db.obter_estado_metas()["erro"] == "lock"


def test_postergadas_schema_e_helpers(banco_temporario):
    from input_module import db
    conn = db.get_db_connection()
    tabelas = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    conn.close()
    assert "metas_postergadas" in tabelas

    metas = pd.DataFrame([
        {"Ano": 2026, "Mes": 1, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Meta": 17.0},
    ])
    depara = pd.DataFrame([
        {"Plano": "POSTES - CAPEX", "Nome_Curto": "POSTE", "Unidade": "Und.",
         "Area": "Construção", "Modular_RS": 6921.0, "Ordem_Exibicao": 1},
    ])
    post = pd.DataFrame([
        {"Ano": 2026, "Mes": 7, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Qtd": 3.0},
        {"Ano": 2026, "Mes": 8, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Qtd": 2.0},
    ])
    db.substituir_metas(metas, depara, post)
    p = db.carregar_postergacoes(2026)
    assert len(p) == 2
    assert db.carregar_postergacoes(2025).empty
    assert p["Qtd"].sum() == 5.0

    # replace: segunda chamada substitui, não acumula
    db.substituir_metas(metas, depara, post.head(1))
    assert len(db.carregar_postergacoes(2026)) == 1

    # df_postergacoes omitido (None) não mexe na tabela de postergadas
    db.substituir_metas(metas, depara)
    assert len(db.carregar_postergacoes(2026)) == 1


# ── Task 2: Sincronização de Metas do Controle Plano de Recomposição ───────
def test_caminho_controle_recomposicao_usa_usuario_da_maquina(monkeypatch):
    from input_module import config

    monkeypatch.delenv("CONTROLE_RECOMPOSICAO_PATH", raising=False)
    monkeypatch.setenv("USER", "usuario-sharepoint")
    monkeypatch.setenv("USERNAME", "outro-usuario")

    esperado = (
        Path("C:/Users")
        / "usuario-sharepoint"
        / "EDP"
        / "O365_Planejamento_Manutencao_EDP_Brasil - Documentos"
        / "PLANO RECOMPOSIÇÃO"
        / "SP"
        / "2026"
        / "Controle Plano de Recomposição 2026.xlsx"
    )
    assert config.caminho_controle_recomposicao() == esperado


def _xlsx_controle(caminho, meta_jan=17.0, com_postergadas=True):
    """Planilha sintética mínima com abas base, dexpara e (opcional) Postergadas."""
    import gc
    base = pd.DataFrame([
        {"Regionais": "Guarulhos", "Mês": pd.Timestamp(2026, 1, 1),
         "Plano": "POSTES - CAPEX", "Meta": meta_jan, "Conjunto": "POSTE"},
        {"Regionais": "Guarulhos", "Mês": pd.Timestamp(2026, 2, 1),
         "Plano": "POSTE DEMANDA - CAPEX", "Meta": 5.0, "Conjunto": "POSTE"},
        {"Regionais": "Poa/Suzano", "Mês": pd.Timestamp(2026, 1, 1),
         "Plano": "RAMAL", "Meta": 100.0, "Conjunto": "RAMAL"},
    ])
    dexpara = pd.DataFrame([
        {"Projeto": "POSTES - CAPEX", "Unidade": "Und.", "Área": "Projeto", "Modular R$": 6921.0},
        {"Projeto": "POSTE DEMANDA - CAPEX", "Unidade": "Und.", "Área": "Projeto", "Modular R$": 6921.0},
        {"Projeto": "RAMAL", "Unidade": "Ponto", "Área": "CSD", "Modular R$": 694.5},
    ])
    postergadas = pd.DataFrame([
        {"Regional": "Guarulhos", "Mês de Execução Planejado - DDPM": pd.Timestamp(2026, 7, 1),
         "Projeto\nConstrução": "POSTES - CAPEX", "Planejado-DDPM": 3.0},
        {"Regional": "Guarulhos", "Mês de Execução Planejado - DDPM": pd.Timestamp(2026, 8, 1),
         "Projeto\nConstrução": "POSTES - CAPEX", "Planejado-DDPM": 2.0},
    ])
    w = pd.ExcelWriter(caminho, engine="openpyxl")
    try:
        base.to_excel(w, sheet_name="base", index=False)
        dexpara.to_excel(w, sheet_name="dexpara", index=False)
        if com_postergadas:
            postergadas.to_excel(w, sheet_name="Postergadas", index=False)
    finally:
        w.close()
        del w
        gc.collect()


def test_metas_sincronizar(banco_temporario, monkeypatch, tmp_path):
    from input_module import config, db, metas
    arquivo = tmp_path / "Controle.xlsx"
    _xlsx_controle(arquivo)
    monkeypatch.setenv("CONTROLE_RECOMPOSICAO_PATH", str(arquivo))

    estado = metas.sincronizar_se_preciso()
    assert estado["sincronizou"] is True and estado["erro"] is None
    m = db.carregar_metas(2026)
    assert len(m) == 3
    assert m[(m["Regional"] == "Poa/Suzano") & (m["Plano"] == "RAMAL")].iloc[0]["Meta"] == 100.0
    dp = db.carregar_planos_depara().set_index("Plano")
    assert dp.loc["POSTES - CAPEX", "Nome_Curto"] == "POSTE"
    # colisão de nome curto: POSTE DEMANDA - CAPEX não pode virar "POSTE" também
    assert dp.loc["POSTE DEMANDA - CAPEX", "Nome_Curto"] == "POSTE DEMANDA"
    assert dp.loc["POSTES - CAPEX", "Area"] == "Construção"   # "Projeto" -> exibição
    assert dp.loc["RAMAL", "Area"] == "CSD"

    # mtime igual: no-op
    assert metas.sincronizar_se_preciso()["sincronizou"] is False
    # arquivo mudou: reimporta
    import time as _t; _t.sleep(0.05)
    _xlsx_controle(arquivo, meta_jan=99.0)
    estado = metas.sincronizar_se_preciso()
    assert estado["sincronizou"] is True
    m = db.carregar_metas(2026)
    assert m[(m["Plano"] == "POSTES - CAPEX") & (m["Mes"] == 1)].iloc[0]["Meta"] == 99.0
    # sync registra em log_arquivos (bumpa a versão do dataset)
    logs = db.carregar_log_arquivos()
    assert (logs["Usuario"] == "metas-sync").any()


def test_metas_sincronizar_falha_preserva(banco_temporario, monkeypatch, tmp_path):
    from input_module import db, metas
    arquivo = tmp_path / "Controle.xlsx"
    _xlsx_controle(arquivo)
    monkeypatch.setenv("CONTROLE_RECOMPOSICAO_PATH", str(arquivo))
    metas.sincronizar_se_preciso()
    assert len(db.carregar_metas(2026)) == 3

    arquivo.unlink()  # arquivo some (rede/OneDrive fora)
    estado = metas.sincronizar_se_preciso(forcar=True)
    assert estado["erro"] is not None
    assert len(db.carregar_metas(2026)) == 3  # última sync preservada


def test_metas_sincronizar_postergadas(banco_temporario, monkeypatch, tmp_path):
    from input_module import db, metas
    arquivo = tmp_path / "Controle.xlsx"
    _xlsx_controle(arquivo)
    monkeypatch.setenv("CONTROLE_RECOMPOSICAO_PATH", str(arquivo))
    metas.sincronizar_se_preciso()
    p = db.carregar_postergacoes(2026)
    assert p["Qtd"].sum() == 5.0
    jul = p[(p["Mes"] == 7) & (p["Plano"] == "POSTES - CAPEX")]
    assert jul.iloc[0]["Qtd"] == 3.0


def test_metas_sincronizar_sem_aba_postergadas_preserva(banco_temporario, monkeypatch, tmp_path):
    from input_module import db, metas
    arquivo = tmp_path / "Controle.xlsx"
    _xlsx_controle(arquivo)
    monkeypatch.setenv("CONTROLE_RECOMPOSICAO_PATH", str(arquivo))
    metas.sincronizar_se_preciso()
    assert len(db.carregar_postergacoes(2026)) == 2

    import time as _t; _t.sleep(0.05)
    _xlsx_controle(arquivo, com_postergadas=False)  # aba Postergadas some
    estado = metas.sincronizar_se_preciso(forcar=True)
    assert estado["erro"] is not None
    assert len(db.carregar_postergacoes(2026)) == 2   # última sync preservada
    assert len(db.carregar_metas(2026)) == 3          # metas também intactas


@pytest.mark.parametrize("valor", [99, 99.0, "99", "99.0", "99 Encerrado"])
def test_relatorios_reconhece_codigo_99_exato(valor):
    from input_module import relatorios

    assert relatorios._status_99(valor) is True


@pytest.mark.parametrize("valor", [None, "-", 9, 98, 999, "999", "99A"])
def test_relatorios_nao_confunde_outros_status_com_99(valor):
    from input_module import relatorios

    assert relatorios._status_99(valor) is False


def test_relatorios_status_final_preenchido_prevalece_sobre_status_nota():
    from input_module import relatorios

    row = pd.Series({
        "Status_Final": "10 Em planejamento",
        "Status_Nota": "99 Encerrado",
        "Export_status": "-",
    })

    assert relatorios._executada(row) is False


def test_relatorios_preserva_fallbacks_existentes():
    from input_module import relatorios

    status_local = pd.Series({
        "Status_Final": "-",
        "Status_Nota": "99 Encerrado",
        "Export_status": "-",
    })
    status_textual_sap = pd.Series({
        "Status_Final": "ENCE EXEC",
        "Status_Nota": "10 Em planejamento",
        "Export_status": "ENCE EXEC",
    })

    assert relatorios._executada(status_local) is True
    assert relatorios._executada(status_textual_sap) is True


def _dashboard_nota_status_final(
    status_final,
    encerramento,
    regional="Guarulhos",
):
    from input_module import relatorios

    df_notas = pd.DataFrame([{
        "Numero_Nota": 9001,
        "Conjunto": "POSTES - CAPEX",
        "Planejado_DDPM": 2.0,
        "Mes_Execucao_Planejado": "jul-2026",
        "Regional": regional,
        "Regional_CSD": regional,
        "Status_Final": status_final,
        "Status_Nota": "10 Em planejamento",
        "Export_status": "-",
        "Encerram.por data": encerramento,
    }])
    _, df_ramal, df_metas, df_depara, df_postergacoes = _fx_relatorios()

    return relatorios.montar_dashboard(
        df_notas,
        df_ramal.iloc[0:0],
        df_metas,
        df_depara,
        df_postergacoes.iloc[0:0],
        ano=2026,
        mes_referencia=7,
        regional=None,
    )


def test_dashboard_status_final_99_usa_mes_real():
    dashboard = _dashboard_nota_status_final(99, "2026-08-03")

    julho = dashboard["mensalizacao"][6]
    agosto = dashboard["mensalizacao"][7]
    assert julho["executado"] == 0.0
    assert agosto["executado"] == 2.0
    assert dashboard["avisos"]["executadas_sem_data"] == 0


def test_dashboard_status_final_99_sem_data_usa_mes_planejado_e_avisa():
    dashboard = _dashboard_nota_status_final("99 Encerrado", None)

    julho = dashboard["mensalizacao"][6]
    assert julho["executado"] == 2.0
    assert dashboard["avisos"]["executadas_sem_data"] == 1


def test_dashboard_aviso_sem_data_respeita_filtro_regional():
    dashboard = _dashboard_nota_status_final(
        "99 Encerrado",
        None,
        regional="Mogi das Cruzes",
    )
    from input_module import relatorios
    _, df_ramal, df_metas, df_depara, df_postergacoes = _fx_relatorios()
    nota_sem_data = pd.DataFrame([{
        "Numero_Nota": 9002,
        "Conjunto": "POSTES - CAPEX",
        "Planejado_DDPM": 1.0,
        "Mes_Execucao_Planejado": "jul-2026",
        "Regional": "Mogi das Cruzes",
        "Regional_CSD": "Mogi das Cruzes",
        "Status_Final": "99",
        "Status_Nota": "10 Em planejamento",
        "Export_status": "-",
        "Encerram.por data": None,
    }])
    filtrado = relatorios.montar_dashboard(
        nota_sem_data,
        df_ramal.iloc[0:0],
        df_metas,
        df_depara,
        df_postergacoes.iloc[0:0],
        ano=2026,
        mes_referencia=7,
        regional="Guarulhos",
    )

    assert dashboard["avisos"]["executadas_sem_data"] == 1
    assert filtrado["avisos"]["executadas_sem_data"] == 0


def _fx_relatorios():
    """Fixtures mínimas para o dashboard: 2 notas + 1 ramal + metas/depara."""
    df_notas = pd.DataFrame([
        # carteira jul/2026, Guarulhos (via Regional_CSD), POSTES: 2 und, uma executada (99)
        {"Numero_Nota": 1, "Conjunto": "POSTES - CAPEX", "Planejado_DDPM": 1.0,
         "Mes_Execucao_Planejado": "jul-2026", "Regional": "Mogi das Cruzes",
         "Regional_CSD": "Guarulhos", "Status_Nota": "99 Encerrado",
         "Export_status": "-", "Encerram.por data": "2026-07-10"},
        {"Numero_Nota": 2, "Conjunto": "POSTES - CAPEX", "Planejado_DDPM": 1.0,
         "Mes_Execucao_Planejado": "jul-2026", "Regional": "Guarulhos",
         "Regional_CSD": "-", "Status_Nota": "10 Em planejamento",
         "Export_status": "ENCE EXEC", "Encerram.por data": "2026-08-02"},
        # conjunto fora do de-para -> balde Outros
        {"Numero_Nota": 3, "Conjunto": "MISTERIOSO", "Planejado_DDPM": 2.0,
         "Mes_Execucao_Planejado": "jan-2026", "Regional": "Guarulhos",
         "Regional_CSD": "Guarulhos", "Status_Nota": "01 Sem providência",
         "Export_status": "-", "Encerram.por data": None},
    ])
    df_ramal = pd.DataFrame([
        # prefixo 160 (Poá) -> Poa/Suzano
        {"Numero_Nota": 9, "Local_Instalacao": "160RL00000001", "Planejado_DDPM": 1.0,
         "Mes_Execucao_Planejado": "jul-2026", "Status_Nota": "ENCE EXEC"},
    ])
    df_metas = pd.DataFrame([
        {"Ano": 2026, "Mes": 7, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Meta": 4.0},
        {"Ano": 2026, "Mes": 7, "Regional": "Poa/Suzano", "Plano": "RAMAL", "Meta": 2.0},
    ])
    df_depara = pd.DataFrame([
        {"Plano": "POSTES - CAPEX", "Nome_Curto": "POSTE", "Unidade": "Und.",
         "Area": "Construção", "Modular_RS": 10.0, "Ordem_Exibicao": 1},
        {"Plano": "RAMAL", "Nome_Curto": "RAMAL", "Unidade": "Ponto",
         "Area": "CSD", "Modular_RS": 2.0, "Ordem_Exibicao": 2},
    ])
    df_postergacoes = pd.DataFrame([
        {"Ano": 2026, "Mes": 7, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Qtd": 2.0},
        {"Ano": 2026, "Mes": 3, "Regional": "Guarulhos", "Plano": "POSTES - CAPEX", "Qtd": 5.0},
        {"Ano": 2026, "Mes": 7, "Regional": "Poa/Suzano", "Plano": "RAMAL", "Qtd": 9.0},
    ])
    return df_notas, df_ramal, df_metas, df_depara, df_postergacoes


def test_dashboard_agregacao_basica(banco_temporario):
    from input_module import relatorios
    d = relatorios.montar_dashboard(*_fx_relatorios(), ano=2026, mes_referencia=7, regional=None)

    # hero de julho: carteira POSTES 2 + RAMAL 1 = 3; meta 4+2=6; executado jul = 1 (nota 1; a nota 2 encerra em ago)
    assert d["hero"]["carteira"] == 3.0
    assert d["hero"]["meta"] == 6.0
    assert d["hero"]["executado"] == 1.0
    assert round(d["hero"]["pct_disp"], 3) == 0.5
    assert d["hero"]["carteira_rs"] == 2 * 10.0 + 1 * 2.0
    assert d["hero"]["postergadas"] == 11.0            # jul: POSTES 2 + RAMAL 9

    anual = {l["plano"]: l for l in d["visao_anual"]}
    assert anual["POSTES - CAPEX"]["area"] == "Construção"
    assert anual["POSTES - CAPEX"]["carteira"] == 2.0
    assert anual["POSTES - CAPEX"]["saldo"] == -2.0
    assert anual["POSTES - CAPEX"]["postergado"] == 7.0  # ano: 2 + 5
    assert anual["RAMAL"]["carteira"] == 1.0
    assert anual["RAMAL"]["postergado"] == 9.0
    assert anual["MISTERIOSO"]["area"] == "Outros"        # nunca some silenciosamente
    assert anual["MISTERIOSO"]["pct_disp"] is None        # meta 0 -> null

    assert len(d["mensalizacao"]) == 12
    jul = next(m for m in d["mensalizacao"] if m["mes"] == 7)
    assert jul["carteira"] == 3.0 and jul["executado"] == 1.0

    regs = {r["regional"]: r for r in d["regionais"]}
    assert len(regs) == 6
    assert regs["Guarulhos"]["carteira"] == 2.0           # Regional_CSD + fallback Regional
    assert regs["Poa/Suzano"]["carteira"] == 1.0          # ramal 160 -> Poa/Suzano


def test_dashboard_filtro_regional(banco_temporario):
    from input_module import relatorios
    d = relatorios.montar_dashboard(*_fx_relatorios(), ano=2026, mes_referencia=7,
                                    regional="Guarulhos")
    assert d["hero"]["carteira"] == 2.0                   # só POSTES; ramal era Poa/Suzano
    assert d["hero"]["meta"] == 4.0
    assert d["hero"]["postergadas"] == 2.0                # só Guarulhos, jul
    regs = {r["regional"]: r for r in d["regionais"]}
    assert len(regs) == 6                                 # bloco regionais não filtra


def test_dashboard_mes_referencia_muda_hero(banco_temporario):
    from input_module import relatorios
    d = relatorios.montar_dashboard(*_fx_relatorios(), ano=2026, mes_referencia=3, regional=None)
    assert d["hero"]["mes_nome"] == "março"
    assert d["hero"]["carteira"] == 0.0                    # nenhuma carteira em março
    assert d["mes_referencia"] == 3


def test_api_relatorios_dashboard(banco_temporario, monkeypatch, tmp_path):
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
    from input_module.routes import router
    arquivo = tmp_path / "Controle.xlsx"
    _xlsx_controle(arquivo)
    monkeypatch.setenv("CONTROLE_RECOMPOSICAO_PATH", str(arquivo))
    app = FastAPI(); app.include_router(router)
    client = TestClient(app)

    r = client.get("/api/input/relatorios/dashboard")
    assert r.status_code == 200
    corpo = r.json()
    assert {"hero", "visao_anual", "mensalizacao", "regionais",
            "financeiro_ano", "metas_info", "regionais_disponiveis"} <= set(corpo)
    assert corpo["metas_info"]["erro"] is None
    assert len(corpo["regionais_disponiveis"]) == 6
    assert "postergadas" in corpo["hero"]
    assert all("postergado" in l for l in corpo["visao_anual"])
    etag = r.headers["etag"]
    assert client.get("/api/input/relatorios/dashboard",
                      headers={"If-None-Match": etag}).status_code == 304
    # filtro por regional aceito
    assert client.get("/api/input/relatorios/dashboard?regional=Guarulhos").status_code == 200

    # mes fora do intervalo 1..12 -> 422
    r_invalido = client.get("/api/input/relatorios/dashboard?mes=13")
    assert r_invalido.status_code == 422

    # mes=3 muda o hero para março e o ETag reflete o mes
    r_marco = client.get("/api/input/relatorios/dashboard?mes=3")
    assert r_marco.status_code == 200
    assert r_marco.json()["hero"]["mes_nome"] == "março"
    assert r_marco.headers["etag"] != etag

    r = client.post("/api/input/metas/sincronizar")
    assert r.status_code == 200 and "sincronizou" in r.json()


def test_criar_notas_grava_origem(banco_temporario):
    from input_module import db, service
    nota = service.NovaNota(Numero_Nota=778001, Status_Nota="01 Sem providência",
                            Prioridade_Nota="Programável", Local_Instalacao="045BF00000123")
    service.criar_notas([nota], usuario="teste", origem="carteira")
    conn = db.get_db_connection()
    row = conn.execute("SELECT origem FROM notas WHERE Numero_Nota=778001").fetchone()
    conn.close()
    assert row[0] == "carteira"


def test_criar_notas_origem_default_manual(banco_temporario):
    from input_module import db, service
    nota = service.NovaNota(Numero_Nota=778002, Status_Nota="01 Sem providência",
                            Prioridade_Nota="Programável")
    service.criar_notas([nota], usuario="teste")
    conn = db.get_db_connection()
    row = conn.execute("SELECT origem FROM notas WHERE Numero_Nota=778002").fetchone()
    conn.close()
    assert row[0] == "manual"


# ── Perfil de execução: local x produção ─────────────────────────────────────
def test_perfil_padrao_e_local(monkeypatch):
    monkeypatch.delenv("EDP_PERFIL", raising=False)
    assert config.perfil() == config.PERFIL_LOCAL
    assert config.em_producao() is False


def test_perfil_producao_reconhecido(monkeypatch):
    monkeypatch.setenv("EDP_PERFIL", "PRODUCAO")
    assert config.perfil() == config.PERFIL_PRODUCAO
    assert config.em_producao() is True


def _liberar_copia_excel_rede(monkeypatch):
    """Remove as envs que fazem gerar_copia_excel_rede() sair antes do corpo.

    Sem isso o teste passaria pelo atalho de ambiente de teste e nunca
    exercitaria a guarda de perfil.
    """
    monkeypatch.delenv("PYTEST_CURRENT_TEST", raising=False)
    monkeypatch.delenv("INPUT_DATA_DIR", raising=False)


def test_copia_excel_rede_nao_toca_rede_no_perfil_local(monkeypatch):
    """Perfil local sai antes de enriquecer dados e de qualquer caminho de rede."""
    from input_module import engine
    _liberar_copia_excel_rede(monkeypatch)
    monkeypatch.delenv("EDP_PERFIL", raising=False)
    chamadas = []
    monkeypatch.setattr(engine, "enriquecer_dados",
                        lambda: chamadas.append("enriquecer_dados"))

    engine.gerar_copia_excel_rede()

    assert chamadas == []


def test_copia_excel_rede_segue_no_perfil_producao(monkeypatch):
    """A guarda é do perfil local — produção continua gerando a cópia."""
    from input_module import engine
    _liberar_copia_excel_rede(monkeypatch)
    monkeypatch.setenv("EDP_PERFIL", "producao")
    chamadas = []

    def _enriquecer_e_parar():
        chamadas.append("enriquecer_dados")
        raise RuntimeError("teste para antes de escrever na rede")

    monkeypatch.setattr(engine, "enriquecer_dados", _enriquecer_e_parar)

    engine.gerar_copia_excel_rede()

    assert chamadas == ["enriquecer_dados"]


def test_perfil_local_usa_banco_do_data_dir(monkeypatch, tmp_path):
    monkeypatch.delenv("EDP_PERFIL", raising=False)
    monkeypatch.delenv("INPUT_DB_PATH", raising=False)
    monkeypatch.setenv("INPUT_DATA_DIR", str(tmp_path))
    assert config.caminho_banco_notas() == str(tmp_path / "notas_departamento.db")


def test_perfil_producao_usa_banco_da_rede(monkeypatch, tmp_path):
    monkeypatch.setenv("EDP_PERFIL", "producao")
    monkeypatch.delenv("INPUT_DB_PATH", raising=False)
    monkeypatch.setenv("INPUT_DATA_DIR", str(tmp_path))
    assert config.caminho_banco_notas() == config.REDE_DB_ORIGEM


def test_input_db_path_vence_o_perfil(monkeypatch, tmp_path):
    alvo = tmp_path / "compartilhado.db"
    monkeypatch.setenv("EDP_PERFIL", "producao")
    monkeypatch.setenv("INPUT_DB_PATH", str(alvo))
    assert config.caminho_banco_notas() == str(alvo)


def test_producao_sem_banco_acessivel_falha_alto(monkeypatch, tmp_path):
    """Produção sem rede levanta erro — jamais cai no banco local silenciosamente."""
    from input_module import db
    monkeypatch.setenv("EDP_PERFIL", "producao")
    monkeypatch.setenv("INPUT_DB_PATH", str(tmp_path / "inexistente.db"))
    with pytest.raises(db.BancoRedeIndisponivelErro) as erro:
        db.migrar_da_rede_se_preciso()
    mensagem = str(erro.value)
    assert "EDP_PERFIL=producao" in mensagem
    assert "inexistente.db" in mensagem          # nome lógico do banco: útil
    assert str(tmp_path) not in mensagem         # diretório completo: não vaza


def test_producao_com_banco_acessivel_nao_copia_nada(monkeypatch, tmp_path):
    from input_module import db
    compartilhado = tmp_path / "compartilhado.db"
    monkeypatch.setenv("INPUT_DB_PATH", str(compartilhado))
    monkeypatch.delenv("EDP_PERFIL", raising=False)
    db.inicializar_banco()
    monkeypatch.setenv("EDP_PERFIL", "producao")
    assert db.migrar_da_rede_se_preciso() == "rede"
    assert not (tmp_path / "notas_departamento.db").exists()


def test_mascarar_caminho_nao_expoe_host_nem_diretorio():
    mascarado = config.mascarar_caminho(config.REDE_DB_ORIGEM)
    assert "notas_departamento.db" in mascarado
    assert "ebeat-fp1" not in mascarado
    assert "Diretoria Tecnica" not in mascarado
    assert config.mascarar_caminho(r"C:\dados\notas.db").startswith("local:")


def test_descrever_conexao_resume_sem_caminho_completo(monkeypatch, tmp_path):
    from input_module import db
    monkeypatch.delenv("EDP_PERFIL", raising=False)
    monkeypatch.delenv("INPUT_DB_PATH", raising=False)
    monkeypatch.setenv("INPUT_DATA_DIR", str(tmp_path))
    db.inicializar_banco()
    db.salvar_em_massa(pd.DataFrame([_nota(9100)]))
    resumo = db.descrever_conexao()
    assert resumo["ambiente"] == "local"
    assert resumo["tipo"] == "sqlite"
    assert resumo["status"] == "ok"
    assert resumo["qtd_notas"] == 1
    assert str(tmp_path) not in resumo["alvo"]


def test_rede_raiz_respeita_env(monkeypatch):
    """INPUT_REDE_RAIZ redireciona todos os caminhos derivados da rede."""
    import importlib
    monkeypatch.setenv("INPUT_REDE_RAIZ", r"\outro-host\Compartilhado")
    try:
        importlib.reload(config)
        assert config.REDE_RAIZ == r"\outro-host\Compartilhado"
        assert config.REDE_DB_ORIGEM.startswith(r"\outro-host\Compartilhado")
        assert config.CAMINHO_BASE_IW28.startswith(r"\outro-host\Compartilhado")
    finally:
        monkeypatch.delenv("INPUT_REDE_RAIZ", raising=False)
        importlib.reload(config)
    assert "outro-host" not in config.REDE_RAIZ


# ── Ramal: idempotência do ID_Cronologia ─────────────────────────────────────
def test_ramal_lote_parcial_preserva_id_cronologia(banco_temporario):
    """Reprocessar uma nota não renumera a cronologia nem colide com as outras."""
    from input_module import db
    db.salvar_ramal_em_massa(pd.DataFrame(
        [_nota_ramal(5201), _nota_ramal(5202), _nota_ramal(5203)]))
    antes = dict(zip(db.carregar_dados_ramal()["Numero_Nota"],
                     db.carregar_dados_ramal()["ID_Cronologia"]))
    assert len(set(antes.values())) == 3  # cronologias distintas

    db.salvar_ramal_em_massa(pd.DataFrame([_nota_ramal(5203, Observacao="editada")]))
    depois = dict(zip(db.carregar_dados_ramal()["Numero_Nota"],
                      db.carregar_dados_ramal()["ID_Cronologia"]))
    assert depois == antes
    assert len(db.carregar_dados_ramal()) == 3


def test_ramal_nota_nova_continua_a_numeracao(banco_temporario):
    from input_module import db
    db.salvar_ramal_em_massa(pd.DataFrame([_nota_ramal(5301), _nota_ramal(5302)]))
    db.salvar_ramal_em_massa(pd.DataFrame([_nota_ramal(5303)]))
    df = db.carregar_dados_ramal()
    cronologias = sorted(int(x) for x in df["ID_Cronologia"])
    assert cronologias == [1, 2, 3]


# ── Notificações Diárias aos Engenheiros ───────────────────────────────────────
def test_notificacoes_resumo_diario_por_regional_e_engenheiro(banco_temporario):
    from input_module import db, service, notificacoes_service
    import datetime

    hoje = datetime.date.today().isoformat()
    # 1. Cria notas em diferentes regionais (Guarulhos -> James, Suzano -> Danilo)
    n1 = service.NovaNota(Numero_Nota=9001, Regional="Guarulhos", Status_Nota="00 Pendente", Prioridade_Nota="Programável", Local_Instalacao="045RL00000001")
    n2 = service.NovaNota(Numero_Nota=9002, Regional="Suzano", Status_Nota="00 Pendente", Prioridade_Nota="Programável", Local_Instalacao="050RL00000001")
    service.criar_notas([n1, n2], usuario="felip")

    # 2. Aplica edição em 9001
    db.aplicar_edicoes([{"Numero_Nota": 9001, "Planejado_DDPM": 3.5}], usuario="felip")

    # 3. Consulta resumo diário
    resumo = notificacoes_service.obter_resumo_alteracoes_diarias(data_referencia=hoje)
    assert resumo["total_alteracoes"] >= 3  # criações + edição
    assert "James" in resumo["engenheiros"]
    assert "Danilo" in resumo["engenheiros"]

    james_data = resumo["engenheiros"]["James"]
    assert james_data["total_alteracoes"] >= 2  # criação + edição da nota de Guarulhos
    assert 9001 in james_data["notas_afetadas"]


def test_notificacoes_api_e_emails_responsaveis(cliente):
    # Teste de endpoints de e-mails dos responsáveis
    r_get = cliente.get("/api/input/responsaveis/emails")
    assert r_get.status_code == 200
    emails_iniciais = r_get.json()
    assert "James" in emails_iniciais

    r_put = cliente.put(
        "/api/input/responsaveis/emails",
        json={"James": "james.novo@edp.com", "Danilo": "danilo.novo@edp.com"},
        headers=CABECALHO_USER,
    )
    assert r_put.status_code == 200
    assert cliente.get("/api/input/responsaveis/emails").json()["James"] == "james.novo@edp.com"

    # Teste de endpoint de resumo diário
    r_resumo = cliente.get("/api/input/notificacoes/resumo-diario")
    assert r_resumo.status_code == 200
    assert "engenheiros" in r_resumo.json()


def test_notificacoes_multiplos_engenheiros_mesma_regional(banco_temporario):
    from input_module import db, service, notificacoes_service
    import datetime

    hoje = datetime.date.today().isoformat()
    # Atribui dois engenheiros para Mogi das Cruzes
    db.salvar_responsaveis({"Mogi das Cruzes": "Fabricio, Danilo", "Guarulhos": "James"})
    db.salvar_emails_responsaveis({
        "Fabricio": "fabricio@edp.com",
        "Danilo": "danilo@edp.com",
        "James": "james@edp.com",
    })

    # Cria nota em Mogi das Cruzes (prefixo BIR)
    n = service.NovaNota(
        Numero_Nota=9550,
        Conjunto="MOGI DAS CRUZES",
        Status_Nota="00 Pendente",
        Prioridade_Nota="Programável",
        Local_Instalacao="BIRRL00000001",
    )
    service.criar_notas([n], usuario="teste_user")
    db.aplicar_edicoes([{"Numero_Nota": 9550, "Status_Nota": "10 Concluída"}], usuario="teste_user")

    resumo = notificacoes_service.obter_resumo_alteracoes_diarias(data_referencia=hoje)
    assert "Fabricio" in resumo["engenheiros"]
    assert "Danilo" in resumo["engenheiros"]

    fab_dados = resumo["engenheiros"]["Fabricio"]
    dan_dados = resumo["engenheiros"]["Danilo"]

    # Ambos os engenheiros recebem as alterações da regional compartilhada
    assert 9550 in fab_dados["notas_afetadas"]
    assert 9550 in dan_dados["notas_afetadas"]
    assert fab_dados["total_alteracoes"] >= 2
    assert dan_dados["total_alteracoes"] >= 2


def test_gerar_planilha_alteracoes_anexo():
    import os
    import openpyxl
    from input_module import notificacoes_service

    alteracoes = [
        {
            "Numero_Nota": 123456,
            "Regional": "Mogi das Cruzes",
            "Conjunto": "SUZANO",
            "Circuito": "MOG-01",
            "Tipo_Evento": "Edição de Campo",
            "Campo_Alterado": "Status_Nota",
            "Valor_Antigo": "00 Pendente",
            "Valor_Novo": "10 Concluída",
            "Detalhe": "Status_Nota: '00 Pendente' ➔ '10 Concluída'",
            "Usuario": "felip",
            "Data_Hora": "2026-08-18 11:00:00",
        },
    ]

    caminho_xlsx = notificacoes_service.gerar_planilha_alteracoes_anexo("Fabricio", alteracoes, "2026-08-18")
    assert os.path.exists(caminho_xlsx)
    assert caminho_xlsx.endswith(".xlsx")

    wb = openpyxl.load_workbook(caminho_xlsx)
    assert "Notas Modificadas" in wb.sheetnames
    assert "Resumo Executivo" in wb.sheetnames

    ws_det = wb["Notas Modificadas"]
    assert ws_det.cell(row=2, column=1).value == 123456
    assert ws_det.cell(row=2, column=2).value == "Mogi das Cruzes"

    wb.close()
    try:
        os.remove(caminho_xlsx)
    except Exception:
        pass


def test_sincronizar_status_sap_para_notas(banco_temporario):
    from input_module import db
    # Cadastra notas no banco
    db.salvar_em_massa(pd.DataFrame([
        _nota(7001, Status_Nota="10 Em planejamento"),
        _nota(7002, Status_Nota="51 Ordem Liberada"),
        _nota(7003, Status_Nota="52 ADS e Viabilizado"),
    ]))

    # Cria base_iw28 simulando retorno do SAP onde 7001 foi para 51 e 7003 foi para 99
    df_iw28 = pd.DataFrame([
        {"Nota": 7001, "Status usuário": "51 Ordem Liberada"},
        {"Nota": 7002, "Status usuário": "51 Ordem Liberada"},
        {"Nota": 7003, "Status usuário": "99 Encerrado"},
    ])
    db.salvar_base_dataframe("base_iw28", df_iw28)

    res = db.sincronizar_status_sap_para_notas(usuario="Robô SAP")
    assert res["atualizadas"] == 2

    # Verifica se notas foram atualizadas
    df_banco = db.carregar_dados()
    linha_7001 = df_banco[df_banco["Numero_Nota"] == 7001].iloc[0]
    assert "51" in str(linha_7001["Status_Nota"])
    assert "10" in str(linha_7001["Status_Anterior"])

    linha_7003 = df_banco[df_banco["Numero_Nota"] == 7003].iloc[0]
    assert "99" in str(linha_7003["Status_Nota"])
    assert "52" in str(linha_7003["Status_Anterior"])

    # Verifica se logs foram gerados
    logs = db.carregar_logs()
    logs_7001 = logs[logs["Numero_Nota"] == 7001]
    assert len(logs_7001) >= 1
    assert logs_7001.iloc[0]["Usuario"] == "Robô SAP"
    assert logs_7001.iloc[0]["Campo_Alterado"] == "Status_Nota"
    assert "10" in str(logs_7001.iloc[0]["Valor_Antigo"])
    assert "51" in str(logs_7001.iloc[0]["Valor_Novo"])


def test_endpoint_sincronizar_status_sap(cliente):
    from input_module import db
    db.salvar_em_massa(pd.DataFrame([
        _nota(8001, Status_Nota="10 Em planejamento"),
    ]))
    db.salvar_base_dataframe("base_iw28", pd.DataFrame([
        {"Nota": 8001, "Status usuário": "51 Ordem Liberada"},
    ]))

    resp = cliente.post("/api/input/bases/sincronizar-status-sap", headers={"X-User": "admin_teste"})
    assert resp.status_code == 200
    dados = resp.json()
    assert dados["atualizadas"] >= 1

    df_banco = db.carregar_dados()
    linhas = df_banco[df_banco["Numero_Nota"] == 8001]
    if not linhas.empty:
        assert "51" in str(linhas.iloc[0]["Status_Nota"])


def test_extrair_data_sap_e_comparacao(banco_temporario):
    from input_module import db, engine
    assert engine.extrair_data_sap("M08/2026-POSTE-GUR1302") == "ago-2026"
    assert engine.extrair_data_sap("M04/24-REDE COMPACTA TRIF") == "abr-2024"
    assert engine.extrair_data_sap("M00/0000-A PLANEJAR") == "-"
    assert engine.extrair_data_sap(None) == "-"

    # Testa enriquecimento do dataset completo
    db.salvar_em_massa(pd.DataFrame([
        _nota(9001, Mes_Execucao_Planejado="ago-2026"),
        _nota(9002, Mes_Execucao_Planejado="2026-08-01"),
        _nota(9003, Mes_Execucao_Planejado="jul-2026"),
    ]))
    db.salvar_base_dataframe("base_iw28", pd.DataFrame([
        {"Nota": 9001, "Descrição": "M08/2026-POSTE-GUR1302", "Status usuário": "10 Aberto", "Data da nota": "2026-01-15"},
        {"Nota": 9002, "Descrição": "M08/2026-POSTE-GUR1302", "Status usuário": "10 Aberto", "Data da nota": "2026-01-15"},
        {"Nota": 9003, "Descrição": "M08/2026-POSTE-GUR1302", "Status usuário": "10 Aberto", "Data da nota": "2026-01-15"},
    ]))

    df = engine.get_dataset()
    r9001 = df[df["Numero_Nota"] == 9001].iloc[0]
    r9002 = df[df["Numero_Nota"] == 9002].iloc[0]
    r9003 = df[df["Numero_Nota"] == 9003].iloc[0]

    assert r9001["Data_programada_SAP"] == "ago-2026"
    assert r9001["Comparacao_Data_SAP"] == "Igual"
    assert r9002["Comparacao_Data_SAP"] == "Igual"
    assert r9003["Comparacao_Data_SAP"] == "Divergente"

