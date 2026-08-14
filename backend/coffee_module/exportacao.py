"""Geração de planilhas para a seção COFFEE."""
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


_CABECALHOS = (
    "ID ONR",
    "ID SAP",
    "Classificação",
    "Local de instalação",
    "Poste(s)",
    "Referência",
    "Componente",
    "Sintoma",
    "Observação",
    "Origem",
    "Concluída em",
)


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    return f"'{texto}" if texto.startswith(("=", "+", "-", "@")) else texto


def _primeiro(dados: dict[str, Any], *campos: str) -> str:
    for campo in campos:
        valor = _texto(dados.get(campo))
        if valor:
            return valor
    return ""


def _local_instalacao(dados: dict[str, Any]) -> str:
    direto = _primeiro(dados, "local_instalacao")
    if direto:
        return direto
    partes = [
        _primeiro(dados, "cidade"),
        _primeiro(dados, "tipo_local_instalacao"),
        _primeiro(dados, "local_instalacao_numero"),
    ]
    return "-".join(parte for parte in partes if parte)


def _linha(nota: dict[str, Any]) -> tuple[Any, ...]:
    dados = nota.get("dados_json") or {}
    classificacao = _texto(nota.get("classificacao"))
    return (
        nota.get("pk"),
        nota.get("id_sap"),
        classificacao.capitalize(),
        _local_instalacao(dados),
        _primeiro(dados, "postes", "poste"),
        _primeiro(dados, "referencia_fisica", "referencia_eletrica", "referencia"),
        _primeiro(dados, "componente", "componente_novo"),
        _primeiro(dados, "sintoma"),
        _primeiro(dados, "observacao", "observacoes"),
        _texto(nota.get("origem")),
        _texto(nota.get("corrigida_em") or nota.get("classificacao_em")),
    )


def gerar_planilha_concluidas(notas: list[dict[str, Any]]) -> bytes:
    """Gera uma planilha XLSX autocontida para o conjunto filtrado de concluídas."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Notas concluídas"
    worksheet.append(_CABECALHOS)
    for nota in notas:
        worksheet.append(_linha(nota))

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, heading in enumerate(_CABECALHOS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(14, min(42, len(heading) + 4))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
