"""Motor de enriquecimento de dados do módulo Input.

Porte de Input/processamento.py (função ``puxar_dados_completos_da_rede``
renomeada para ``enriquecer_dados``) e da auditoria ``avaliar_prazo_sap``
(Input/app.py). Sem dependência de Streamlit.

Diferenças relevantes em relação ao porte original:
- Caminhos de rede lidos de ``config`` em TEMPO DE CHAMADA (``config.CAMINHO_*``),
  para que o monkeypatch dos testes tenha efeito.
- ``st.error(...)`` vira ``print(...)``.
- Cache em memória de ``get_dataset`` revalidado por versão do dataset
  (``db.obter_versao_dataset()``), com TTL de fallback e lock; ``status_bases``
  com seu próprio TTL e lock; e ``gerar_copia_excel_rede`` (que nunca derruba
  a request: corpo inteiro em try/except).
"""
import concurrent.futures
import datetime
import os
import re
import threading
import time
from typing import Any

import numpy as np
import pandas as pd

from input_module import config, db
from input_module.db import carregar_dados, carregar_projeto_construcao


import unicodedata

meses_pt_rev = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "maio": 5, "jun": 6,
                "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}


def normalizar_conjunto(texto: object) -> str:
    """
    Padroniza nomes de conjuntos para cruzamentos de dados (PROCV / Map), eliminando:
    - Quebras de linha e múltiplos espaços
    - Espaçamento inconsistente em hífens e barras (ex: ' - ', '-', ' -', '- ')
    - Variações de acentos (NFKD -> ASCII)
    """
    if pd.isna(texto) or texto is None:
        return "-"
    s = str(texto).strip()
    if not s or s in ["-", "nan", "None", "NAN", "NONE", "<NA>"]:
        return "-"
    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
    s = re.sub(r'\s*-\s*', ' - ', s)
    s = re.sub(r'\s*/\s*', ' / ', s)
    return " ".join(s.split()).upper().strip()


def limpar_exibicao_conjunto(texto: object) -> str:
    """Limpa quebras de linha e espaços múltiplos do Conjunto preservando a grafia."""
    if pd.isna(texto) or texto is None:
        return "-"
    s = str(texto).strip()
    if not s or s in ["-", "nan", "None", "NAN", "NONE", "<NA>"]:
        return "-"
    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r'\s*-\s*', ' - ', s)
    return " ".join(s.split()).strip()


def normalizar_prioridade_sap(val):
    if pd.isna(val) or str(val).strip() in ["", "-", "nan", "None"]:
        return None
    val_str = str(val).strip().lower()
    
    val_str = ''.join(c for c in unicodedata.normalize('NFD', val_str) if unicodedata.category(c) != 'Mn')
    
    if val_str.startswith("1"): return "Emergente"
    if val_str.startswith("2"): return "Urgente"
    if val_str.startswith("3"): return "Importante"
    if val_str.startswith("4"): return "Prioritário"
    if val_str.startswith("5"): return "Programável"
    if val_str.startswith("6"): return "Informativo"
    
    if "emerg" in val_str: return "Emergente"
    if "urg" in val_str: return "Urgente"
    if "imp" in val_str: return "Importante"
    if "prio" in val_str: return "Prioritário"
    if "prog" in val_str: return "Programável"
    if "info" in val_str: return "Informativo"
    if "prot" in val_str: return "Protheus"
    if "proj" in val_str: return "Nota Projetos"
    
    return None


def normalizar_status_sap(val):
    if pd.isna(val) or str(val).strip() in ["", "-", "nan", "None", "Fora SAP", "Pendente Extração SAP", "Erro na leitura"]:
        return None
    val_str = str(val).strip()
    
    match = re.match(r'^(\d+)', val_str)
    if match:
        num = int(match.group(1))
        from input_module.config import STATUS_MAP
        if num in STATUS_MAP:
            return STATUS_MAP[num]
            
    val_upper = val_str.upper()
    if "SUPR CANC" in val_upper or "ENCE CANC" in val_upper: return "55 Cancelado"
    if "SUPR" in val_upper: return "SUPR"
    if "ENCE EXEC" in val_upper: return "ENCE EXEC"
    
    return None


def parse_data_encerramento(val: Any) -> datetime.datetime | None:
    """Interpreta datas de encerramento/SAP com precisão evitando o bug de dayfirst=True em formato ISO."""
    if val is None or pd.isna(val):
        return None
    if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)):
        return pd.to_datetime(val)
    s = str(val).strip()
    if s in ["", "-", "None", "nan", "<NA>", "null"]:
        return None

    # 1. Formato ISO YYYY-MM-DD (com ou sem horário) -> SEMPRE Ano-Mês-Dia (NÃO dayfirst)
    m_iso = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m_iso:
        ano, mes, dia = int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3))
        try:
            return datetime.datetime(ano, mes, dia)
        except Exception:
            return None

    # 2. Formato Brasileiro DD/MM/YYYY -> Dia-Mês-Ano
    m_br = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m_br:
        dia, mes, ano = int(m_br.group(1)), int(m_br.group(2)), int(m_br.group(3))
        try:
            return datetime.datetime(ano, mes, dia)
        except Exception:
            return None

    # 3. Fallback
    try:
        dt = pd.to_datetime(s, errors="coerce")
        return None if pd.isna(dt) else dt
    except Exception:
        return None


def extrair_data_sap(descricao: Any) -> str:
    """Extrai o mês e ano programados no SAP a partir do texto da descrição (ex: M04/2026-SUBTERRANEO -> abr-2026)."""
    if pd.isna(descricao) or not descricao:
        return "-"
    desc_str = str(descricao).strip()
    if desc_str in ["", "-", "None", "nan", "<NA>", "null"]:
        return "-"

    # 1. Padrão M04/2026, M04-2026, M 04/2026, 04/2026
    match = re.search(r'M?\s*(\d{2})[/-](\d{4}|\d{2})', desc_str, re.IGNORECASE)
    if match:
        mes_num = int(match.group(1))
        ano_str = match.group(2)
        ano_num = int(ano_str)
        if len(ano_str) == 2:
            ano_num += 2000

        meses_pt = {
            1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr',
            5: 'maio', 6: 'jun', 7: 'jul', 8: 'ago',
            9: 'set', 10: 'out', 11: 'nov', 12: 'dez'
        }
        if 1 <= mes_num <= 12 and 2000 <= ano_num <= 2099:
            return f"{meses_pt[mes_num]}-{ano_num}"

    # 2. Padrão ISO 2026-04 ou 2026/04
    match_iso = re.search(r'(\d{4})[/-](\d{2})', desc_str)
    if match_iso:
        ano_num = int(match_iso.group(1))
        mes_num = int(match_iso.group(2))
        meses_pt = {
            1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr',
            5: 'maio', 6: 'jun', 7: 'jul', 8: 'ago',
            9: 'set', 10: 'out', 11: 'nov', 12: 'dez'
        }
        if 1 <= mes_num <= 12 and 2000 <= ano_num <= 2099:
            return f"{meses_pt[mes_num]}-{ano_num}"

    return "-"



def _ler_export_medidas() -> pd.DataFrame | None:
    return db.carregar_base_dataframe("base_iw66")


def _comparar_medida_planejado(medida_str: str, planejado_val) -> str:
    medida_str = str(medida_str).strip().replace(",", ".")
    if pd.isna(planejado_val) or medida_str == "-":
        return "-"
    try:
        planejado_val = float(planejado_val)
    except (TypeError, ValueError):
        return "-"

    km_val, un_val = 0.0, 0.0
    has_km, has_un = False, False
    m = re.search(r"([\d.]+)\s*km", medida_str.lower())
    if m:
        try:
            km_val = float(m.group(1))
            has_km = True
        except ValueError:
            pass
    m = re.search(r"([\d.]+)\s*un", medida_str.lower())
    if m:
        try:
            un_val = float(m.group(1))
            has_un = True
        except ValueError:
            pass

    if not has_km and not has_un:
        return "-"

    match_km = has_km and (
        abs(km_val * 1000.0 - planejado_val) < 0.1 or abs(km_val - planejado_val) < 0.1
    )
    match_un = has_un and abs(un_val - planejado_val) < 0.1

    if has_km and has_un:
        return "Sim" if (match_km or match_un) else "Não"
    if has_km:
        return "Sim" if match_km else "Não"
    return "Sim" if match_un else "Não"


# --- FUNÇÃO DE REGRA DE NEGÓCIO: CONJUNTO CRÍTICO ---
# Avalia a criticidade do conjunto com base no Delta do Indicador (12 meses)
def regra_conjunto_critico(valor):
    if pd.isna(valor): return "-"
    elif valor > -0.5: return "Violado"
    elif valor == -0.5: return "Crítico"
    else: return 'Dentro'


# ====================================================================
# AUDITORIA CRONOLÓGICA (DDPM vs SAP) — porte de Input/app.py:925
# ====================================================================
def avaliar_prazo_sap(row):
    try:
        status_final = str(row.get('Status_Final', ''))
        # Reconciliação: o engine produz Status_Final = Export_status (status SAP).
        # O indicador "Encerrado (99)" da lógica original referia-se ao status da
        # nota; usamos também Status_Nota para detectar o 99 de forma robusta.
        status_nota = str(row.get('Status_Nota', ''))
        ordem_executada = str(row.get('Ordem_Executada', 'NÃO')).strip().upper()
        # Desvio intencional: o original verificava apenas Status_Final; aqui também
        # lemos Status_Nota para reconhecer notas logicamente encerradas (99) que
        # casaram no IW28 com outro status SAP (ex.: LIBE).
        is_99 = ('99' in status_final) or ('99' in status_nota)

        # 1. TRATAMENTO ULTRA-ROBUSTO DO PLANEJADO (DDPM)
        val_plan = str(row.get('Mes_Execucao_Planejado', '')).strip()
        mes_planejado, ano_planejado = None, None

        if val_plan not in ["", "-", "None", "nan"]:
            # Testa se o Pandas autoconverteu o planejamento para data completa (Ex: 2024-02-01 00:00:00)
            match_iso = re.match(r'^(\d{4})[-/](\d{2})[-/](\d{2})', val_plan)
            if match_iso:
                ano_planejado = int(match_iso.group(1))
                mes_planejado = int(match_iso.group(2))
            elif '-' in val_plan:
                partes = val_plan.split('-')
                if partes[0].lower() in meses_pt_rev:
                    mes_planejado = meses_pt_rev[partes[0].lower()]
                    ano_planejado = int(partes[1])
                    if ano_planejado < 100: ano_planejado += 2000
                elif partes[1].lower() in meses_pt_rev:
                    mes_planejado = meses_pt_rev[partes[1].lower()]
                    ano_planejado = int(partes[0])
                    if ano_planejado < 100: ano_planejado += 2000

        hoje = datetime.datetime.now()

        # --- NOVA REGRA 1: PASSÍVEL DE ENCERRAMENTO ---
        # Notas com Ordem_Executada == 'SIM' que não possuem status 99
        if not is_99 and ordem_executada == 'SIM':
            return "⚠️ Passível de Encerramento"

        if not mes_planejado or not ano_planejado:
            if is_99: return "⚠️ Sem Mês Planejado Válido"
            return "⚪ Sem Planejamento"

        # REGRA: Realizada Fora do Plano (Mês Planejado > Atual)
        if ano_planejado > hoje.year:
            return "🟣 Fora do Plano"

        # --- REGRA 2: NOTAS NÃO ENCERRADAS (AVALIAÇÃO DE ATRASO) ---
        if not is_99:
            desvio_hoje = (hoje.year - ano_planejado) * 12 + (hoje.month - mes_planejado)
            if desvio_hoje > 1:
                return "🔴 Pendente Atrasado"
            else:
                return "⚪ Em Andamento (No Prazo)"

        # 2. TRATAMENTO DO REALIZADO (SAP - Ex: 2026-08-03 ou 03/08/2026)
        val_real = row.get('Encerram.por data', '-')
        if pd.isna(val_real) or str(val_real).strip() in ["", "-", "None", "nan", "<NA>", "null"]:
            return "⏳Sem Data SAP"

        dt_real = parse_data_encerramento(val_real)
        if not dt_real:
            return "⚠️ Data SAP Inválida"

        mes_real = dt_real.month
        ano_real = dt_real.year

        # 3. COMPARAÇÃO MATEMÁTICA DO DESVIO
        # Notas adiantadas sempre marcadas como Adiantado (Azul).
        # Notas com atraso de até 1 mês são consideradas No Prazo (Verde).
        # Notas executadas com atraso >= 2 meses são Executado com atraso (Laranja).
        desvio_meses = (ano_real - ano_planejado) * 12 + (mes_real - mes_planejado)
        if desvio_meses < 0:
            return "🔵 Adiantado"
        elif desvio_meses <= 1:
            return "🟢 No Prazo"
        else:
            return "🟠 Executado com atraso"
    except:
        return "⚠️ Erro na Análise"


# ====================================================================
# MOTOR DE DADOS (Carregamento e Cruzamentos)
# ====================================================================
def enriquecer_dados():
    # 3.1. Carrega a base bruta (notas cadastradas) do SQLite
    df = carregar_dados()

    # Tratamento de limpeza do Conjunto (remove quebras de linha e múltiplos espaços)
    if 'Conjunto' in df.columns:
        df['Conjunto'] = df['Conjunto'].apply(limpar_exibicao_conjunto)

    # Tratamento de segurança para garantir que o Status Anterior seja numérico/texto legível
    if 'Status_Anterior' in df.columns:
        df['Status_Anterior'] = pd.to_numeric(df['Status_Anterior'], errors='coerce')
        df['Status_Anterior'] = df['Status_Anterior'].apply(lambda x: str(int(x)) if pd.notna(x) else "-")
    else:
        df['Status_Anterior'] = "-"

    # --- MAPEAMENTOS GEOGRÁFICOS BÁSICOS ---
    # Extrai os primeiros 3 caracteres do Circuito/Local Instalação para cruzar com os dicionários de configuração
    df['Cidade'] = df['Local_Instalacao'].astype(str).str[:3].map(config.DE_PARA_CIDADES).fillna("Desconhecido")
    df['CJ_Aneel'] = df['Circuito'].astype(str).str[:3].map(config.DE_PARA_CJ_ANEEL).fillna("Desconhecido")
    df["substacao_conjunto"] = df['Circuito'].astype(str).str[:3].fillna("Desconhecido") + " - " + df['Conjunto'].astype(str).fillna("Desconhecido")

    # --- 3.2. PROCV DO INDICADOR DE CONTINUIDADE (CRITICIDADE E RANKING) ---
    # Lê a tabela do banco e verifica quais conjuntos estão próximos da violação (Limite ANEEL)
    df_hierarquia = db.carregar_base_dataframe("base_indicador_continuidade")
    if df_hierarquia is not None and not df_hierarquia.empty:
        try:
            df_hierarquia.columns = df_hierarquia.columns.astype(str).str.replace('\n', ' ').str.replace('[', '').str.replace(']', '').str.strip()

            col_alvo = 'DELTA_INDICADOR _12MM_CONJUNTO' if 'DELTA_INDICADOR _12MM_CONJUNTO' in df_hierarquia.columns else 'DELTA_INDICADOR_12MM_CONJUNTO'
            df_hierarquia['DELTA_INDICADOR_12MM_CONJUNTO'] = pd.to_numeric(df_hierarquia[col_alvo], errors='coerce')

            col_nome_conjunto = 'TBL_HIERARQUIA_CONJUNTO CJ_NOME' if 'TBL_HIERARQUIA_CONJUNTO CJ_NOME' in df_hierarquia.columns else ('TBL_HIERARQUIA_CONJUNTO[CJ_NOME]' if 'TBL_HIERARQUIA_CONJUNTO[CJ_NOME]' in df_hierarquia.columns else df_hierarquia.columns[0])

            # Padroniza nomes (remove acentos, espaços) para garantir que o cruzamento de dados (Merge/Map) funcione perfeitamente
            df_hierarquia['Conj.Corrijido'] = df_hierarquia[col_nome_conjunto].astype(str).str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.strip().str.upper()
            df_hierarquia['Conjunto Crítico'] = df_hierarquia['DELTA_INDICADOR_12MM_CONJUNTO'].apply(regra_conjunto_critico)
            df_hierarquia['ranking'] = df_hierarquia['DELTA_INDICADOR_12MM_CONJUNTO'].rank(ascending=False, method='min', na_option='bottom').fillna(99).astype(int)

            mapeamento_conjunto_critico = dict(zip(df_hierarquia['Conj.Corrijido'], df_hierarquia['Conjunto Crítico']))
            mapeamento_ranking = dict(zip(df_hierarquia['Conj.Corrijido'], df_hierarquia['ranking']))

            col_regional_excel = 'Regional_1' if 'Regional_1' in df_hierarquia.columns else 'Regional'
            if col_regional_excel in df_hierarquia.columns:
                mapeamento_regional = dict(zip(df_hierarquia['Conj.Corrijido'], df_hierarquia[col_regional_excel]))
            else:
                mapeamento_regional = {}
                print("Aviso: Coluna Regional_1 não encontrada na planilha de indicadores.")

            SINONIMOS_CONJUNTO = {"CARAGUATATUBA":"CARAGUA", "FERRAZ":"FERRAZ DE VASCONCELOS", "BRAS CUBAS": "BRAZ CUBAS"}
            chave_busca_base = df['CJ_Aneel'].astype(str).str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.strip().str.upper().replace(SINONIMOS_CONJUNTO)

            df['Conj.critico'] = chave_busca_base.map(mapeamento_conjunto_critico).fillna("-")
            df['ranking'] = chave_busca_base.map(mapeamento_ranking).fillna(0).astype(int)

        except Exception as e:
            df['Conj.critico'] = "-"
            df['ranking'] = 0
            print(f"Erro ao ler Indicador de Continuidade: {e}")
    else:
        df['Conj.critico'] = "-"
        df['ranking'] = 0

    # --- 3.3. MAPEAMENTO ESTÁTICO: REGIONAL CSD ---
    chave_busca_regional = df['CJ_Aneel'].astype(str).str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.strip().str.upper()
    df['Regional_CSD'] = chave_busca_regional.map(config.MAP_REGIONAL_CSD).fillna("-")

    # --- 3.4. INTEGRAÇÃO SAP: EXTRAÇÃO IW28 (STATUS E DATAS REAIS) ---
    # Puxa os dados gerados pelo Robô RPA para atualizar o status final e data de encerramento da nota
    df['Centro_Responsavel_Banco'] = df['Centro_Responsavel'].fillna("-")

    df_sap = db.carregar_base_dataframe("base_iw28")
    if df_sap is not None and not df_sap.empty:
        try:
            # Mapeamento resiliente de colunas do IW28 (resistente a variações de acentuação e encodings)
            def _achar_coluna(padroes: list[str]) -> str | None:
                for col in df_sap.columns:
                    col_norm = unicodedata.normalize("NFKD", str(col)).encode("ascii", "ignore").decode("utf-8").lower()
                    if any(p in col_norm for p in padroes):
                        return col
                return None

            col_nota = _achar_coluna(["nota"])
            col_status = _achar_coluna(["status usu", "status ordem", "status"])
            col_centro = _achar_coluna(["centrabalho", "centro p/cen", "centro"])
            col_ordem = _achar_coluna(["ordem"])
            col_encerram = _achar_coluna(["encerram"])
            col_desc = _achar_coluna(["descri"])
            col_prio = _achar_coluna(["prioridade"])
            col_data_nota = _achar_coluna(["data da nota", "data nota"])

            # Normaliza Nota do SAP como string numérica pura para cruzamento seguro
            if col_nota:
                df_sap['Nota_str'] = pd.to_numeric(df_sap[col_nota], errors='coerce').dropna().astype(int).astype(str).str.strip()
            else:
                df_sap['Nota_str'] = df_sap.index.astype(str)
            df['Numero_Nota_str'] = pd.to_numeric(df['Numero_Nota'], errors='coerce').fillna(0).astype(int).astype(str).str.strip()

            if col_status:
                dicionario_status_sap = dict(zip(df_sap['Nota_str'], df_sap[col_status]))
                df['Export_status'] = df['Numero_Nota_str'].map(dicionario_status_sap).fillna("Fora SAP")
            else:
                df['Export_status'] = "Fora SAP"

            if col_centro:
                dicionario_centro_sap = dict(zip(df_sap['Nota_str'], df_sap[col_centro]))
                df['Centro_SAP'] = df['Numero_Nota_str'].map(dicionario_centro_sap)
            else:
                df['Centro_SAP'] = None

            # Identifica a qual Ordem aquela Nota pertence para podermos cruzar com o financeiro (IW38)
            if col_ordem:
                df_sap['Ordem_Texto'] = pd.to_numeric(df_sap[col_ordem], errors='coerce').apply(lambda x: str(int(x)) if pd.notna(x) else "-")
                dicionario_ordem_sap = dict(zip(df_sap['Nota_str'], df_sap['Ordem_Texto']))
                df['Ordem'] = df['Numero_Nota_str'].map(dicionario_ordem_sap).fillna("-")
            else:
                df['Ordem'] = "-"

            def _fmt_dt_sap(val):
                if pd.isna(val) or str(val).strip().lower() in ["none", "nan", "-", "", "<na>"]:
                    return "-"
                dt = parse_data_encerramento(val)
                if dt:
                    return dt.strftime("%d/%m/%Y")
                return str(val).strip()

            if col_data_nota:
                dicionario_data_nota = dict(zip(df_sap['Nota_str'], df_sap[col_data_nota]))
                df['Data_Nota_SAP'] = df['Numero_Nota_str'].map(dicionario_data_nota).apply(_fmt_dt_sap)
                df['Data da nota'] = df['Data_Nota_SAP']
            else:
                df['Data_Nota_SAP'] = "-"
                df['Data da nota'] = "-"

            if col_encerram:
                dicionario_encerram_data = dict(zip(df_sap['Nota_str'], df_sap[col_encerram]))
                df['Encerram.por data'] = df['Numero_Nota_str'].map(dicionario_encerram_data).apply(_fmt_dt_sap)
            else:
                df['Encerram.por data'] = "-"

            # Integração da Descrição para extrair a data programada do SAP
            if col_desc:
                dicionario_desc_sap = dict(zip(df_sap['Nota_str'], df_sap[col_desc]))
                df['Descricao_SAP'] = df['Numero_Nota_str'].map(dicionario_desc_sap)
                df['Data programada SAP'] = df['Descricao_SAP'].apply(extrair_data_sap)
            else:
                df['Data programada SAP'] = "-"
            df['Data_programada_SAP'] = df['Data programada SAP']

            def parse_ano_mes(val):
                if val is None or pd.isna(val):
                    return None
                s = str(val).strip().lower()
                if s in ["-", "", "nan", "none", "<na>", "null"]:
                    return None
                m = re.match(r'^(\d{4})[-/](\d{1,2})', s)
                if m:
                    return int(m.group(1)), int(m.group(2))
                m2 = re.search(r'([a-zç]+)[-/\s]+(\d{2,4})', s)
                if m2:
                    m_txt, y_txt = m2.group(1), int(m2.group(2))
                    if y_txt < 100:
                        y_txt += 2000
                    if m_txt in meses_pt_rev:
                        return y_txt, meses_pt_rev[m_txt]
                return None

            def comparar_datas_sap(row):
                dt_sap = parse_ano_mes(row.get('Data programada SAP', row.get('Data_programada_SAP', '-')))
                dt_local = parse_ano_mes(row.get('Mes_Execucao_Planejado', '-'))
                if not dt_sap or not dt_local:
                    return "-"
                if dt_sap == dt_local:
                    return "Igual"
                return "Divergente"

            df['Comparação Data SAP'] = df.apply(comparar_datas_sap, axis=1)
            df['Comparacao_Data_SAP'] = df['Comparação Data SAP']

            # Normalização do Status do SAP para atualizar o Status local
            df['Status_SAP_Norm'] = df['Export_status'].apply(normalizar_status_sap)
            df['Status_Nota'] = df['Status_SAP_Norm'].fillna(df['Status_Nota'])

            # Integração da Prioridade da Nota vinda do SAP
            if col_prio:
                dicionario_prio_sap = dict(zip(df_sap['Nota_str'], df_sap[col_prio]))
                df['Prioridade_SAP'] = df['Numero_Nota_str'].map(dicionario_prio_sap)
                df['Prioridade_SAP_Norm'] = df['Prioridade_SAP'].apply(normalizar_prioridade_sap)
                df['Prioridade_Nota'] = df['Prioridade_SAP_Norm'].fillna(df['Prioridade_Nota'])

            df['Centro_Responsavel'] = df['Centro_SAP'].fillna(df['Centro_Responsavel_Banco']).fillna("-")

        except Exception as e:
            df['Export_status'] = "Erro na leitura"
            df['Centro_Responsavel'] = df['Centro_Responsavel_Banco']
            df['Data_Nota_SAP'] = "-"
            df['Data da nota'] = "-"
            df['Encerram.por data'] = "-"
            df['Data programada SAP'] = "-"
            df['Data_programada_SAP'] = "-"
            df['Comparação Data SAP'] = "-"
            df['Comparacao_Data_SAP'] = "-"
            print(f"Erro ao ler IW28: {e}")
    else:
        df['Export_status'] = "Pendente Extração SAP"
        df['Centro_Responsavel'] = df['Centro_Responsavel_Banco']
        df['Data_Nota_SAP'] = "-"
        df['Data da nota'] = "-"
        df['Encerram.por data'] = "-"
        df['Data programada SAP'] = "-"
        df['Data_programada_SAP'] = "-"
        df['Comparação Data SAP'] = "-"
        df['Comparacao_Data_SAP'] = "-"
        df['Encerram.por data'] = "-"
        df['Data programada SAP'] = "-"
        df['Data_programada_SAP'] = "-"
        df['Comparação Data SAP'] = "-"
        df['Comparacao_Data_SAP'] = "-"

    df = df.drop(columns=['Centro_Responsavel_Banco', 'Descricao_SAP', 'Status_SAP_Norm', 'Prioridade_SAP', 'Prioridade_SAP_Norm', 'Numero_Nota_str'], errors='ignore')
    if 'Centro_SAP' in df.columns:
        df = df.drop(columns=['Centro_SAP'], errors='ignore')

    # Lógica fallback: Se a nota ainda não foi enviada pro SAP, o Status Final reflete o Status local da Engenharia
    df['Status_Final'] = df['Export_status']
    mascara_fora_sap = df['Export_status'] == "Fora SAP"
    df.loc[mascara_fora_sap, 'Status_Final'] = df.loc[mascara_fora_sap, 'Status_Nota']
    df['Status_Usuário_Ordem'] = "-"

    # --- 3.5. PROCV DA QUANTIDADE DE CLIENTES POR CONJUNTO ---
    # Utilizado mais a frente como denominador para calcular o DEC e FEC
    df_clientes = db.carregar_base_dataframe("base_clientes")
    if df_clientes is not None and not df_clientes.empty:
        try:
            col_chave_excel = 'CONJUNTO_DESC'
            col_valor_excel = 'QTDE_CONJUNTO'
            df_clientes[col_chave_excel] = df_clientes[col_chave_excel].astype(str).str.strip().str.upper()

            def converter_clientes_inteiro(valor):
                if pd.isna(valor): return 0
                if isinstance(valor, (int, float)): return int(valor)
                v_str = str(valor).strip().replace('.', '')
                try: return int(v_str)
                except: return 0

            df_clientes[col_valor_excel] = df_clientes[col_valor_excel].apply(converter_clientes_inteiro)
            dict_clientes_dinamico = dict(zip(df_clientes[col_chave_excel], df_clientes[col_valor_excel]))

            chave_busca_regional = df['CJ_Aneel'].astype(str).str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.strip().str.upper()
            df['N_Clientes_Conjunto'] = chave_busca_regional.map(dict_clientes_dinamico).fillna(0).astype(int)

        except Exception as e:
            df['N_Clientes_Conjunto'] = 0
            print(f"Erro ao ler Planilha de Clientes: {e}")
    else:
        df['N_Clientes_Conjunto'] = 0

    # --- 3.6. INTEGRAÇÃO SAP: CUSTO E EXECUÇÃO DE ORDENS (IW38) ---
    # Compara o valor Orçado (Planejado) contra o que realmente foi Gasto (Real)
    df_ordem = db.carregar_base_dataframe("base_iw38")
    if df_ordem is not None and not df_ordem.empty:
        try:
            colunas_ordem = ['Ordem', 'Status usuário', 'Status do sistema', 'Total planejado','Total real']
            col_validas = [c for c in colunas_ordem if c in df_ordem.columns]
            df_ordem = df_ordem[col_validas]
            df_ordem['Ordem'] = df_ordem['Ordem'].dropna().astype(int).astype(str).str.strip()
            dicionario_centro_sap = dict(zip(df_ordem['Ordem'], df_ordem['Status usuário']))
            dicionario_status_sistema_sap = dict(zip(df_ordem['Ordem'], df_ordem['Status do sistema']))
            dicionario_total_planejado_ordem = dict(zip(df_ordem['Ordem'], df_ordem['Total planejado']))
            dicionario_total_real_ordem = dict(zip(df_ordem['Ordem'], df_ordem['Total real']))

            if 'Ordem' in df.columns:
                chave_busca_ordem = pd.to_numeric(df['Ordem'], errors='coerce').dropna().astype(int).astype(str).str.strip()
                df.loc[df['Ordem'] != "Fora SAP", 'Status_Usuário_Ordem'] = chave_busca_ordem.map(dicionario_centro_sap).fillna("-")
                df['Status_Usuário_Ordem'] = df['Status_Usuário_Ordem'].fillna("-")

                df.loc[df['Ordem'] != "Fora SAP", 'Status_Sistema'] = chave_busca_ordem.map(dicionario_status_sistema_sap).fillna("-")
                df['Status_Sistema'] = df['Status_Sistema'].fillna("-")

                df.loc[df['Ordem'] != "Fora SAP", 'Total_planejado_ordem'] = chave_busca_ordem.map(dicionario_total_planejado_ordem).fillna(0.0)
                df['Total_planejado_ordem'] = pd.to_numeric(df['Total_planejado_ordem'], errors='coerce').fillna(0.0)

                df.loc[df['Ordem'] != "Fora SAP", 'Total_real_ordem'] = chave_busca_ordem.map(dicionario_total_real_ordem).fillna(0.0)
                df['Total_real_ordem'] = pd.to_numeric(df['Total_real_ordem'], errors='coerce').fillna(0.0)

                # Cálculo percentual de avanço financeiro da obra
                def calcular_exec_percentagem(row):
                    try:
                        planejado = float(row['Total_planejado_ordem'])
                        real = float(row['Total_real_ordem'])
                        if planejado > 0: return (real / planejado) * 100
                        elif real > 0 and planejado == 0: return 100.0
                        else: return 0.0
                    except:
                        return "-"

                df['Exec_percentagem_ordem'] = df.apply(calcular_exec_percentagem, axis=1)

                # Traduz a sopa de letrinhas do SAP (JAND INVE, etc) em um simples SIM ou NÃO para facilitar a auditoria
                df['Ordem_Executada'] = df['Status_Usuário_Ordem'].map(config.MAP_ORDEM_EXECUTADA).fillna("NÃO")

        except Exception as e:
            df['Status_Usuário_Ordem'] = "Erro na leitura"
            print(f"Erro ao ler IW38: {e}")
    else:
        df['Status_Usuário_Ordem'] = "Pendente Extração IW38"
        df['Status_Sistema'] = "Pendente Extração IW38"
        df['Total_planejado_ordem'] = "Pendente Extração IW38"
        df['Total_real_ordem'] = "Pendente Extração IW38"
        df['Exec_percentagem_ordem'] = "Pendente Extração IW38"
        df['Ordem_Executada'] = "Pendente Extração IW38"

    map_projeto_construcao = carregar_projeto_construcao()
    df['Projeto_Construcao'] = df['CJ_Aneel'].map(map_projeto_construcao).fillna("-")
    df["Modular"] = df["Conjunto"].astype(str).str.contains("MODULAR", case=False, na=False).map({True: "Sim", False: "Não"})

    # --- 3.7. PROCV COMPLEXO: CUSTOS MODULARES, CHI, CI E SAZONALIDADE ---
    # Lê os custos padrão dos conjuntos modulares e multiplica pelo volume planejado (Planejado_DDPM)
    colunas_modulo_9 = ['Modular', 'CHI', 'CI', 'Ocorrencia', 'DEC_PROG_CHI', 'CHI_Sazonal_2025', 'Total_planejado_modular']
    for col in colunas_modulo_9: df[col] = 0.0

    df_custo_raw = db.carregar_base_dataframe("base_custo_modular")
    df_sazonal_excel = db.carregar_base_dataframe("base_sazonal")
    
    if df_custo_raw is not None and not df_custo_raw.empty:
        try:
            df_custo_raw.columns = df_custo_raw.columns.astype(str).str.strip()

            col_chave_excel = [c for c in df_custo_raw.columns if 'Conjunto' in c][0]
            col_valor_excel = [c for c in df_custo_raw.columns if 'Custo Modular' in c][0]
            chi_col = [c for c in df_custo_raw.columns if 'CHI' in c][0]
            ci_col = [c for c in df_custo_raw.columns if 'CI' in c][0]
            ocor_col = [c for c in df_custo_raw.columns if 'Ocor' in c][0]
            col_m_excel = df_custo_raw.columns[12] if len(df_custo_raw.columns) > 12 else None

            df_custo = df_custo_raw[[col_chave_excel, col_valor_excel, chi_col, ci_col, ocor_col]].copy()
            df_custo.columns = ['chave', 'valor', 'chi_b', 'ci_b', 'ocor_b']
            df_custo['chave'] = df_custo['chave'].apply(normalizar_conjunto)

            def limpar_numero_br(valor):
                if pd.isna(valor): return 0.0
                if isinstance(valor, (int, float)): return float(valor)
                v_str = str(valor).upper().replace('R$', '').strip()
                if ',' in v_str: v_str = v_str.replace('.', '').replace(',', '.')
                try: return float(v_str)
                except: return 0.0

            df_custo['valor'] = df_custo['valor'].apply(limpar_numero_br)
            df_custo['chi_b'] = df_custo['chi_b'].apply(limpar_numero_br)
            df_custo['ci_b'] = df_custo['ci_b'].apply(limpar_numero_br)
            df_custo['ocor_b'] = df_custo['ocor_b'].apply(limpar_numero_br)

            dict_custo = dict(zip(df_custo['chave'], df_custo['valor']))
            dict_chi = dict(zip(df_custo['chave'], df_custo['chi_b']))
            dict_ci = dict(zip(df_custo['chave'], df_custo['ci_b']))
            dict_ocor = dict(zip(df_custo['chave'], df_custo['ocor_b']))

            dict_dec_prog = {}
            if col_m_excel:
                dict_dec_prog = dict(zip(df_custo_raw[col_chave_excel].apply(normalizar_conjunto), df_custo_raw[col_m_excel].fillna(0.0)))

            dict_sazonal = {}
            try:
                if df_sazonal_excel is not None and len(df_sazonal_excel.columns) >= 21:
                    df_saz = df_sazonal_excel.iloc[:, 20:32]
                    if not df_saz.empty:
                        dict_sazonal = dict(zip(df_saz.iloc[0].astype(int), df_saz.iloc[3].astype(float)))
            except Exception as e_saz:
                print(f"Sazonalidade não carregada: {e_saz}")

            if 'Conjunto' in df.columns:
                chave_busca = df['Conjunto'].apply(normalizar_conjunto)
                quantidade_g2 = pd.to_numeric(df['Planejado_DDPM'], errors='coerce').fillna(0.0)

                # A quantidade planejada atua como multiplicador das métricas unitárias
                df['Modular'] = chave_busca.map(dict_custo).fillna(0.0)
                df['Total_planejado_modular'] = df['Modular'] * quantidade_g2
                df['CHI'] = chave_busca.map(dict_chi).fillna(0.0) * quantidade_g2
                df['CI'] = chave_busca.map(dict_ci).fillna(0.0) * quantidade_g2
                df['Ocorrencia'] = chave_busca.map(dict_ocor).fillna(0.0) * quantidade_g2
                df['DEC_PROG_CHI'] = chave_busca.map(dict_dec_prog).fillna(0.0) * quantidade_g2

                df['Data_H2'] = pd.to_datetime(df['Data_Envio_Projeto'], errors='coerce', dayfirst=True)
                df['Ano_H2'] = df['Data_H2'].dt.year
                df['Mes_H2'] = df['Data_H2'].dt.month

                fator_proch = df['Mes_H2'].map(dict_sazonal).fillna(0.0)
                col_cc_multiplicador = pd.to_numeric(df['CC'], errors='coerce').fillna(0.0) if 'CC' in df.columns else 0.0

                df['CHI_Sazonal_2025'] = np.where(df['Ano_H2'] == 2025, fator_proch * col_cc_multiplicador, 0.0)
                df = df.drop(columns=['Data_H2', 'Ano_H2', 'Mes_H2'], errors='ignore')

        except Exception as e:
            print(f"Erro Crítico no Bloco 9 (Modulares): {e}")

    # --- 3.8. REGRAS DE NEGÓCIO: CÁLCULO DEC E FEC ---
    # DEC = Duração das interrupções / Nº de Clientes. FEC = Frequência / Nº de Clientes
    col_bw_clientes = 'N_Clientes_Conjunto'
    col_bx_duracao = 'CHI'
    col_by_freq = 'CI'

    df['DEC'] = 0.0
    df['FEC'] = 0.0

    if all(col in df.columns for col in [col_bw_clientes, col_bx_duracao, col_by_freq]):
        cond_divisao_valida = (df[col_bw_clientes] != 0) & (df[col_bw_clientes].notna())
        df.loc[cond_divisao_valida, 'DEC'] = df[col_bx_duracao] / df[col_bw_clientes]
        df.loc[cond_divisao_valida, 'FEC'] = df[col_by_freq] / df[col_bw_clientes]

    # --- 3.9. PROCV COMPOSTO: FATOR DE GANHOS CHI-CONJ ---
    # Avaliação de Ganhos utilizando duas colunas como chave (Conjunto + Circuito Aneel)
    df['CHI_Conj'] = 0.0

    df_ganhos = db.carregar_base_dataframe("base_ganhos")
    if df_ganhos is not None and not df_ganhos.empty and len(df_ganhos.columns) > 10:
        try:
            df_ganhos.columns = df_ganhos.columns.astype(str).str.strip()

            col_c_excel = df_ganhos.columns[2]
            col_b_excel = df_ganhos.columns[1]
            col_k_excel = df_ganhos.columns[10]

            df_ganhos['chave_composta'] = (
                df_ganhos[col_c_excel].apply(normalizar_conjunto) + "_" +
                df_ganhos[col_b_excel].astype(str).str.strip().str.upper()
            )

            dict_ganhos = dict(zip(df_ganhos['chave_composta'], df_ganhos[col_k_excel].fillna(0.0)))

            chave_busca_sistema = (
                df['Conjunto'].apply(normalizar_conjunto) + "_" +
                df['CJ_Aneel'].astype(str).str.strip().str.upper()
            )

            fator_ganhos = chave_busca_sistema.map(dict_ganhos).fillna(0.0)
            quantidade_g2 = pd.to_numeric(df['Planejado_DDPM'], errors='coerce').fillna(0.0)

            df['CHI_Conj'] = fator_ganhos * quantidade_g2

        except Exception as e:
            print(f"Erro ao ler planilha de Ganhos: {e}")

    # --- 3.10. HISTÓRICOS: 12 MESES E 3 MESES (fonte Table1 descontinuada) ---
    for col in ['CI_12M', 'CHI_12M', 'OCO_12M', 'OCO_3M']: df[col] = "-"

    # --- 3.11. LÓGICA DE TOPOLOGIA DE PROTEÇÃO ---
    # O código da topologia está embutido (escondido) dentro do nome do Local de Instalação.
    col_f = 'Local_Instalacao'
    local_limpo = df[col_f].astype(str).str.strip().str.upper()
    # Fatia (slicing) a string do Local Instalação para montar o código-chave de busca
    parte1 = local_limpo.str[0:3].str.strip()
    parte2 = local_limpo.str[4:6].str.strip()
    parte3 = local_limpo.str[7:17].str.strip()
    chave_protecao = parte1 + parte2 + parte3 + "9"
    equipamento_protecao_direto = parte2.isin(['RL', 'BR', 'BF', 'DJ'])

    # Atribui diretamente a chave extraída se for um equipamento de proteção, caso contrário "-"
    df['Equipamento_Protecao'] = np.where(equipamento_protecao_direto, chave_protecao, "-")

    # --- 3.12. INTEGRAÇÃO SAP: MEDIDAS IW66 ---
    df['Medida_SAP'] = "-"
    df_medidas_raw = _ler_export_medidas()
    if df_medidas_raw is not None:
        try:
            df_m = df_medidas_raw.copy()

            # Normalização resiliente das colunas do IW66 (trata encoding do SQLite ex: Denominao / N de ordenao)
            novas_cols = {}
            for col in df_m.columns:
                c_norm = unicodedata.normalize('NFKD', str(col)).encode('ascii', 'ignore').decode('utf-8').lower()
                if 'denomina' in c_norm and 'conjunto' in c_norm:
                    novas_cols[col] = 'Denominação do conjunto'
                elif 'ordena' in c_norm:
                    novas_cols[col] = 'Nº de ordenação'
                elif 'texto' in c_norm and 'medida' in c_norm:
                    novas_cols[col] = 'Texto medida'
                elif 'descri' in c_norm:
                    novas_cols[col] = 'Descrição'

            df_m = df_m.rename(columns=novas_cols)
            df_m['Nota'] = df_m['Nota'].dropna().astype(int).astype(str).str.strip()

            _UN_DENOMS = {"POSTE", "TRANSFORMADOR", "TRANSF", "TRAFO", "SUBST", "CHAVE",
                          "RELIGADOR", "SECCIONALIZADOR", "DISJUNTOR", "DJ", "BF", "LBS",
                          "MONITORAMENTO", "MANUT. CIRC"}
            _M_DENOMS = {"REDE", "RDS", "BLINDAGEM", "MELHORIA OPERATIVA"}
            _UN_KW = {"POSTE", "TRANSF", "TRAFO", "RELIG", "CHAVE", "SECCIONALIZADOR",
                      "DISJUNTOR", "DJ"}
            _M_KW = {"CONDUTOR", "CABO", "SPACER", "RECOND", "CONSTR",
                     "BLINDAR", "EXTENSAO", "REDE"}

            def _classificar(row):
                denom = str(row.get("Denominação do conjunto", "")).strip().upper()
                texto = str(row.get("Texto medida", "")).strip().upper()
                desc = str(row.get("Descrição", "")).strip().upper()
                try:
                    val = float(row.get("Nº de ordenação", 0) or 0)
                except (TypeError, ValueError):
                    val = 0.0
                is_un = any(kw in denom for kw in _UN_DENOMS)
                is_m = any(kw in denom for kw in _M_DENOMS)
                has_un = any(kw in texto or kw in desc for kw in _UN_KW)
                has_m = any(kw in texto or kw in desc for kw in _M_KW)
                if val > 20:
                    return val, "m"
                if is_un:
                    return val, "un"
                if is_m:
                    if has_un and not has_m and val <= 20:
                        return val, "un"
                    return val, "m"
                if has_un and not has_m and val <= 20:
                    return val, "un"
                if has_m and not has_un:
                    return val, "m"
                return val, ("m" if val >= 10 else "un")

            df_m[["val_class", "unit_class"]] = pd.DataFrame(
                df_m.apply(_classificar, axis=1).tolist(), index=df_m.index
            )
            df_m["val_m"] = np.where(df_m["unit_class"] == "m", df_m["val_class"], 0.0)
            df_m["val_un"] = np.where(df_m["unit_class"] == "un", df_m["val_class"], 0.0)
            grouped = df_m.groupby("Nota")[["val_m", "val_un"]].sum().reset_index()

            def _format_medida(row):
                parts = []
                if row["val_m"] > 0:
                    km = row["val_m"] / 1000.0
                    parts.append(f"{km:.3f}".rstrip("0").rstrip(".") + " km")
                if row["val_un"] > 0:
                    un = int(row["val_un"]) if float(row["val_un"]).is_integer() else f"{row['val_un']:.1f}"
                    parts.append(f"{un} un")
                return " / ".join(parts) if parts else "-"

            grouped["Medida_SAP_Str"] = grouped.apply(_format_medida, axis=1)
            dict_medidas = dict(zip(grouped["Nota"], grouped["Medida_SAP_Str"]))
            df["Medida_SAP"] = df["Numero_Nota"].astype(str).str.strip().map(dict_medidas).fillna("-")

            # Agregação de Medidas das Filhas para Notas Mães
            # Em grupos hierárquicos no SAP (IW66), as medidas físicas ficam alocadas nas notas filhas.
            # Para notas mães sem medida direta, agregamos a soma das medidas das suas filhas ativas.
            try:
                df_temp = df[['Numero_Nota', 'Nota_Mae', 'Status_Nota', 'Medida_SAP']].copy()
                df_temp['Nota_Mae_Limpa'] = df_temp['Nota_Mae'].fillna('').astype(str).str.strip()
                df_temp['Status_Nota_Str'] = df_temp['Status_Nota'].fillna('').astype(str).str.strip()

                inativos = {'ENCE CANC', 'SUPR CANC', 'ENCE EXEC', 'SUPR', '999', '998', '997', '55', '99'}
                def _eh_ativa(st):
                    st_u = st.upper()
                    return not (st_u in inativos or st_u.startswith('55') or st_u.startswith('99'))

                df_temp['Ativa'] = df_temp['Status_Nota_Str'].apply(_eh_ativa)
                filhas_ativas = df_temp[
                    df_temp['Ativa'] &
                    df_temp['Nota_Mae_Limpa'].str.isdigit() &
                    (df_temp['Nota_Mae_Limpa'] != '0')
                ].copy()

                if not filhas_ativas.empty:
                    def _extrair_val_m_un(medida_str):
                        if not medida_str or medida_str == '-': return 0.0, 0.0
                        val_m, val_un = 0.0, 0.0
                        m_km = re.search(r'([\d.]+)\s*km', str(medida_str).lower())
                        if m_km:
                            try: val_m = float(m_km.group(1)) * 1000.0
                            except ValueError: pass
                        m_un = re.search(r'([\d.]+)\s*un', str(medida_str).lower())
                        if m_un:
                            try: val_un = float(m_un.group(1))
                            except ValueError: pass
                        return val_m, val_un

                    vals = filhas_ativas['Medida_SAP'].apply(_extrair_val_m_un).tolist()
                    filhas_ativas['val_m'] = [v[0] for v in vals]
                    filhas_ativas['val_un'] = [v[1] for v in vals]

                    agg_maes = filhas_ativas.groupby('Nota_Mae_Limpa')[['val_m', 'val_un']].sum().reset_index()
                    agg_maes['Medida_Agregada'] = agg_maes.apply(_format_medida, axis=1)
                    dict_agg = dict(zip(agg_maes['Nota_Mae_Limpa'], agg_maes['Medida_Agregada']))

                    num_nota_str = df['Numero_Nota'].astype(str).str.strip()
                    medida_agregada_col = num_nota_str.map(dict_agg).fillna('-')

                    df['Medida_SAP'] = np.where(
                        (df['Medida_SAP'] == '-') & (medida_agregada_col != '-'),
                        medida_agregada_col,
                        df['Medida_SAP']
                    )
            except Exception as e_agg:
                print(f"Aviso ao agregar medidas de notas filhas para mães: {e_agg}")
        except Exception as e:
            print(f"Erro ao processar medidas IW66: {e}")
            df["Medida_SAP"] = "Erro"

    # --- 3.13. COMPARAÇÃO: MEDIDA VS PLANEJADO ---
    df["Medida_vs_Planejado"] = df.apply(
        lambda r: _comparar_medida_planejado(r.get("Medida_SAP", "-"), r.get("Planejado_DDPM")),
        axis=1,
    )

    if 'Ordem' in df.columns:
        df['Ordem'] = df['Ordem'].replace("Fora SAP", "-")
    df["Auditoria_Cronograma"] = df.apply(avaliar_prazo_sap, axis=1)
    return df


# ====================================================================
# CACHE EM MEMÓRIA (validado por versão do dataset + TTL de fallback) E
# METADADOS DAS BASES
# ====================================================================
_CACHE_TTL_SEGUNDOS = 600  # fallback para escritas que não passem pelos logs
_cache = {"df": None, "quando": 0.0, "versao": None}
_cache_lock = threading.Lock()

_sincronizando_rede = False
_sincronizando_lock = threading.Lock()

def esta_sincronizando_rede() -> bool:
    with _sincronizando_lock:
        return _sincronizando_rede


def get_dataset(forcar: bool = False) -> pd.DataFrame:
    with _cache_lock:
        versao = db.obter_versao_dataset()
        expirado = time.time() - _cache["quando"] > _CACHE_TTL_SEGUNDOS
        if (forcar or _cache["df"] is None or expirado
                or _cache["versao"] != versao):
            df_res = enriquecer_dados()
            colunas_existentes = [col for col in config.COLUNAS_PAINEL if col in df_res.columns]
            colunas_extras = [col for col in df_res.columns if col not in colunas_existentes]
            _cache["df"] = df_res[colunas_existentes + colunas_extras]
            _cache["quando"] = time.time()
            _cache["versao"] = versao
        return _cache["df"].copy()


def invalidar_cache() -> None:
    with _cache_lock:
        _cache["df"] = None


_STATUS_BASES_TTL_SEGUNDOS = 60
_status_bases_cache = {"quando": 0.0, "valor": None}
_status_bases_lock = threading.Lock()


def _checar_base_com_timeout(item):
    nome, caminho = item
    try:
        if not caminho:
            return {"nome": nome, "arquivo": "", "encontrada": False, "modificada": None}
        if "IW38" in nome and not os.path.exists(caminho):
            pasta = os.path.dirname(caminho)
            alt = os.path.join(pasta, "Gerada_custo_ord_IW38.XLSX" if "Gerada_ord" in caminho else "Gerada_ord_IW38.XLSX")
            if os.path.exists(alt):
                caminho = alt
        existe = os.path.exists(caminho)
        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(caminho)).isoformat() if existe else None
        return {
            "nome": nome,
            "arquivo": os.path.basename(caminho),
            "encontrada": existe,
            "modificada": mtime,
        }
    except Exception:
        return {
            "nome": nome,
            "arquivo": os.path.basename(caminho) if caminho else "",
            "encontrada": False,
            "modificada": None,
        }


def status_bases() -> list:
    """Stats dos 7 caminhos SMB, cacheados 60s com checagem paralela e timeout."""
    with _status_bases_lock:
        agora = time.time()
        if (_status_bases_cache["valor"] is not None
                and agora - _status_bases_cache["quando"] < _STATUS_BASES_TTL_SEGUNDOS):
            return _status_bases_cache["valor"]

        bases_map = {}
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=min(8, len(config.BASES_REDE) or 1)) as executor:
                futuros = {
                    executor.submit(_checar_base_com_timeout, item): item[0]
                    for item in config.BASES_REDE.items()
                }
                concluidos, _ = concurrent.futures.wait(futuros.keys(), timeout=4.0)
                for f in concluidos:
                    try:
                        res = f.result()
                        bases_map[res["nome"]] = res
                    except Exception:
                        pass
        except Exception:
            pass

        bases = []
        for nome, caminho in config.BASES_REDE.items():
            if nome in bases_map:
                bases.append(bases_map[nome])
            else:
                bases.append({
                    "nome": nome,
                    "arquivo": os.path.basename(caminho) if caminho else "",
                    "encontrada": False,
                    "modificada": None,
                })

        _status_bases_cache["quando"] = agora
        _status_bases_cache["valor"] = bases
        return bases


def invalidar_status_bases() -> None:
    """Força status_bases() a reler o filesystem na próxima chamada.

    Chamar após qualquer escrita que troque um arquivo de BASES_REDE
    (upload manual, sync SAP) — o TTL de 60s por si só não detecta isso.
    """
    with _status_bases_lock:
        _status_bases_cache["valor"] = None


# ====================================================================
# CÓPIA EXCEL NA REDE — porte de Input/processamento.py:387
# (nunca derruba a request: corpo inteiro em try/except)
# ====================================================================
def gerar_copia_excel_rede():
    """Puxa a base mais recente, executa todos os cruzamentos e gera o Excel
    sincronizado na rede para alimentar as planilhas laterais.

    Toda a lógica está protegida por try/except: se a rede estiver indisponível
    o erro é apenas registrado, sem derrubar a request que disparou a tarefa.
    """
    if os.environ.get("PYTEST_CURRENT_TEST") or os.environ.get("INPUT_DATA_DIR"):
        return

    # Perfil local não publica nada: sai ANTES de enriquecer dados, de checar/
    # remover locks "~$" e de qualquer escrita em caminho de rede.
    if not config.em_producao():
        print("⚠️ [input] Perfil LOCAL: cópia Excel da rede NÃO gerada e banco "
              "NÃO espelhado — as alterações ficam apenas nesta máquina. "
              "Rode o servidor com EDP_PERFIL=producao para publicar.")
        return

    global _sincronizando_rede
    with _sincronizando_lock:
        if _sincronizando_rede:
            print("Sincronização com a rede já em andamento. Ignorando chamada duplicada.")
            return
        _sincronizando_rede = True
    try:
        # 1. Puxa os dados atualizados com todos os cálculos automáticos prontos
        df_fresco = enriquecer_dados()

        # Filtra e renomeia as colunas para o mesmo padrão amigável do painel
        colunas_exportar = [col for col in config.COLUNAS_PAINEL if col in df_fresco.columns]
        df_export = df_fresco[colunas_exportar].copy()
        df_export = df_export.rename(columns=config.MAPA_NOMES_EXCEL_LEGADO)

        # Função auxiliar para formatar e salvar o Excel de forma idêntica ao legado
        def salvar_excel_formatado(caminho):
            dir_name = os.path.dirname(caminho)
            base_name = os.path.basename(caminho)
            caminho_owner = os.path.join(dir_name, f"~${base_name}")
            caminho_tmp = caminho.replace('.xlsx', '_TEMP.xlsx') if caminho.endswith('.xlsx') else caminho + '_TEMP.xlsx'

            if os.path.exists(caminho_owner):
                try:
                    os.remove(caminho_owner)
                    print(f"🧹 Lock fantasma '{caminho_owner}' removido automaticamente.")
                except Exception:
                    pass

            if os.path.exists(caminho_tmp):
                try: os.remove(caminho_tmp)
                except Exception: pass

            with pd.ExcelWriter(caminho_tmp, engine='openpyxl') as writer:
                nome_aba = 'Input de Notas'
                df_export.to_excel(writer, sheet_name=nome_aba, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets[nome_aba]
                
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                # Estilo do cabeçalho
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                data_font = Font(name='Calibri', size=11)
                
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin', color='A6A6A6'),
                    right=Side(style='thin', color='A6A6A6'),
                    top=Side(style='thin', color='A6A6A6'),
                    bottom=Side(style='thin', color='A6A6A6')
                )
                
                # Formata linha do cabeçalho
                worksheet.row_dimensions[1].height = 45
                for col_num in range(1, len(df_export.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                # Formata dados e configura largura das colunas
                for col_num, col_name in enumerate(df_export.columns, 1):
                    if "Observação" in col_name or "Observacao" in col_name:
                        width = 45
                    elif "Local" in col_name or "Conjunto" in col_name or "Circuito" in col_name:
                        width = 20
                    else:
                        width = 14
                    
                    letter = chr(64 + col_num) if col_num <= 26 else f"{chr(64 + col_num // 26)}{chr(64 + col_num % 26)}"
                    worksheet.column_dimensions[letter].width = width
                    
                    for row_num in range(2, len(df_export) + 2):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.alignment = center_align
                        cell.font = data_font
                
                # Adiciona autofiltro e congela cabeçalho
                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.freeze_panes = 'A2'

            if os.path.exists(caminho_tmp):
                try:
                    os.replace(caminho_tmp, caminho)
                except Exception:
                    import shutil
                    shutil.move(caminho_tmp, caminho)

        # 2. Salva na rede a planilha principal
        salvar_excel_formatado(config.CAMINHO_COPIA_EXCEL)
        
        # 3. Salva também como "Input Nota.xlsx" na raiz da rede para compatibilidade externa
        try:
            salvar_excel_formatado(config.CAMINHO_INPUT_NOTA_RAIZ)
        except Exception as e2:
            print(f"Erro ao gerar cópia de compatibilidade Input Nota.xlsx na rede: {e2}")
            
        # 4. Banco de notas: nada a fazer.
        # Aqui só chega o perfil de produção, onde o banco EM USO já é o da rede
        # (config.caminho_banco_notas) — a escrita já caiu no arquivo
        # compartilhado. Nunca reintroduzir src.backup(dst) neste ponto: ele
        # sobrescreve o arquivo inteiro da rede e apaga o que os outros usuários
        # gravaram (removido em ef19f4f). Se o perfil local precisar publicar, o
        # caminho é UPSERT por Numero_Nota, nunca cópia de arquivo.

    except Exception as e:
        print(f"Erro ao gerar cópia Excel na rede: {e}")
    finally:
        with _sincronizando_lock:
            _sincronizando_rede = False
