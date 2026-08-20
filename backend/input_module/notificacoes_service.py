import datetime
import logging
import os
import re
import tempfile
import pandas as pd
from input_module import db

logger = logging.getLogger("NotificacoesService")


def _extrair_data_str(dt_val) -> str:
    if dt_val is None or pd.isna(dt_val):
        return ""
    if isinstance(dt_val, (datetime.date, datetime.datetime)):
        return dt_val.strftime("%Y-%m-%d")
    s = str(dt_val).strip()
    return s[:10] if len(s) >= 10 else s


def _extrair_lista_engenheiros(resp_val) -> list[str]:
    """Extrai lista de nomes de engenheiros (suporta string com vírgula, barra, ponto e vírgula ou lista)."""
    if resp_val is None or pd.isna(resp_val):
        return []
    if isinstance(resp_val, (list, tuple, set)):
        return [str(e).strip() for e in resp_val if str(e).strip()]
    s = str(resp_val).strip()
    if not s or s == "-":
        return []
    # Divide por vírgula, ponto e vírgula, barra ou ' e ' / ' and '
    partes = re.split(r"[,;/]|\be\b|\band\b", s, flags=re.IGNORECASE)
    return [p.strip() for p in partes if p.strip() and p.strip() != "-"]


def obter_resumo_alteracoes_diarias(data_referencia: str | None = None) -> dict:
    """Consolida as alterações do log por Engenheiro e Regional para a data de referência (padrão: hoje).
    Suporta múltiplos engenheiros atribuídos à mesma regional/conjunto.
    """
    if not data_referencia:
        data_referencia = datetime.date.today().isoformat()

    df_logs = db.carregar_logs()
    df_notas = db.carregar_dados()
    de_para_resp = db.carregar_responsaveis()
    emails_resp = db.carregar_emails_responsaveis()

    # Mapeia todos os engenheiros conhecidos (seja nas chaves de email ou no de_para)
    todos_engenheiros = set(emails_resp.keys())
    for reg, resp_val in de_para_resp.items():
        for eng in _extrair_lista_engenheiros(resp_val):
            todos_engenheiros.add(eng)

    engenheiros_map: dict[str, dict] = {}
    for eng in sorted(todos_engenheiros):
        if not eng or eng == "-":
            continue
        # Identifica todas as regionais/conjuntos que incluem este engenheiro
        reg_list = []
        for reg_key, resp_val in de_para_resp.items():
            if eng in _extrair_lista_engenheiros(resp_val):
                reg_list.append(reg_key)

        engenheiros_map[eng] = {
            "engenheiro": eng,
            "email": emails_resp.get(eng, ""),
            "regionais": sorted(list(set(reg_list))),
            "total_alteracoes": 0,
            "total_notas_afetadas": 0,
            "notas_afetadas": [],
            "alteracoes": [],
        }

    if df_logs.empty:
        return {
            "data_referencia": data_referencia,
            "total_alteracoes": 0,
            "total_notas_afetadas": 0,
            "engenheiros": engenheiros_map,
        }

    # Filtra logs pela data de referência
    df_logs["Data_Str"] = df_logs["Data_Hora"].apply(_extrair_data_str)
    df_dia = df_logs[df_logs["Data_Str"] == data_referencia].copy()

    if df_dia.empty:
        return {
            "data_referencia": data_referencia,
            "total_alteracoes": 0,
            "total_notas_afetadas": 0,
            "engenheiros": engenheiros_map,
        }

    # Metadados das notas
    dict_notas = {}
    if not df_notas.empty:
        for _, row in df_notas.iterrows():
            num = int(row["Numero_Nota"]) if pd.notna(row["Numero_Nota"]) else None
            if num:
                dict_notas[num] = {
                    "Regional": str(row.get("Regional", "-") or "-").strip(),
                    "Conjunto": str(row.get("Conjunto", "-") or "-").strip(),
                    "Circuito": str(row.get("Circuito", "-") or "-").strip(),
                    "Local_Instalacao": str(row.get("Local_Instalacao", "-") or "-").strip(),
                }

    for _, row in df_dia.iterrows():
        try:
            num_nota = int(row["Numero_Nota"])
        except (ValueError, TypeError):
            continue

        info_nota = dict_notas.get(num_nota, {})
        regional = info_nota.get("Regional", "-")
        conjunto = info_nota.get("Conjunto", "-")
        circuito = info_nota.get("Circuito", "-")

        # Procura engenheiros atribuídos à regional ou ao conjunto
        resp_raw = de_para_resp.get(regional) or de_para_resp.get(conjunto)
        engs_atribuidos = _extrair_lista_engenheiros(resp_raw)
        if not engs_atribuidos:
            engs_atribuidos = ["Não Definido"]

        campo = str(row.get("Campo_Alterado", "-"))
        antigo = str(row.get("Valor_Antigo", "-"))
        novo = str(row.get("Valor_Novo", "-"))
        usuario = str(row.get("Usuario", "sistema"))
        dt_hora_str = str(row.get("Data_Hora", ""))

        if campo == "Nota_Mae":
            tipo_evento = "Vínculo / Hierarquia"
            detalhe = f"Nota Mãe alterada de '{antigo}' para '{novo}'"
        elif campo == "CRIAÇÃO DE NOTA":
            tipo_evento = "Criação de Nota"
            detalhe = novo
        elif campo == "EXCLUSÃO DE NOTA":
            tipo_evento = "Exclusão de Nota"
            detalhe = "Nota removida do plano"
        else:
            tipo_evento = "Edição de Campo"
            detalhe = f"{campo}: '{antigo}' ➔ '{novo}'"

        item = {
            "ID_Log": int(row.get("ID_Log", 0)) if pd.notna(row.get("ID_Log")) else 0,
            "Numero_Nota": num_nota,
            "Regional": regional,
            "Conjunto": conjunto,
            "Circuito": circuito,
            "Tipo_Evento": tipo_evento,
            "Campo_Alterado": campo,
            "Valor_Antigo": antigo,
            "Valor_Novo": novo,
            "Detalhe": detalhe,
            "Usuario": usuario,
            "Data_Hora": dt_hora_str,
        }

        # Registra a alteração para CADA engenheiro atribuído
        for eng in engs_atribuidos:
            if eng not in engenheiros_map:
                engenheiros_map[eng] = {
                    "engenheiro": eng,
                    "email": emails_resp.get(eng, ""),
                    "regionais": [regional] if regional != "-" else [],
                    "total_alteracoes": 0,
                    "total_notas_afetadas": 0,
                    "notas_afetadas": [],
                    "alteracoes": [],
                }
            eng_entry = engenheiros_map[eng]
            eng_entry["total_alteracoes"] += 1
            if num_nota not in eng_entry["notas_afetadas"]:
                eng_entry["notas_afetadas"].append(num_nota)
            eng_entry["alteracoes"].append(item)

    for eng, dados in engenheiros_map.items():
        dados["total_notas_afetadas"] = len(dados["notas_afetadas"])

    total_alt = len(df_dia)
    total_notas_af = len(set(df_dia["Numero_Nota"].astype(int, errors="ignore")))

    return {
        "data_referencia": data_referencia,
        "total_alteracoes": total_alt,
        "total_notas_afetadas": total_notas_af,
        "engenheiros": engenheiros_map,
    }


def gerar_planilha_alteracoes_anexo(engenheiro: str, alteracoes: list[dict], data_referencia: str) -> str:
    """Gera um arquivo Excel (.xlsx) formatado com as notas modificadas do engenheiro e retorna o caminho temporário."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # 1. Aba de Notas Modificadas
    ws_det = wb.active
    ws_det.title = "Notas Modificadas"
    ws_det.views.sheetView[0].showGridLines = True

    headers = [
        "Nº Nota", "Regional", "Conjunto", "Circuito",
        "Tipo de Evento", "Campo Alterado", "Valor Anterior",
        "Valor Novo", "Detalhe da Alteração", "Responsável Alteração", "Data/Hora"
    ]
    ws_det.append(headers)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws_det.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    ws_det.row_dimensions[1].height = 28

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    data_font = Font(name="Calibri", size=10)

    for row_idx, item in enumerate(alteracoes, 2):
        row_data = [
            item.get("Numero_Nota"),
            item.get("Regional", "-"),
            item.get("Conjunto", "-"),
            item.get("Circuito", "-"),
            item.get("Tipo_Evento", "-"),
            item.get("Campo_Alterado", "-"),
            item.get("Valor_Antigo", "-"),
            item.get("Valor_Novo", "-"),
            item.get("Detalhe", "-"),
            item.get("Usuario", "-"),
            item.get("Data_Hora", "-"),
        ]
        ws_det.append(row_data)
        is_even = row_idx % 2 == 0
        for col_idx in range(1, len(row_data) + 1):
            c = ws_det.cell(row=row_idx, column=col_idx)
            c.font = data_font
            c.border = thin_border
            if is_even:
                c.fill = zebra_fill
            if col_idx in (1, 11):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws_det.row_dimensions[row_idx].height = 20

    # Auto largura de colunas
    for col in ws_det.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_det.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 2. Aba de Resumo Executivo
    ws_res = wb.create_sheet(title="Resumo Executivo")
    ws_res.views.sheetView[0].showGridLines = True

    ws_res.append(["EDP-Helios • Relatório Diário de Alterações em Notas"])
    ws_res.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    ws_res.append([])
    ws_res.append(["Data de Referência:", data_referencia])
    ws_res.append(["Engenheiro Destinatário:", engenheiro])
    ws_res.append(["Total de Notas Modificadas:", len({i["Numero_Nota"] for i in alteracoes})])
    ws_res.append(["Total de Alterações Registradas:", len(alteracoes)])
    ws_res.append([])

    ws_res.append(["Categoria de Alteração", "Qtd Eventos"])
    ws_res.cell(row=8, column=1).fill = header_fill
    ws_res.cell(row=8, column=1).font = header_font
    ws_res.cell(row=8, column=2).fill = header_fill
    ws_res.cell(row=8, column=2).font = header_font

    from collections import Counter
    cat_counter = Counter(i.get("Tipo_Evento", "Edição de Campo") for i in alteracoes)
    for r_i, (cat, count) in enumerate(cat_counter.most_common(), 9):
        ws_res.append([cat, count])
        ws_res.cell(row=r_i, column=1).font = data_font
        ws_res.cell(row=r_i, column=2).font = data_font
        ws_res.cell(row=r_i, column=1).border = thin_border
        ws_res.cell(row=r_i, column=2).border = thin_border

    for col in ws_res.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_res.column_dimensions[col_letter].width = max(max_len + 4, 18)

    temp_dir = tempfile.gettempdir()
    eng_slug = "".join(c for c in engenheiro if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_")
    nome_arquivo = f"EDP-Helios_Alteracoes_{eng_slug}_{data_referencia}.xlsx"
    caminho_final = os.path.join(temp_dir, nome_arquivo)
    wb.save(caminho_final)
    return caminho_final


def gerar_email_outlook_engenheiro(engenheiro: str, data_referencia: str | None = None, usuario: str = "sistema") -> dict:
    """Gera o rascunho de e-mail diário com resumo executivo, link do EDP-Helios e anexo Excel no Outlook."""
    resumo = obter_resumo_alteracoes_diarias(data_referencia)
    dados_eng = resumo["engenheiros"].get(engenheiro)

    if not dados_eng or dados_eng["total_alteracoes"] == 0:
        return {
            "ok": False,
            "mensagem": f"Nenhuma alteração registrada para as regionais do engenheiro {engenheiro} em {resumo['data_referencia']}.",
        }

    email_to = dados_eng.get("email") or ""
    regionais_str = ", ".join(dados_eng["regionais"]) or "Geral"
    total_notas = dados_eng["total_notas_afetadas"]
    total_alt = dados_eng["total_alteracoes"]
    data_ref = resumo["data_referencia"]

    try:
        dt_obj = datetime.date.fromisoformat(data_ref)
        data_formatada = dt_obj.strftime("%d/%m/%Y")
    except Exception:
        data_formatada = data_ref

    # Gera o arquivo Excel formatado para anexo
    caminho_excel = gerar_planilha_alteracoes_anexo(engenheiro, dados_eng["alteracoes"], data_ref)
    nome_excel = os.path.basename(caminho_excel)

    # Agrupa alterações por categoria para o resumo executivo
    from collections import Counter
    cat_counter = Counter(i.get("Tipo_Evento", "Edição de Campo") for i in dados_eng["alteracoes"])

    linhas_categorias = []
    for cat, qtd in cat_counter.most_common():
        cor_badge = "#0284c7"
        if "Criação" in cat:
            cor_badge = "#16a34a"
        elif "Exclusão" in cat:
            cor_badge = "#dc2626"
        elif "Hierarquia" in cat or "Vínculo" in cat:
            cor_badge = "#7c3aed"

        linhas_categorias.append(f"""
        <tr style="border-bottom: 1px solid #f1f5f9;">
            <td style="padding: 8px 12px;">
                <span style="display: inline-block; background-color: {cor_badge}; color: #ffffff; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600;">{cat}</span>
            </td>
            <td style="padding: 8px 12px; text-align: right; font-weight: 700; font-family: monospace; color: #0f172a;">{qtd} evento(s)</td>
        </tr>
        """)

    # Amostra das primeiras 5 alterações para preview visual
    amostra = dados_eng["alteracoes"][:5]
    linhas_amostra = []
    for item in amostra:
        linhas_amostra.append(f"""
        <tr style="border-bottom: 1px solid #e2e8f0;">
            <td style="padding: 6px 10px; font-family: monospace; font-weight: 700; color: #0284c7;">{item['Numero_Nota']}</td>
            <td style="padding: 6px 10px; color: #475569;">{item['Regional']} / {item['Conjunto']}</td>
            <td style="padding: 6px 10px; font-size: 11.5px;">{item['Tipo_Evento']}</td>
            <td style="padding: 6px 10px; font-family: monospace; font-size: 11.5px; color: #1e293b;">{item['Detalhe']}</td>
        </tr>
        """)

    aviso_mais = ""
    if total_alt > 5:
        sobraram = total_alt - 5
        aviso_mais = f"""
        <div style="padding: 8px 10px; font-size: 11.5px; color: #64748b; background-color: #f8fafc; text-align: center; border-top: 1px solid #e2e8f0;">
            + {sobraram} alteração(ões) listadas detalhadamente no arquivo Excel em anexo.
        </div>
        """

    url_sistema = os.environ.get("EDP_HELIOS_URL", "http://localhost:6328").strip()

    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        outlook = win32com.client.Dispatch("Outlook.Application")
        message = outlook.CreateItem(0)

        hora = datetime.datetime.now().hour
        saudacao = "bom dia" if 6 <= hora < 12 else ("boa tarde" if 12 <= hora < 18 else "boa noite")

        message.To = email_to
        message.CC = "felipeg.bezerra@edp.com"
        message.Subject = f"EDP-Helios • Resumo de Alterações em Notas - {engenheiro} ({regionais_str}) - {data_formatada}"
        message.BodyFormat = 2  # HTML

        message.HTMLBody = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="utf-8"></head>
        <body style="margin: 0; padding: 0; background-color: #f8fafc; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; color: #1e293b;">
            <div style="max-width: 650px; margin: 16px auto; background-color: #ffffff; border-radius: 12px; overflow: hidden; border: 1px solid #e2e8f0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);">
                <!-- Header -->
                <div style="background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%); padding: 22px 26px; color: #ffffff;">
                    <div style="font-size: 11px; font-weight: 700; letter-spacing: 2px; text-transform: uppercase; color: #38bdf8; margin-bottom: 4px;">
                        Plataforma de Engenharia & Planejamento
                    </div>
                    <div style="font-size: 20px; font-weight: 700; color: #ffffff;">
                        EDP-Helios • Resumo de Alterações
                    </div>
                    <div style="font-size: 12.5px; color: #94a3b8; margin-top: 2px;">
                        Informativo diário de notas do plano de distribuição ({data_formatada})
                    </div>
                </div>

                <!-- Corpo -->
                <div style="padding: 22px 26px;">
                    <p style="font-size: 14.5px; margin-top: 0; color: #0f172a;">
                        Olá, <strong>{engenheiro}</strong>, {saudacao}!
                    </p>
                    <p style="font-size: 13.5px; color: #334155; line-height: 1.5;">
                        Informamos que foram registradas <strong>{total_alt} alteração(ões)</strong> em <strong>{total_notas} nota(s)</strong> vinculadas às suas regionais (<strong>{regionais_str}</strong>) nesta data.
                    </p>

                    <!-- Cards de KPI -->
                    <table style="width: 100%; border-collapse: separate; border-spacing: 8px; margin: 16px -8px;">
                        <tr>
                            <td style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center; width: 33%;">
                                <div style="font-size: 10.5px; text-transform: uppercase; letter-spacing: 0.5px; color: #64748b; font-weight: 600;">Notas Modificadas</div>
                                <div style="font-size: 20px; font-weight: 700; color: #0284c7; margin-top: 2px;">{total_notas}</div>
                            </td>
                            <td style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center; width: 33%;">
                                <div style="font-size: 10.5px; text-transform: uppercase; letter-spacing: 0.5px; color: #64748b; font-weight: 600;">Total de Alterações</div>
                                <div style="font-size: 20px; font-weight: 700; color: #0f172a; margin-top: 2px;">{total_alt}</div>
                            </td>
                            <td style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center; width: 33%;">
                                <div style="font-size: 10.5px; text-transform: uppercase; letter-spacing: 0.5px; color: #64748b; font-weight: 600;">Regionais</div>
                                <div style="font-size: 12px; font-weight: 700; color: #334155; margin-top: 4px;">{regionais_str}</div>
                            </td>
                        </tr>
                    </table>

                    <!-- Tabela de Categorias -->
                    <div style="margin-top: 16px; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden;">
                        <div style="background-color: #f8fafc; padding: 9px 12px; border-bottom: 1px solid #e2e8f0; font-size: 11.5px; font-weight: 700; color: #334155; text-transform: uppercase; letter-spacing: 0.5px;">
                            📊 Resumo por Categoria de Evento
                        </div>
                        <table style="width: 100%; border-collapse: collapse; font-size: 12.5px;">
                            <tbody>
                                {''.join(linhas_categorias)}
                            </tbody>
                        </table>
                    </div>

                    <!-- Amostra de Alterações -->
                    <div style="margin-top: 18px;">
                        <div style="font-size: 11.5px; font-weight: 700; color: #334155; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px;">
                            🔍 Amostra das Principais Alterações:
                        </div>
                        <table style="width: 100%; border-collapse: collapse; font-size: 11.5px; border: 1px solid #e2e8f0; border-radius: 6px; overflow: hidden;">
                            <thead>
                                <tr style="background-color: #f1f5f9; color: #475569; text-align: left;">
                                    <th style="padding: 7px 10px; border-bottom: 1px solid #e2e8f0;">Nota</th>
                                    <th style="padding: 7px 10px; border-bottom: 1px solid #e2e8f0;">Regional/Conjunto</th>
                                    <th style="padding: 7px 10px; border-bottom: 1px solid #e2e8f0;">Tipo</th>
                                    <th style="padding: 7px 10px; border-bottom: 1px solid #e2e8f0;">Detalhe</th>
                                </tr>
                            </thead>
                            <tbody>
                                {''.join(linhas_amostra)}
                            </tbody>
                        </table>
                        {aviso_mais}
                    </div>

                    <!-- Card do Arquivo Anexo -->
                    <div style="margin-top: 20px; background-color: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 8px; padding: 12px 16px;">
                        <div style="font-size: 12.5px; color: #166534; line-height: 1.4;">
                            📎 <strong>Relatório Completo em Anexo:</strong> O arquivo <code>{nome_excel}</code> está anexado a este e-mail com a relação detalhada de todas as notas modificadas para conferência.
                        </div>
                    </div>

                    <!-- Botão de Acesso ao Sistema -->
                    <div style="margin-top: 22px; text-align: center;">
                        <a href="{url_sistema}" target="_blank" style="display: inline-block; background-color: #0284c7; color: #ffffff; text-decoration: none; padding: 10px 24px; border-radius: 6px; font-weight: 600; font-size: 13px;">
                            Abrir o EDP-Helios ➜
                        </a>
                        <div style="margin-top: 6px; font-size: 11px; color: #94a3b8;">
                            Link: <a href="{url_sistema}" style="color: #0284c7;">{url_sistema}</a>
                        </div>
                    </div>
                </div>

                <!-- Footer -->
                <div style="background-color: #f8fafc; border-top: 1px solid #e2e8f0; padding: 14px 26px; font-size: 11px; color: #64748b; line-height: 1.4;">
                    <p style="margin: 0;">
                        Informativo automático gerado pelo <strong>EDP-Helios</strong> por solicitação de <strong>{usuario}</strong>.
                    </p>
                    <p style="margin: 2px 0 0 0; color: #94a3b8;">
                        EDP Brasil • Engenharia DSPM / DDPM
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

        if os.path.exists(caminho_excel):
            message.Attachments.Add(caminho_excel)

        message.Display()
        logger.info(f"✅ E-mail diário com anexo Excel gerado para {engenheiro} por {usuario}.")
        return {
            "ok": True,
            "mensagem": f"E-mail e planilha Excel gerados com sucesso no Outlook para {engenheiro} ({email_to})!",
        }
    except Exception as e:
        logger.error(f"Erro ao interagir com Outlook para {engenheiro}: {e}")
        return {
            "ok": False,
            "mensagem": f"Não foi possível abrir o Outlook para {engenheiro}: {e}",
        }
    finally:
        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except Exception:
            pass


def gerar_todos_emails_outlook(data_referencia: str | None = None, usuario: str = "sistema") -> dict:
    """Gera e-mails no Outlook para todos os engenheiros que tiveram alterações no dia."""
    resumo = obter_resumo_alteracoes_diarias(data_referencia)
    engenheiros_com_alt = [
        eng for eng, d in resumo["engenheiros"].items()
        if d["total_alteracoes"] > 0
    ]

    if not engenheiros_com_alt:
        return {
            "ok": False,
            "mensagem": "Nenhuma alteração encontrada nas regionais para gerar notificações hoje.",
            "enviados": 0,
        }

    resultados = []
    for eng in engenheiros_com_alt:
        res = gerar_email_outlook_engenheiro(eng, data_referencia=data_referencia, usuario=usuario)
        resultados.append({"engenheiro": eng, "resultado": res})

    qtd_ok = sum(1 for r in resultados if r["resultado"]["ok"])
    return {
        "ok": qtd_ok > 0,
        "mensagem": f"{qtd_ok} e-mail(s) e planilha(s) gerados com sucesso no Outlook!",
        "enviados": qtd_ok,
        "total_elegiveis": len(engenheiros_com_alt),
        "detalhes": resultados,
    }
